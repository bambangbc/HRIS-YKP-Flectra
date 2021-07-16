# from flectra import api, fields, models
##import time
# from datetime import datetime, timedelta
# from dateutil.relativedelta import relativedelta
# import dateutil.parser
# from flectra.exceptions import ValidationError, RedirectWarning, UserError
import math
import os
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

from flectra import _
from flectra import api, fields, models
from flectra.exceptions import UserError

SESSION_STATES = [('draft', 'Draft'), ('confirm', 'Waiting Approval'), ("validate", "Approved"), ("refuse", "Refused"),
                  ("cancel", "Cancelled"), ("validate_realisasi", "Realisasi Approved")]
HARI = [('senin', 'Senin'), ('selasa', 'Selasa'), ('rabu', 'rabu'), ('kamis', 'kamis'), ('jumat', 'jumat'),
        ('sabtu', 'sabtu'), ('Minggu', 'Minggu')]


class HRovertimeReportWizard(models.TransientModel):
    _name = "hr.overtime.report.wizard"

    periode_awal = fields.Date("Periode Awal")
    periode_akhir = fields.Date("Periode Akhir")

    def create_sheet_lembur(self, sheet, tipe):
        d1 = "{} 00:00:00".format(self.periode_awal)
        d2 = "{} 23:59:59".format(self.periode_akhir)
        overtimes = self.env['hr.overtime'].search(
            [('date_from', '>=', d1), ('date_from', '<=', d2), ('tipe_lembur', '=', tipe)],
            order='employee_id asc')
        sheet['A3'] = 'Periode tanggal {} s.d {}'.format(
            datetime.strptime(self.periode_awal, '%Y-%m-%d').strftime('%d %B %Y'),
            datetime.strptime(self.periode_akhir, '%Y-%m-%d').strftime('%d %B %Y'))
        no = 1
        start_row = 7
        employee_name = ''
        total_lembur = 0
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for overtime in overtimes:
            if overtime.realisasi_date_to and overtime.realisasi_date_from:
                sheet.insert_rows(start_row)
                if employee_name == '':
                    employee_name = overtime.employee_id.name
                    sheet['A{}'.format(start_row)] = no
                    sheet['A{}'.format(start_row)].border = thin_border
                    sheet['B{}'.format(start_row)] = employee_name
                    sheet['B{}'.format(start_row)].border = thin_border
                elif employee_name != overtime.employee_id.name:
                    no += 1
                    employee_name = overtime.employee_id.name
                    sheet['A{}'.format(start_row)] = no
                    sheet['A{}'.format(start_row)].border = thin_border
                    sheet['B{}'.format(start_row)] = employee_name
                    sheet['B{}'.format(start_row)].border = thin_border
                    sheet['K{}'.format(start_row - 1)] = total_lembur
                    sheet['K{}'.format(start_row - 1)].number_format = '#,##0.00'
                    sheet['K{}'.format(start_row - 1)].border = thin_border
                    total_lembur = 0
                sheet['C{}'.format(start_row)] = datetime.strftime(fields.Datetime.from_string(overtime.date_to),
                                                                   '%d-%b-%Y')
                sheet['C{}'.format(start_row)].border = thin_border
                sheet['D{}'.format(start_row)] = datetime.strftime(fields.Datetime.from_string(overtime.date_to),
                                                                   '%A')
                sheet['D{}'.format(start_row)].border = thin_border
                date_from = fields.Datetime.from_string(overtime.realisasi_date_from)
                if date_from:
                    date_from = date_from + timedelta(hours=7)
                    sheet['E{}'.format(start_row)] = datetime.strftime(date_from, '%H:%M')
                else:
                    sheet['E{}'.format(start_row)] = '-'
                sheet['E{}'.format(start_row)].border = thin_border
                date_to = fields.Datetime.from_string(overtime.realisasi_date_to)
                date_to = date_to + timedelta(hours=7)
                if date_to:
                    sheet['F{}'.format(start_row)] = datetime.strftime(date_to, '%H:%M')
                else:
                    sheet['F{}'.format(start_row)] = '-'
                sheet['F{}'.format(start_row)].border = thin_border
                sheet['G{}'.format(start_row)] = overtime.jam_lembur
                sheet['G{}'.format(start_row)].border = thin_border
                sheet['H{}'.format(start_row)] = overtime.jam_realisasi_lembur
                sheet['H{}'.format(start_row)].border = thin_border
                sheet['I{}'.format(start_row)] = 'Ada' if overtime.employee_id.npwp else 'Tidak'
                sheet['I{}'.format(start_row)].border = thin_border
                sheet['J{}'.format(start_row)] = overtime.uang_lembur
                sheet['J{}'.format(start_row)].number_format = '#,##0.00'
                sheet['J{}'.format(start_row)].border = thin_border
                sheet['L{}'.format(start_row)].border = thin_border
                total_lembur += overtime.uang_lembur
                start_row += 1
        if start_row > 7:
            sheet['K{}'.format(start_row - 1)] = total_lembur
            sheet['K{}'.format(start_row - 1)].number_format = '#,##0.00'
            sheet['K{}'.format(start_row - 1)].border = thin_border
            # total
            sheet.insert_rows(start_row)
            sheet.merge_cells('A{}:J{}'.format(start_row, start_row))
            sheet['A{}'.format(start_row)] = 'Total'
            sheet['A{}'.format(start_row)].font = Font(bold=True)
            sheet['A{}'.format(start_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet['A{}'.format(start_row)].border = thin_border
            sheet['B{}'.format(start_row)].border = thin_border
            sheet['C{}'.format(start_row)].border = thin_border
            sheet['D{}'.format(start_row)].border = thin_border
            sheet['E{}'.format(start_row)].border = thin_border
            sheet['F{}'.format(start_row)].border = thin_border
            sheet['G{}'.format(start_row)].border = thin_border
            sheet['H{}'.format(start_row)].border = thin_border
            sheet['I{}'.format(start_row)].border = thin_border
            sheet['J{}'.format(start_row)].border = thin_border

            sheet['K{}'.format(start_row)] = "=SUM(K{}:K{})".format(7, start_row - 1)
            sheet['K{}'.format(start_row)].number_format = '#,##0.00'
            sheet['K{}'.format(start_row)].border = thin_border
            sheet['L{}'.format(start_row)].border = thin_border

            sheet['H{}'.format(start_row + 2)] = 'Bandung, {}'.format(datetime.today().strftime('%d %B %Y'))

        sheet.protection.password = "YakinBisa123!"
        sheet.protection.sheet = True
        sheet.protection.enable()

    def generate_excel(self):
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'id_ID')
        except:
            print('error setting locale')

        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/report_lembur.xlsx".format(path)
        wb = load_workbook(filename)
        self.create_sheet_lembur(wb['Lembur'], 'lembur')
        self.create_sheet_lembur(wb['Lembur Libur'], 'lembur_libur')
        self.create_sheet_lembur(wb['Lembur Khusus'], 'lembur_khusus')
        wb.save(filename="{}/../reports/report_lembur_{}_{}.xlsx".format(path, str(self.periode_awal).replace('-', '_'),
                                                                         str(self.periode_awal).replace('-', '_')))
        return {
            'name': 'Lembur',
            'type': 'ir.actions.act_url',
            'url': '/hr_ykp_lembur/download/{}{}'.format(str(self.periode_awal).replace('-', '_'),
                                                         str(self.periode_awal).replace('-', '_')),
            'target': 'self',
        }


class hr_overtime(models.Model):
    _name = "hr.overtime"
    _inherit = ['mail.thread']
    _description = "Overtime"
    _order = "date_from asc"
    _rec_name = 'tujuan'

    def _default_employee(self):
        return self.env.context.get('default_employee_id') or self.env['hr.employee'].search(
            [('user_id', '=', self.env.uid)], limit=1)

    def _default_date_from(self):
        return fields.datetime.now().replace(hour=10, minute=30, second=0)

    @api.onchange('tipe_lembur')
    def onchange_tipelembur(self):
        if self.tipe_lembur == 'lembur' or self.tipe_lembur == 'lembur_khusus':
            self.date_from = fields.datetime.now().replace(hour=10, minute=30, second=0)

    name = fields.Text("Description", size=64)
    state = fields.Selection(string="State",
                             selection=SESSION_STATES,
                             required=True,
                             readonly=True,
                             default=SESSION_STATES[0][0], track_visibility='onchange')
    tipe_lembur = fields.Selection([
        ('lembur', 'Lembur Biasa'),
        ('lembur_libur', 'Lembur Hari Libur'),
        ('lembur_khusus', 'Lembur Khusus')
    ], string='Tipe Lembur', default='lembur')
    employee_id = fields.Many2one(comodel_name="hr.employee", string="Nama", track_visibility='onchange',
                                  default=_default_employee)

    unit_kerja = fields.Many2one(comodel_name="hr.department", strin="Unit Kerja")
    hari = fields.Selection(string="Hari",
                            selection=HARI)
    date_from = fields.Datetime("Start Date", track_visibility='onchange', default=_default_date_from)
    date_to = fields.Datetime("Jam Rencana Selesai", track_visibility='onchange')
    realisasi_date_from = fields.Datetime("Realisasi Jam mulai", track_visibility='onchange')
    realisasi_date_to = fields.Datetime("Realisasi Jam selesai", track_visibility='onchange')
    tujuan = fields.Char("Tujuan")
    target = fields.Char("Target")
    notes = fields.Text("Keterangan", )
    number_of_hours_temp = fields.Float("Overtime Hours", readonly=True, track_visibility='onchange')
    jam_realisasi_lembur = fields.Float('Jam Lembur', track_visibility='onchange', default=0.0)
    jam_lembur = fields.Float('Jumlah Lembur', track_visibility='onchange', default=0.0)
    approve1 = fields.Many2one(comodel_name="res.users", string="Approve 1")
    approve2 = fields.Many2one(comodel_name="res.users", string="Approve 2")
    user_id = fields.Many2one(related='employee_id.user_id', string="User")
    uang_lembur = fields.Float('Uang Lembur', default=0)

    @api.model
    def _needaction_domain_get(self, domain=None):
        return [('state', '=', 'draft')]

    @api.onchange('employee_id')
    def onchange_department(self):
        self.unit_kerja = self.employee_id.department_id.id
        self.user_id = self.employee_id.user_id.id

    @api.model
    def create(self, vals):
        DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
        from_dt = datetime.strptime(vals['date_from'], DATETIME_FORMAT)
        to_dt = datetime.strptime(vals['date_to'], DATETIME_FORMAT)
        if to_dt < from_dt:
            raise UserError(_('Waktu rencana Selesai lebih kecil dibanding waktu mulai.'))

        timedelta = to_dt - from_dt
        diff_day = (float(timedelta.seconds) / 3600)
        if vals.get('tipe_lember') == 'lembur_libur' and diff_day > 5:
            raise UserError(_('Jumlah Jam lembur untuk lembur Hari libur maksimal 5 jam'))
        vals['number_of_hours_temp'] = diff_day
        return super(hr_overtime, self).create(vals)

    @api.multi
    def unlink(self):
        for i in self:
            if i.state != 'draft' and self.env.uid != 1:
                raise UserError(_('Data yang bisa dihapus hanya yg berstatus draft'))
                # if self.env.uid == i.user_id.id and self.env.uid != 1:
                #     raise UserError(_('Anda Tidak Bisa Menghapus Lembur Atas Nama Diri Anda Sendiri.'))
        return super(hr_overtime, self).unlink()

    def _hitung_lembur(self, vals):
        if vals.get('realisasi_date_from') and vals.get('realisasi_date_to'):
            from_dt = vals['realisasi_date_from']
            to_dt = vals['realisasi_date_to']
            timedelta = to_dt - from_dt
            diff_hour = math.floor((float(timedelta.seconds) / 3600))
            vals['jam_lembur'] = diff_hour
            # hitung uang lembur
            if self.tipe_lembur == 'lembur' or self.tipe_lembur == 'lembur_khusus':
                if diff_hour > 5:
                    diff_hour = 3
                if diff_hour == 1:
                    koefsien = 1.5
                elif diff_hour == 2:
                    koefsien = 3.5
                else:
                    koefsien = 5.5
            else:
                if diff_hour > 5:
                    diff_hour = 5
                koefsien = diff_hour * 2

            if self.tipe_lembur == 'lembur_khusus' or self.employee_id.user_id.has_group(
                    'hr_ykp_employees.group_hr_kasi') or self.employee_id.user_id.has_group(
                'hr_ykp_training.group_pengurus') or self.employee_id.user_id.has_group(
                'hr_ykp_employees.group_hr_pengurus'):
                vals['uang_lembur'] = 200000
            else:
                # hitung karyawan kontrak
                imbalan = 0
                contract = self.env['hr.contract'].search(
                    [('employee_id', '=', self.employee_id.id), ('type_id', 'in', [2, 3])],
                    limit=1)  # 2 -> kontrak 3 -> partnership
                if contract:
                    if contract.type_id.name == 'PARTNERSHIP':
                        vals['uang_lembur'] = 200000 * diff_hour
                    elif contract.type_id.name == 'KONTRAK':
                        imbalan = contract.wage - 800000
                else:
                    salary = self.env['hr.master.salary'].search(
                        [('grade_id', '=', self.employee_id.grade.id), ('name', '=', 'C')], limit=1)
                    if salary:
                        imbalan = salary.nilai
                uang_lembur = (imbalan / 173) * koefsien
                if uang_lembur > 200000:
                    uang_lembur = 200000
                vals['uang_lembur'] = uang_lembur
            vals['jam_realisasi_lembur'] = diff_hour

    @api.multi
    def write(self, vals):
        """Override default Odoo write function and extend."""
        # Do your custom logic here
        if self.env.uid == self.user_id.id and self.env.uid != 1:
            raise UserError(_('Anda Tidak Bisa Menghapus Lembur Atas Nama Diri Anda Sendiri.'))
        self._hitung_lembur(vals)
        return super(hr_overtime, self).write(vals)

    @api.multi
    def action_draft(self):
        if self.approve2.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[0][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_confirm(self):
        if self.approve1.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[1][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_validate(self):
        if self.approve2.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[2][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_validate_realisasi(self):
        self.state = SESSION_STATES[5][0]
        self.action_calculate_overtime()

    @api.multi
    def action_refuse(self):
        if self.approve1.id == self.env.uid or self.approve2.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[3][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_cancel(self):
        if self.approve1.id == self.env.uid or self.approve2.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[4][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    def action_calculate_overtime(self):
        DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
        employee = self.employee_id

        d1 = datetime.strptime(self.date_from, DATETIME_FORMAT).replace(hour=0, minute=0, second=0)
        d2 = datetime.strptime(self.date_from, DATETIME_FORMAT).replace(hour=23, minute=59, second=59)
        domain = [('check_in', '>=', datetime.strftime(d1, DATETIME_FORMAT)),
                  ('check_in', '<=', datetime.strftime(d2, DATETIME_FORMAT)), ('employee_id', '=', employee.id)]
        attendance = self.env['hr.attendance'].search(domain, limit=1)
        if attendance:
            if self.tipe_lembur == 'lembur' or self.tipe_lembur == 'lembur_khusus':
                from_dt = datetime.strptime(self.date_from, DATETIME_FORMAT).replace(hour=10, minute=30)
            else:
                from_dt = datetime.strptime(attendance.check_in, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.date_to, DATETIME_FORMAT)
            to_dt_attendance = datetime.strptime(attendance.check_out, DATETIME_FORMAT)
            print(to_dt)
            print(to_dt_attendance)
            if to_dt_attendance < to_dt:
                to_dt = to_dt_attendance
            self.write({'realisasi_date_from': from_dt, 'realisasi_date_to': to_dt})

    @api.one
    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        raise UserError(_('Data tidak bisa di duplikasi !'))
        return super(hr_overtime, self).copy(default=default)

    @api.onchange('date_to', 'break_hour')
    def _get_number_of_hours(self):
        """Returns an overtime hours."""
        if self.date_to and self.date_from:
            DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
            from_dt = datetime.strptime(self.date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = (float(timedelta.seconds) / 3600)
            self.number_of_hours_temp = diff_day

    @api.onchange('date_from')
    def _calc_date(self):
        if self.date_from and self.date_to:
            DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
            from_dt = datetime.strptime(self.date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = (float(timedelta.seconds) / 3600)
            self.number_of_hours_temp = diff_day

    @api.onchange('realisasi_date_from')
    def _calc_date(self):
        if self.date_from and self.date_to:
            DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
            from_dt = datetime.strptime(self.realisasi_date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.realisasi_date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = (float(timedelta.seconds) / 3600)
            self.jam_realisasi_lembur = diff_day

    @api.onchange('realisasi_date_to')
    def _get_number_of_hours(self):
        """Returns an overtime hours."""
        if self.date_to and self.date_from:
            DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
            from_dt = datetime.strptime(self.realisasi_date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.realisasi_date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = (float(timedelta.seconds) / 3600)
            self.jam_realisasi_lembur = diff_day

    @api.multi
    def unlink(self):
        for i in self:
            if i.state != 'draft':
                raise UserError(_('Data yang bisa dihapus hanya yg berstatus draft'))
        return super(hr_overtime, self).unlink()


hr_overtime()
