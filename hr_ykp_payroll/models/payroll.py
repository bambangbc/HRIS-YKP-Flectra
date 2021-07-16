import calendar
import datetime
import logging
import os,stat
import pdb
import codecs

import time
import math
from datetime import datetime, timedelta
from datetime import time as datetime_time
from dateutil import relativedelta

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

import babel

from flectra import api, fields, models, tools, _
from flectra.tools import float_utils
from flectra.addons import decimal_precision as dp
from flectra.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


class JenisPegawai(models.Model):
    _inherit = 'jenis.pegawai'

    jkk = fields.Float('Tunjangan JKK')
    jkm = fields.Float('Tunjangan JKM')
    jht_tk = fields.Float('Tunjangan JHT TK')
    jp = fields.Float('Tunjangan JP')
    pjht_tk = fields.Float('Potongan JHT TK')
    pjp = fields.Float('Potongan JP')
    tunj_bpjskes = fields.Float("Tunjangan BPJS Kes")
    pot_bpjskes = fields.Float("Potongan BPJS Kes")
    umk = fields.Float("UMK")
    batas_max_bpjskes = fields.Float('Max BPJS Kes')
    batas_max_bpjstk = fields.Float('Max BPJS TK')


class hr_grade(models.Model):
    _inherit = 'hr.grade'

    level_id = fields.Selection([('1', '1'), ('2', '2'), ('3', '3'), ('4', '4'), ('5', '5')], string="Level",
                                required=True, default="1")

    @api.multi
    @api.depends('name', 'level_id')
    def name_get(self):
        result = []
        for i in self:
            name = i.name
            if i.level_id:
                name = name + '/' + i.level_id
            result.append((i.id, name))
        return result


class HRmanual(models.Model):
    _name = "hr.payslip.manual"
    _description = "input manual kopen, ziebar dan BJB Syariah"

    name = fields.Date(string="Tanggal Gajian", required="True")
    employee_id = fields.Many2one(comodel_name="hr.employee", string="employee", required=True,
                                  track_visibikity="onchange")
    iuran_wajib_kopen = fields.Integer(string="Iuran Wajib Kopen")
    angsuran_kopen = fields.Integer(string="Angsuran Kopen")
    kopen = fields.Integer(string="Kopen")
    iuran_wajib_ziebar = fields.Integer(string="Iurang Wajib Ziebar")
    angsuran_ziebar = fields.Integer(string="Angsuran Ziebar")
    ziebar = fields.Integer(string="Ziebar")
    bjb_syariah = fields.Integer(string="bjb Syariah")


class hr_master_salary(models.Model):
    _name = "hr.master.salary"
    _description = "Master Salary"

    name = fields.Selection(
        [('A', 'Perhitungan A'), ('B', 'Perhitungan B'), ('C', 'Perhitungan C'), ('D', 'Perhitungan D')],
        string="Jenis Perhitungan")
    grade_id = fields.Many2one('hr.grade', string='Grade')
    jenis_pegawai = fields.Many2one('jenis.pegawai', string="Jenis Pegawai")
    nilai = fields.Integer('Nilai')


hr_master_salary()


class HRPayslip(models.Model):
    _inherit = "hr.payslip"

    single_salary = fields.Integer('Single Salary')
    tunjangan_dplk = fields.Integer('Tunjangan DPLK')
    potongan_dplk = fields.Integer('Potongan DPLK')
    tunj_bpjstk = fields.Integer('Tunjangan BPJS TK')
    pot_bpjstk = fields.Integer('Potongan BPJS TK')
    tunj_bpjskes = fields.Integer('Tunjangan BPJS Kes')
    pot_bpjskes = fields.Integer('Potongan BPJS Kes')
    pot_kehadiran = fields.Integer('Potongan Kehadiran')
    pot_absensi = fields.Integer('Potongan absensi')
    pot_bjbs = fields.Integer('Potongan bjb')
    pot_kopen = fields.Integer('Potongan kopen')
    pot_ziebar = fields.Integer('Potongan ziebar')
    jum_tunjangan = fields.Integer('jum tunjangan')
    jum_potongan = fields.Integer('jum_potongan')
    gaji_bersih = fields.Integer("gaji Bersih")
    masa = fields.Char("masa kerja")
    tunj_pajak = fields.Integer("Tunjangan Pajak")
    tgl_bayar = fields.Date("Tanggal Bayar")

    @api.multi
    def compute_sheet(self):
        for payslip in self:
            tot_tunj = 0
            tot_pot = 0
            pk = 0
            pa = 0
            pbjbs = 0
            pkop = 0
            pzieb = 0
            pot_bpjskes = 0
            pot_bpjstk = 0
            tunj_bpjskes = 0
            tunj_bpjstk = 0
            pot = 0
            net = 0
            tunj = 0
            pot = 0
            basic = 0
            salarys = 0
            number = payslip.number or self.env['ir.sequence'].next_by_code('salary.slip')
            # delete old payslip lines
            payslip.line_ids.unlink()

            # mencari single salary
            if payslip.employee_id.jenis_pegawai.name == "KONTRAK":
                single = self.env['hr.master.salary'].search(
                    [('name', '=', 'A')], limit=1)
            elif payslip.employee_id.jenis_pegawai.name == "KONTRAK PARTNERSHIP":
                single = self.env['hr.master.salary'].search(
                    [('name', '=', 'A')], limit=1)
            else:
                #import pdb;pdb.set_trace()
                single = self.env['hr.master.salary'].search(
                    [('name', '=', 'A'), ('grade_id', '=', payslip.employee_id.grade.id),
                     ('jenis_pegawai', '=', payslip.employee_id.jenis_pegawai.id)])
                # import pdb;pdb.set_trace()
                if not single:
                    single = self.env['hr.master.salary'].search(
                        [('name', '=', 'B'), ('grade_id', '=', payslip.employee_id.grade.id),
                         ('jenis_pegawai', '=', payslip.employee_id.jenis_pegawai.id)])
            for salary in single:
                # tunjangan BPJS TK dan potongan BPJS TK
                if payslip.contract_id.type_id.name == "CAPEG LAMA" or payslip.contract_id.type_id.name == "KONTRAK" or payslip.contract_id.type_id.name == "PEGAWAI TETAP NON GRADING" or payslip.contract_id.type_id.name == "PARTNERSHIP" :
                    umk = payslip.employee_id.jenis_pegawai.umk
                    if payslip.contract_id.wage <= umk:
                        tunj_bpjstk = ((umk * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                (umk * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (umk * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (umk * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((umk * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                (umk * payslip.employee_id.jenis_pegawai.pjp) / 100)
                    elif payslip.contract_id.wage >= umk and payslip.contract_id.wage <= payslip.employee_id.jenis_pegawai.batas_max_bpjstk:
                        tunj_bpjstk = ((payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                (payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (
                                                      payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                (payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.pjp) / 100)
                    else:
                        tunj_bpjstk = ((
                                               payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                              (
                                                      payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (
                                                      payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (
                                                      payslip.employee_id.jenis_pegawai.batas_max_bpjstk * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((
                                              payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                             (
                                                     payslip.employee_id.jenis_pegawai.batas_max_bpjstk * payslip.employee_id.jenis_pegawai.pjp) / 100)

                else:
                    umk = payslip.employee_id.jenis_pegawai.umk
                    if salary.nilai <= umk:
                        tunj_bpjstk = ((umk * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                (umk * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (umk * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (umk * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((umk * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                (umk * payslip.employee_id.jenis_pegawai.pjp) / 100)
                    elif salary.nilai >= umk and salary.nilai <= payslip.employee_id.jenis_pegawai.batas_max_bpjstk:
                        tunj_bpjstk = ((salary.nilai * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                (salary.nilai * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (salary.nilai * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (salary.nilai * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((salary.nilai * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                (salary.nilai * payslip.employee_id.jenis_pegawai.pjp) / 100)
                    else:
                        tunj_bpjstk = ((salary.nilai * payslip.employee_id.jenis_pegawai.jkk) / 100) + (
                                (salary.nilai * payslip.employee_id.jenis_pegawai.jkm) / 100) + (
                                              (salary.nilai * payslip.employee_id.jenis_pegawai.jht_tk) / 100) + (
                                              (payslip.employee_id.jenis_pegawai.batas_max_bpjstk * payslip.employee_id.jenis_pegawai.jp) / 100)
                        pot_bpjstk = ((salary.nilai * payslip.employee_id.jenis_pegawai.pjht_tk) / 100) + (
                                (payslip.employee_id.jenis_pegawai.batas_max_bpjstk * payslip.employee_id.jenis_pegawai.pjp) / 100)
                if payslip.contract_id.type_id.name != "KONTRAK" and payslip.contract_id.type_id.name != "CAPEG LAMA" and payslip.contract_id.type_id.name != "PEGAWAI TETAP NON GRADING" and payslip.contract_id.type_id.name != "PARTNERSHIP":
                    umk = payslip.employee_id.jenis_pegawai.umk
                    if salary.nilai <= umk:
                        tunj_bpjskes = ((umk * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                    elif salary.nilai >= umk and salary.nilai <= payslip.employee_id.jenis_pegawai.batas_max_bpjskes:
                        tunj_bpjskes = ((salary.nilai * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                    else:
                        tunj_bpjskes = ((
                                                payslip.employee_id.jenis_pegawai.batas_max_bpjskes * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)

                    if salary.nilai <= umk:
                        pot_bpjskes = ((umk * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                    elif salary.nilai >= umk and salary.nilai <= payslip.employee_id.jenis_pegawai.batas_max_bpjskes:
                        pot_bpjskes = ((salary.nilai * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                    else:
                        pot_bpjskes = ((
                                               payslip.employee_id.jenis_pegawai.batas_max_bpjskes * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                else:
                    umk = payslip.employee_id.jenis_pegawai.umk
                    if payslip.contract_id.wage <= umk:
                        tunj_bpjskes = ((umk * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                    elif payslip.contract_id.wage >= umk and payslip.contract_id.wage <= payslip.employee_id.jenis_pegawai.batas_max_bpjskes:
                        tunj_bpjskes = (
                                (payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                    else:
                        tunj_bpjskes = ((
                                                payslip.employee_id.jenis_pegawai.batas_max_bpjskes * payslip.employee_id.jenis_pegawai.tunj_bpjskes) / 100)

                    if payslip.contract_id.wage <= umk:
                        pot_bpjskes = ((umk * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                    elif payslip.contract_id.wage >= umk and payslip.contract_id.wage <= payslip.employee_id.jenis_pegawai.batas_max_bpjskes:
                        pot_bpjskes = ((payslip.contract_id.wage * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                    else:
                        pot_bpjskes = ((
                                               payslip.employee_id.jenis_pegawai.batas_max_bpjskes * payslip.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                if payslip.contract_id.wage == 0:
                    salarys = salary.nilai
                else:
                    salarys = payslip.contract_id.wage
                payslip.write({'tunj_bpjstk': math.ceil(tunj_bpjstk), 'pot_bpjstk': math.ceil(pot_bpjstk),
                               'tunj_bpjskes': math.ceil(tunj_bpjskes), 'pot_bpjskes': math.ceil(pot_bpjskes)})
            #import pdb;pdb.set_trace()
            if salarys == 0:
                salarys = payslip.contract_id.wage
            payslip.write({'single_salary': salarys})
            # mencari tunjangan DPLK dan potongan dplk
            if payslip.contract_id.type_id.name == "PEGAWAI TETAP NON GRADING" and payslip.employee_id.name != "Ivansyah Wahyu" and payslip.employee_id.name != "Dwi Permana":
                tunj = ((payslip.contract_id.wage - 1300000) * 10) / 100
                pot = ((payslip.contract_id.wage - 1300000) * 5) / 100
                # payslip.write({'tunjangan_dplk' : tunj, 'potongan_dplk' : pot})
            elif payslip.contract_id.type_id.name == "PEGAWAI TETAP NON GRADING" and payslip.employee_id.name == "Ivansyah Wahyu":
                tunj = ((payslip.contract_id.wage - 2050000) * 10) / 100
                pot = ((payslip.contract_id.wage - 2050000) * 5) / 100
                # payslip.write({'tunjangan_dplk' : tunj, 'potongan_dplk' : pot})
            elif payslip.contract_id.type_id.name == "PEGAWAI TETAP" or payslip.contract_id.type_id.name == "PEGAWAI TETAP NON GRADING":
                single = self.env['hr.master.salary'].search(
                    [('name', '=', 'C'), ('grade_id', '=', payslip.employee_id.grade.id)])
                for salary in single:
                    tunj = (salary.nilai * 10) / 100
                    pot = (salary.nilai * 5) / 100
            payslip.write({'tunjangan_dplk': tunj, 'potongan_dplk': pot})

            # set the list of contract for which the rules have to be applied
            # if we don't give the contract, then the rules to apply should be for all current contracts of the employee
            payslip.write(
                {'pot_kehadiran': pk, 'pot_absensi': pa, 'pot_bjbs': pbjbs, 'pot_kopen': pkop, 'pot_ziebar': pzieb})
            contract_ids = payslip.contract_id.ids or \
                           self.get_contract(payslip.employee_id, payslip.date_from, payslip.date_to)
            lines = [(0, 0, line) for line in self._get_payslip_lines(contract_ids, payslip.id)]
            payslip.write({'line_ids': lines, 'number': number})
            # for lin in self :
            pk = 0
            pa = 0
            pbjbs = 0
            pkop = 0
            pzieb = 0
            for lines in payslip.line_ids:
                if lines['code'] == 'PK':
                    pk = lines['amount']
                if lines['code'] == 'PA':
                    pa = lines['amount']
                if lines['code'] == 'PBJBS':
                    pbjbs = lines['amount']
                if lines['code'] == 'PKOP':
                    pkop = lines['amount']
                if lines['code'] == 'PZIEB':
                    pzieb = lines['amount']
                if lines['code'] == 'NET':
                    net = lines['amount']
                if lines['code'] == 'BASIC':
                    basic = lines['amount']
            if payslip.employee_id.tanggal_masuk != False:
                masa = ((datetime.now() - datetime.strptime(payslip.employee_id.tanggal_masuk,
                                                            "%Y-%M-%d")).days / 30) / 12
                tahun = int(masa)
                bulan = str(masa - tahun)[2:3]
            else:
                tahun = 0
                bulan = 0
            payslip.write({'pot_kopen': -pkop, 'pot_ziebar': -pzieb,
                           'pot_kehadiran': -pk, 'pot_absensi': -pa, 'pot_bjbs': -pbjbs,
                           'jum_tunjangan': tunj + tunj_bpjskes + tunj_bpjstk,
                           'jum_potongan': -(int(pk) + int(pa) - int(pot) - int(pot_bpjskes) - int(pot_bpjstk) + int(
                               pbjbs) + int(pkop) + int(pzieb)), 'gaji_bersih': net,
                           'masa': str(tahun) + " Tahun " + str(bulan) + " Bulan"})
            ### Menghitung pajak pph21 ###
            gaji = 0
            asuransi = 0
            potongan = 0
            salary = 0
            dplk = 0
            bpjstk = 0
            bpjskes = 0
            potongan_dplk = 0
            pot_bpjstk = 0
            pot_dplk_i = 0
            pot_bpjstk_i = 0
            pot_bpjskes = 0
            bpjstk_jkm = 0
            bpjstk_jkk = 0
            pot_bpjskes_i = 0
            bpjstk_jkm_i = 0
            bpjstk_jkk_i = 0
            pot_kehadiran = 0
            pot_absensi = 0
            pot_dplk = 0
            mth = 0
            mth_dari = 12   
            month = datetime.strptime(payslip.date_to,"%Y-%m-%d").month
            year = datetime.strptime(payslip.date_to,"%Y-%m-%d").year
            date_to = datetime(int(year), month, 26).strftime('%Y-%m-%d')
            date_one = datetime(int(year), 1, 1).strftime('%Y-%m-%d')
            pysl = self.env['hr.payslip'].search(
                [('employee_id', '=', payslip.employee_id.id), ('date_to', '<=', date_to),('date_to','>=', date_one)])
            for pyslip in pysl :
                
                salary = payslip.single_salary
                pot_kehadiran = 0
                pot_absensi = 0
                for detail in payslip.line_ids :
                    if detail.code == "PK" :
                        pot_kehadiran =detail.total
                    if detail.code == "PA" :
                        pot_absensi = detail.total

                #faktor Penambah
                gaji += salary + pot_kehadiran + pot_absensi
                bpjstk_jkk_i = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkk/100)
                bpjstk_jkm_i = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkm/100)
                bpjskes_i = payslip.tunj_bpjskes

                bpjstk_jkk += payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkk/100)
                bpjstk_jkm += payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkm/100)
                bpjskes += payslip.tunj_bpjskes

                #faktor pengurang
                pot_dplk_i = payslip.potongan_dplk
                pot_bpjstk_i = payslip.pot_bpjstk

                pot_dplk += payslip.potongan_dplk
                pot_bpjstk += payslip.pot_bpjstk

            for i in range(month, 12):

                #faktor penambah
                gaji += salary
                bpjstk_jkk += bpjstk_jkk_i
                bpjstk_jkm += bpjstk_jkm_i
                bpjskes += bpjskes_i
                
                #faktor pengurang
                pot_dplk += pot_dplk_i
                pot_bpjstk += pot_bpjstk_i
            biaya_jabatan = gaji * 0.05 
            ## hitung pajak
            pajak_real = 0
            pkp_setahun = 0
            cek = True
            i = 0
            pajak_sebelum = 0
            while cek == True and i < 100 :
                jum_bruto = gaji + bpjstk_jkk + bpjstk_jkm + bpjskes + pajak_sebelum
                biaya_jabatan = jum_bruto * 0.05
                jum_pengurang = biaya_jabatan + pot_dplk + pot_bpjstk
                jum_neto = jum_bruto - jum_pengurang
                pkp_sthn = jum_neto - payslip.employee_id.status_pajak.nominal_tahun
                if pkp_sthn > 1000 :
                    pkp_setahun = int(str(int(pkp_sthn))[:len(str(int(pkp_sthn)))-3]+"000")
                pajak_real = 0
                for pkp in self.env['hr.pkp'].search([]) :
                    pajak = 0
                    if pkp_setahun > pkp.nominal_max :
                        total_ptkp = pkp.nominal_max - pkp.nominal_mix
                        pajak = total_ptkp * pkp.pajak
                    elif pkp_setahun <= pkp.nominal_max and pkp_setahun >= pkp.nominal_mix :
                        pjk = pkp_setahun - pkp.nominal_mix
                        pajak = pjk * pkp.pajak
                    pajak_real = pajak_real + pajak
                pph21 = pajak_real
                #self.pph21 = (self.pkp_setahun - rumus_pkp.nominal_mix) * rumus_pkp.pajak
                if pph21 == pajak_sebelum :
                    cek_rutin = False
                if pph21 != pajak_sebelum :
                    pajak_sebelum = pph21
                i += 1
            #import pdb;pdb.set_trace()
            tgl_bayar = str(datetime.strptime((payslip.date_to),"%Y-%m-%d") + timedelta(days=1))
            ttyme = datetime.fromtimestamp(time.mktime(time.strptime(payslip.date_to, "%Y-%m-%d")))
            locale = self.env.context.get('lang') or 'en_US'
            names = _('Slip Gaji Untuk %s Bulan %s') % (payslip.employee_id.name, tools.ustr(babel.dates.format_date(date=ttyme, format='MMMM-y', locale=locale)))
            payslip.write({'tunj_pajak': pph21/12,'tgl_bayar':tgl_bayar,'name':names})

        return True

    @api.model
    def get_worked_day_lines(self, contracts, date_from, date_to):
        """
        @param contract: Browse record of contracts
        @return: returns a list of dict containing the input that should be applied for the given contract between date_from and date_to
        """
        res = []
        # fill only if the contract as a working schedule linked
        for contract in contracts.filtered(lambda contract: contract.resource_calendar_id):
            day_from = datetime.combine(fields.Date.from_string(date_from), datetime_time.min)
            day_to = datetime.combine(fields.Date.from_string(date_to), datetime_time.max)

            # compute leave days
            leaves = {}
            # day_leave_intervals = contract.employee_id.iter_leaves(day_from, day_to, calendar=contract.resource_calendar_id)
            # for day_intervals in day_leave_intervals:
            #	for interval in day_intervals:
            #		holiday = interval[2]['leaves'].holiday_id
            #		current_leave_struct = leaves.setdefault(holiday.holiday_status_id, {
            #			'name': holiday.holiday_status_id.name or _('Global Leaves'),
            #			'sequence': 5,
            #			'code': holiday.holiday_status_id.name or 'GLOBAL',
            #			'number_of_days': 0.0,
            #			'number_of_hours': 0.0,
            #			'contract_id': contract.id,
            #		})
            #		leave_time = (interval[1] - interval[0]).seconds / 3600
            #		current_leave_struct['number_of_hours'] += leave_time
            #		work_hours = contract.employee_id.get_day_work_hours_count(interval[0].date(), calendar=contract.resource_calendar_id)
            #		if work_hours:
            #			current_leave_struct['number_of_days'] += leave_time / work_hours

            # compute worked days
            work_data = contract.employee_id.with_context(no_tz_convert=True).get_work_days_data(day_from, day_to,
                                                                                                 calendar=contract.resource_calendar_id)
            attendances = {
                'name': _("Normal Working Days paid at 100%"),
                'sequence': 1,
                'code': 'WORK100',
                'number_of_days': work_data['days'],
                'number_of_hours': work_data['hours'],
                'contract_id': contract.id,
            }
            att = self.env['hr.attendance'].search(
                [('employee_id', '=', contract.employee_id.id), ('check_in', '>=', str(
                    datetime.strptime(date_from + " 00:00:01", "%Y-%m-%d %H:%M:%S") - timedelta(hours=7))), (
                     'check_in', '<=',
                     str(datetime.strptime(date_to + " 23:59:59", "%Y-%m-%d %H:%M:%S") - timedelta(hours=7)))])
            # ,('check_out', '<=', str(datetime.strptime(date_to+" 23:59:59","%Y-%m-%d %H:%M:%S")-timedelta(hours=7)))])
            presen_day = 0
            presen_hour = 0
            terlambat = 0
            perdin_hari = 0
            diklat_hari = 0
            terlambat2jam = 0
            for att_in in att:
                if (datetime.strptime(att_in.check_in, "%Y-%m-%d %H:%M:%S") + timedelta(
                        hours=7)).isoweekday() <= 5:
                    presen_day += 1
                    if att_in.check_out != False:
                        presen_hour += (datetime.strptime(att_in.check_out,
                                                          "%Y-%m-%d %H:%M:%S") - datetime.strptime(
                            att_in.check_in, "%Y-%m-%d %H:%M:%S")).seconds / 3600
                    # import pdb;pdb.set_trace()
                    if att_in.terlambat >= 0.25 and att_in.terlambat < 2:
                        terlambat += 1
                    elif att_in.terlambat >= 2:
                        terlambat2jam += 1
            akum_terlambat = {
                'name': _("Akum Terlambat"),
                'sequence': 3,
                'code': 'AT',
                'number_of_days': 0.0,
                'number_of_hours': 0.0,
                'contract_id': contract.id,
            }
            day_from = datetime.strptime(date_from + " 00:00:00", "%Y-%m-%d %H:%M:%S")
            day_to = datetime.strptime(date_to + " 23:59:59", "%Y-%m-%d %H:%M:%S")
            nb_of_days = (day_to - day_from).days + 1

            xx = 0
            week = 1
            weeks = 0
            tampung = 0.0
            for day in range(0, nb_of_days):
                week = (day_from + timedelta(days=day)).isoweekday()
                if week == 7 or week == 6:
                    xx += 1
                dates = day_from + timedelta(days=day)
                absent = self.env['hr.attendance'].search([('employee_id', '=', contract.employee_id.id),
                                                           ('check_in', '>=', str(dates + timedelta(minutes=1))), (
                                                               'check_out', '<=',
                                                               str(dates + timedelta(hours=23, minutes=59)))])
                if absent.terlambat < 0.25:
                    tampung += absent.terlambat
                if week == 7 or day == nb_of_days:
                    if tampung >= 0.249:
                        akum_terlambat['number_of_days'] += 1
                    tampung = 0.0
                    # if week < weeks :
                    # weeks = (day_from + timedelta(days=day)).isoweekday()
                # perdin
                # import pdb;pdb.set_trace()
                perdin = self.env['hr.perdin.employee'].search(
                    [('employee_id', '=', contract.employee_id.id), ('state', '=', 'validate'),
                     ('date_from', '<=', str(day_from + timedelta(days=day))),
                     ('date_to', '>=', str(day_from + timedelta(days=day)))])
                if perdin and (day_from + timedelta(days=day)).isoweekday() < 6:
                    perdin_hari += 1
                diklat = self.env['hr.training.participant'].search(
                    [('name', '=', contract.employee_id.id), ('state', '=', 'approve'),
                     ('start_date', '<=', str(day_from + timedelta(days=day))),
                     ('end_date', '>=', str(day_from + timedelta(days=day)))])
                if diklat and (day_from + timedelta(days=day)).isoweekday() < 6:
                    diklat_hari += 1
            attendances['number_of_days'] = nb_of_days - xx
            attendances['number_of_hours'] = (nb_of_days - xx) * 8

            # perdin = self.env['hr.perdin'].search(
            #    [('state', '=', 'validate'), ('date_from', '>=', date_from), ('date_to', '>=', date_from),
            #     ('date_from', '<=', date_to)])
            # for per in perdin:
            #    if per.date_from >= date_from and per.date_to <= date_to:
            #        for emp in per.employee_ids:
            #            if emp.employee_id.id == contract.employee_id.id:
            #                perdin_hari += per.number_of_days_temp
            #    elif per.date_to >= date_from and per.date_from < date_from:
            #        for emp in per.employee_ids:
            #            if emp.employee_id.id == contract.employee_id.id:
            #                perdin_hari += (datetime.strptime(per.date_to,
            #                                                           "%Y-%m-%d") - datetime.strptime(
            #                    date_from,
            #                    "%Y-%m-%d")).days + 1
            #    elif per.date_from <= date_to and per.date_to >= date_to:
            #        for emp in per.employee_ids:
            #            if emp.employee_id.id == contract.employee_id.id:
            #                perdin_hari += (
            #                               datetime.strptime(date_to, "%Y-%m-%d") - datetime.strptime(
            #                                   per.date_from,
            #                                   "%Y-%m-%d")).days + 1
            presences = {
                'name': _("Kehadiran"),
                'sequence': 2,
                'code': 'presences',
                'number_of_days': presen_day,
                'number_of_hours': presen_hour,
                'contract_id': contract.id,
            }
            terlambat = {
                'name': _("Terlambat"),
                'sequence': 3,
                'code': 'terlambat',
                'number_of_days': terlambat,
                'number_of_hours': 0.0,
                'contract_id': contract.id,
            }
            terlambat_lebih = {
                'name': _("Terlambat Lebih 2 Jam"),
                'sequence': 3,
                'code': 'terlambat2jam',
                'number_of_days': terlambat2jam,
                'number_of_hours': 0.0,
                'contract_id': contract.id,

            }
            perdin = {
                'name': _("Perdin/Diklat"),
                'sequence': 4,
                'code': 'perdin',
                'number_of_days': perdin_hari + diklat_hari,
                'number_of_hours': 0.0,
                'contract_id': contract.id,
            }
            alfa = {
                'name': _("Tidak Hadir"),
                'sequence': 4,
                'code': 'TH',
                'number_of_days': 0.0,
                'number_of_hours': 0.0,
                'contract_id': contract.id,
            }
            cuti = self.env['hr.holidays'].search(
                [('employee_id', '=', contract.employee_id.id), ('state', '=', 'validate'),
                 ('date_from', '>=', date_from), ('date_to', '>=', date_from), ('date_from', '<=', date_to)])
            leaves = {}
            tk = 0
            for cut in cuti:
                if cut.holiday_status_id.name:
                    if cut.holiday_status_id.name in leaves:
                        if cut.date_from >= date_from and cut.date_to <= date_to:
                            leaves[cut.holiday_status_id.name]['number_of_days'] += (datetime.strptime(
                                cut.date_to,
                                "%Y-%m-%d") - datetime.strptime(
                                cut.date_from, "%Y-%m-%d")).days + 1
                            if cut.holiday_status_id.name == 'Cuti Tahunan' or cut.holiday_status_id.name == 'Cuti Besar':
                                tk += (datetime.strptime(cut.date_to, "%Y-%m-%d") - datetime.strptime(
                                    cut.date_from,
                                    "%Y-%m-%d")).days + 1
                        elif cut.date_to >= date_from and cut.date_from <= date_from:
                            leaves[cut.holiday_status_id.name]['number_of_days'] += (datetime.strptime(
                                cut.date_to,
                                "%Y-%m-%d") - datetime.strptime(
                                date_from, "%Y-%m-%d")).days + 1
                            if cut.holiday_status_id.name == 'Cuti Tahunan' or cut.holiday_status_id.name == 'Cuti Besar':
                                tk += (datetime.strptime(cut.date_to, "%Y-%m-%d") - datetime.strptime(
                                    date_from,
                                    "%Y-%m-%d")).days + 1
                        elif cut.date_from <= date_to and cut.date_to >= date_to:
                            leaves[cut.holiday_status_id.name]['number_of_days'] += (datetime.strptime(date_to,
                                                                                                       "%Y-%m-%d") - datetime.strptime(
                                cut.date_from, "%Y-%m-%d")).days + 1
                            if cut.holiday_status_id.name == 'Cuti Tahunan' or cut.holiday_status_id.name == 'Cuti Besar':
                                tk += (datetime.strptime(date_to, "%Y-%m-%d") - datetime.strptime(
                                    cut.date_from,
                                    "%Y-%m-%d")).days + 1
                    else:
                        leaves[cut.holiday_status_id.name] = {
                            'name': cut.holiday_status_id.name,
                            'sequence': 5,
                            'code': cut.holiday_status_id.name,
                            'number_of_days': cut.number_of_days_temp,
                            'number_of_hours': 0.0,
                            'contract_id': contract.id,
                        }
                    if cut.holiday_status_id.name != 'SAKIT DENGAN SURAT DOKTER' and cut.holiday_status_id.name != 'SAKIT TANPA SURAT DOKTER':
                        tk += cut.number_of_days_temp
            res.append(attendances)
            # res.extend(leaves.values())
            leaves = [value for key, value in leaves.items()]
            # import pdb;pdb.set_trace()
            alfa['number_of_days'] = attendances['number_of_days'] - presences['number_of_days'] - tk - perdin[
                'number_of_days']
            if alfa['number_of_days'] <= 0:
                alfa['number_of_days'] = 0
            # for day in range(0, nb_of_days):
            res += [presences] + [terlambat] + [terlambat_lebih] + [akum_terlambat] + [perdin] + leaves + [alfa]

        return res

    @api.model
    def get_inputs(self, contracts, date_from, date_to):
        res = []

        structure_ids = contracts.get_all_structures()
        rule_ids = self.env['hr.payroll.structure'].browse(structure_ids).get_all_rules()
        sorted_rule_ids = [id for id, sequence in sorted(rule_ids, key=lambda x: x[1])]
        inputs = self.env['hr.salary.rule'].browse(sorted_rule_ids).mapped('input_ids')
        ziebar = {
            'name': _("Ziebar"),
            'sequence': 1,
            'code': 'ZIEBAR',
            'amount': 0.0,
            'contract_id': contracts.id,
        }
        kopen = {
            'name': _("Simpanan Wajib"),
            'sequence': 2,
            'code': 'KOPEN',
            'amount': 0.0,
            'contract_id': contracts.id,
        }
        bjb_syariah = {
            'name': _("BJB Syariah"),
            'sequence': 3,
            'code': 'BJBS',
            'amount': 0.0,
            'contract_id': contracts.id,
        }
        tunj_dplk = {
            'name': _("Tunj DPLK"),
            'sequence': 4,
            'code': 'TDPLK',
            'amount': 0.0,
            'contract_id': contracts.id,
        }
        import_payslip = self.env['hr.payslip.manual']
        check_import = import_payslip.search([('employee_id', '=', contracts.employee_id.id), ('name', '=', date_to)],
                                             limit=1)

        if check_import:
            for result in check_import:
                kopen['amount'] = result.kopen
                ziebar['amount'] = result.ziebar
                bjb_syariah['amount'] = result.bjb_syariah
        res += [ziebar] + [kopen] + [bjb_syariah] + [tunj_dplk]

        for contract in contracts:
            for input in inputs:
                input_data = {
                    'name': input.name,
                    'code': input.code,
                    'contract_id': contract.id,
                }
                res += [input_data]
        return res


class HRPayslipReportWizard(models.TransientModel):
    _name = "hr.payslip.report.wizard"

    MONTHS = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli', 'Agustus', 'September', 'Oktober',
              'November', 'Desember']

    month = fields.Selection([
        (1, 'Januari'),
        (2, 'Februari'),
        (3, 'Maret'),
        (4, 'April'),
        (5, 'Mei'),
        (6, 'Juni'),
        (7, 'Juli'),
        (8, 'Agustus'),
        (9, 'September'),
        (10, 'Oktober'),
        (11, 'November'),
        (12, 'Desember'),
    ], default=datetime.now().month, string='Bulan')
    year = fields.Char('Tahun', default=datetime.now().year)
    nomor = fields.Char("Nomor")

    def setupSheetDPLK(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                 self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            if p.contract_id.type_id.name == "CAPEG LAMA":
                dplk = 0
                pot_dplk = 0
            else:
                dplk = p.tunjangan_dplk
                pot_dplk = p.potongan_dplk

            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = p.single_salary
            sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['E{}'.format(10 + idx)].border = thin_border
            sheet['E{}'.format(10 + idx)] = dplk
            sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)] = pot_dplk
            sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['G{}'.format(10 + idx)].border = thin_border
            sheet['G{}'.format(10 + idx)] = "=E{} + F{}".format(10 + idx, 10 + idx)
            idx += 1
        if idx > 0:
            sheet['E{}'.format(10 + idx)] = "=SUM(E{}:E{})".format(10, 10 + idx - 1)
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['G{}'.format(10 + idx)] = "=SUM(G{}:G{})".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'
        return {
            'col': 'E',
            'row': 10 + idx,
            'col2': 'F',
            'row2': 10 + idx,
        }

    def setupSheetBPJSTK(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            if p.contract_id.type_id.name == "PARTNERSHIP":
                single = p.contract_id.wage
            elif p.contract_id.type_id.name == "PEGAWAI TETAP":
                single = p.single_salary
            elif p.contract_id.type_id.name == "CALON PEGAWAI":
                single = (p.single_salary * 80) / 100
            else:
                single = p.contract_id.wage

            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = single
            if single >= p.employee_id.jenis_pegawai.umk or single == 0:
                sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['E{}'.format(10 + idx)].border = thin_border
                sheet['E{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx, p.employee_id.jenis_pegawai.jkk)
                sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['F{}'.format(10 + idx)].border = thin_border
                sheet['F{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx, p.employee_id.jenis_pegawai.jkm)
                sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['G{}'.format(10 + idx)].border = thin_border
                sheet['G{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx, p.employee_id.jenis_pegawai.jht_tk)
                if single <= p.employee_id.jenis_pegawai.batas_max_bpjstk :
                    sheet['H{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    sheet['H{}'.format(10 + idx)].border = thin_border
                    sheet['H{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx, p.employee_id.jenis_pegawai.jp)
                    sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    sheet['K{}'.format(10 + idx)].border = thin_border
                    sheet['K{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx,
                                                                               p.employee_id.jenis_pegawai.pjp)
                else :
                    sheet['H{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    sheet['H{}'.format(10 + idx)].border = thin_border
                    sheet['H{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.batas_max_bpjstk * p.employee_id.jenis_pegawai.jp) / 100)
                    sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    sheet['K{}'.format(10 + idx)].border = thin_border
                    sheet['K{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.batas_max_bpjstk * p.employee_id.jenis_pegawai.pjp) / 100)
                sheet['I{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['I{}'.format(10 + idx)].border = thin_border
                sheet['I{}'.format(10 + idx)] = "=ROUND(SUM(E{}:H{}),0)".format(10 + idx, 10 + idx)
                #if p.employee_id.name == "Dindin Achmad S":
                #    sheet['J{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                #    sheet['J{}'.format(10 + idx)].border = thin_border
                #    sheet['J{}'.format(10 + idx)] = "='Sheet1'!I7".format(10 + idx)
                #    sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                #    sheet['K{}'.format(10 + idx)].border = thin_border
                #    sheet['K{}'.format(10 + idx)] = "='Sheet1'!J7".format(10 + idx)
                #else:
                sheet['J{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['J{}'.format(10 + idx)].border = thin_border
                sheet['J{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx,
                                                                               p.employee_id.jenis_pegawai.pjht_tk)
                sheet['L{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['L{}'.format(10 + idx)].border = thin_border
                sheet['L{}'.format(10 + idx)] = "=ROUND(SUM(J{}:K{}),0)".format(10 + idx, 10 + idx)
                sheet['M{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['M{}'.format(10 + idx)].border = thin_border
                sheet['M{}'.format(10 + idx)] = "=I{} + L{}".format(10 + idx, 10 + idx)
            else:
                sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['E{}'.format(10 + idx)].border = thin_border
                sheet['E{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.jkk) / 100)
                sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['F{}'.format(10 + idx)].border = thin_border
                sheet['F{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.jkm) / 100)
                sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['G{}'.format(10 + idx)].border = thin_border
                sheet['G{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.jht_tk) / 100)
                sheet['H{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['H{}'.format(10 + idx)].border = thin_border
                sheet['H{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.jp) / 100)
                sheet['I{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['I{}'.format(10 + idx)].border = thin_border
                sheet['I{}'.format(10 + idx)] = "=SUM(E{}:H{})".format(10 + idx, 10 + idx)
                sheet['J{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['J{}'.format(10 + idx)].border = thin_border
                sheet['J{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.pjht_tk) / 100)
                sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['K{}'.format(10 + idx)].border = thin_border
                sheet['K{}'.format(10 + idx)] = "=ROUND({},0)".format(
                    (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.pjp) / 100)
                sheet['L{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['L{}'.format(10 + idx)].border = thin_border
                sheet['L{}'.format(10 + idx)] = "=SUM(J{}:K{})".format(10 + idx, 10 + idx)
                sheet['M{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['M{}'.format(10 + idx)].border = thin_border
                sheet['M{}'.format(10 + idx)] = "=I{} + L{}".format(10 + idx, 10 + idx)
            idx += 1
        if idx > 0:
            sheet['E{}'.format(10 + idx)] = "=ROUND(SUM(E{}:E{}),0)".format(10, 10 + idx - 1)
            sheet['F{}'.format(10 + idx)] = "=ROUND(SUM(F{}:F{}),0)".format(10, 10 + idx - 1)
            sheet['G{}'.format(10 + idx)] = "=ROUND(SUM(G{}:G{}),0)".format(10, 10 + idx - 1)
            sheet['H{}'.format(10 + idx)] = "=ROUND(SUM(H{}:H{}),0)".format(10, 10 + idx - 1)
            sheet['I{}'.format(10 + idx)] = "=ROUND(SUM(I{}:I{}),0)".format(10, 10 + idx - 1)
            sheet['J{}'.format(10 + idx)] = "=ROUND(SUM(J{}:J{}),0)".format(10, 10 + idx - 1)
            sheet['K{}'.format(10 + idx)] = "=ROUND(SUM(K{}:K{}),0)".format(10, 10 + idx - 1)
            sheet['L{}'.format(10 + idx)] = "=ROUND(SUM(L{}:L{}),0)".format(10, 10 + idx - 1)
            sheet['M{}'.format(10 + idx)] = "=ROUND(SUM(M{}:M{}),0)".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'
        return {
            'col': 'I',
            'row': 10 + idx,
            'col2': 'L',
            'row2': 10 + idx
        }

    def setupSheetBpjsKes(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            if p.contract_id.type_id.name == "PEGAWAI TETAP":
                single = p.single_salary
            elif p.contract_id.type_id.name == "CALON PEGAWAI":
                single = (p.single_salary * 80) / 100
            else:
                single = p.contract_id.wage

            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = single
            sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['E{}'.format(10 + idx)].border = thin_border
            if p.employee_id.jenis_pegawai.name != "KONTRAK PARTNERSHIP NON BPJS KES":
                if single <= p.employee_id.jenis_pegawai.umk:
                    sheet['E{}'.format(10 + idx)] = "=ROUND({},0)".format(
                        (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                elif single >= p.employee_id.jenis_pegawai.batas_max_bpjskes:
                    sheet['E{}'.format(10 + idx)] = "=ROUND({},0)".format((
                                                                                  p.employee_id.jenis_pegawai.batas_max_bpjskes * p.employee_id.jenis_pegawai.tunj_bpjskes) / 100)
                else:
                    sheet['E{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx,
                                                                               p.employee_id.jenis_pegawai.tunj_bpjskes)
                sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                sheet['F{}'.format(10 + idx)].border = thin_border
                if single <= p.employee_id.jenis_pegawai.umk:
                    sheet['F{}'.format(10 + idx)] = "=ROUND({},0)".format(
                        (p.employee_id.jenis_pegawai.umk * p.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                elif single >= p.employee_id.jenis_pegawai.batas_max_bpjskes:
                    sheet['F{}'.format(10 + idx)] = "=ROUND({},0)".format(
                        (p.employee_id.jenis_pegawai.batas_max_bpjskes * p.employee_id.jenis_pegawai.pot_bpjskes) / 100)
                else:
                    sheet['F{}'.format(10 + idx)] = "=ROUND(D{}*{}%,0)".format(10 + idx,
                                                                               p.employee_id.jenis_pegawai.pot_bpjskes)
            sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['G{}'.format(10 + idx)].border = thin_border
            sheet['G{}'.format(10 + idx)] = "=E{}+F{}".format(10 + idx, 10 + idx)
            idx += 1
        if idx > 0:
            sheet['E{}'.format(10 + idx)] = "=SUM(E{}:E{})".format(10, 10 + idx - 1)
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['G{}'.format(10 + idx)] = "=SUM(G{}:G{})".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'
        return {
            'col': 'E',
            'row': 10 + idx,
            'col2': 'F',
            'row2': 10 + idx,
        }

    def setupSheetPenguranKehadiran(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            numTH = 0
            numPerdin = 0
            for d in p.worked_days_line_ids:
                if d.code == 'TH':
                    numTH = d.number_of_days
                if d.code == 'perdin':
                    numPerdin = d.number_of_days

            sheet.insert_rows(11 + idx)
            sheet['A{}'.format(11 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(11 + idx)].border = thin_border
            sheet['B{}'.format(11 + idx)] = p.employee_id.name
            sheet['B{}'.format(11 + idx)].border = thin_border
            sheet['C{}'.format(11 + idx)] = numTH
            sheet['C{}'.format(11 + idx)].border = thin_border
            sheet['D{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(11 + idx)].border = thin_border
            sheet['D{}'.format(11 + idx)] = "=C{}*65000".format(11 + idx)
            sheet['E{}'.format(11 + idx)] = numPerdin
            sheet['E{}'.format(11 + idx)].border = thin_border
            sheet['F{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(11 + idx)].border = thin_border
            sheet['F{}'.format(11 + idx)] = "=E{}*40000".format(11 + idx)
            sheet['G{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['G{}'.format(11 + idx)].border = thin_border
            sheet['G{}'.format(11 + idx)] = "=D{}+F{}".format(11 + idx, 11 + idx)
            idx += 1
        if idx > 0:
            sheet['C{}'.format(11 + idx)] = "=SUM(C{}:C{})".format(11, 11 + idx - 1)
            sheet['D{}'.format(11 + idx)] = "=SUM(D{}:D{})".format(11, 11 + idx - 1)
            sheet['E{}'.format(11 + idx)] = "=SUM(E{}:E{})".format(11, 11 + idx - 1)
            sheet['F{}'.format(11 + idx)] = "=SUM(F{}:F{})".format(11, 11 + idx - 1)
            sheet['G{}'.format(11 + idx)] = "=SUM(G{}:G{})".format(11, 11 + idx - 1)
            sheet['B{}'.format(
                11 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'
        return {
            'col': 'G',
            'row': 11 + idx
        }

    def setupSheetPenguranAbsensi(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            numLate = 0
            numLate2 = 0
            numAT = 0
            for d in p.worked_days_line_ids:
                if d.code == 'terlambat':
                    numLate = d.number_of_days
                if d.code == 'terlambat2jam':
                    numLate2 = d.number_of_days
                if d.code == 'AT':
                    numAT = d.number_of_days

            sheet.insert_rows(11 + idx)
            sheet['A{}'.format(11 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(11 + idx)].border = thin_border
            sheet['B{}'.format(11 + idx)] = p.employee_id.name
            sheet['B{}'.format(11 + idx)].border = thin_border
            sheet['C{}'.format(11 + idx)] = numLate
            sheet['C{}'.format(11 + idx)].border = thin_border
            sheet['D{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(11 + idx)].border = thin_border
            sheet['D{}'.format(11 + idx)] = "=C{}*20000".format(11 + idx)
            sheet['E{}'.format(11 + idx)] = numLate2
            sheet['E{}'.format(11 + idx)].border = thin_border
            sheet['F{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(11 + idx)].border = thin_border
            sheet['F{}'.format(11 + idx)] = "=E{}*40000".format(11 + idx)
            sheet['G{}'.format(11 + idx)] = numAT
            sheet['G{}'.format(11 + idx)].border = thin_border
            sheet['H{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['H{}'.format(11 + idx)].border = thin_border
            sheet['H{}'.format(11 + idx)] = "=G{}*50000".format(11 + idx)
            sheet['I{}'.format(11 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['I{}'.format(11 + idx)].border = thin_border
            sheet['I{}'.format(11 + idx)] = "=D{}+F{}+H{}".format(11 + idx, 11 + idx, 11 + idx)
            idx += 1
        if idx > 0:
            sheet['C{}'.format(11 + idx)] = "=SUM(C{}:C{})".format(11, 11 + idx - 1)
            sheet['D{}'.format(11 + idx)] = "=SUM(D{}:D{})".format(11, 11 + idx - 1)
            sheet['E{}'.format(11 + idx)] = "=SUM(E{}:E{})".format(11, 11 + idx - 1)
            sheet['F{}'.format(11 + idx)] = "=SUM(F{}:F{})".format(11, 11 + idx - 1)
            sheet['G{}'.format(11 + idx)] = "=SUM(G{}:G{})".format(11, 11 + idx - 1)
            sheet['H{}'.format(11 + idx)] = "=SUM(H{}:H{})".format(11, 11 + idx - 1)
            sheet['I{}'.format(11 + idx)] = "=SUM(I{}:I{})".format(11, 11 + idx - 1)
            sheet['B{}'.format(
                11 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'
        return {
            'col': 'I',
            'row': 11 + idx
        }

    def setupSheetKopen(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            amount = 0
            for line in p.input_line_ids:
                if line.code == 'KOPEN':
                    amount = line.amount
            iuran = 0
            angsuran = 0
            kopen = self.env["hr.payslip.manual"].search(
                [('name', '=', p.date_to), ('employee_id', '=', p.employee_id.id)])
            for kop in kopen:
                iuran = kop.iuran_wajib_kopen
                angsuran = kop.angsuran_kopen
            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = iuran
            sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['E{}'.format(10 + idx)].border = thin_border
            sheet['E{}'.format(10 + idx)] = angsuran
            sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)] = amount
            idx += 1
        if idx > 0:
            sheet['D{}'.format(10 + idx)] = "=SUM(D{}:D{})".format(10, 10 + idx - 1)
            sheet['E{}'.format(10 + idx)] = "=SUM(E{}:E{})".format(10, 10 + idx - 1)
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

    def setupSheetZiebar(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            amount = 0
            for line in p.input_line_ids:
                if line.code == 'ZIEBAR':
                    amount = line.amount
            iuran = 0
            angsuran = 0
            zieb = self.env["hr.payslip.manual"].search(
                [('name', '=', p.date_to), ('employee_id', '=', p.employee_id.id)])
            for zie in zieb:
                iuran = zie.iuran_wajib_ziebar
                angsuran = zie.angsuran_ziebar
            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = iuran
            sheet['E{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['E{}'.format(10 + idx)].border = thin_border
            sheet['E{}'.format(10 + idx)] = angsuran
            sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)] = amount
            idx += 1
        if idx > 0:
            sheet['D{}'.format(10 + idx)] = "=SUM(D{}:D{})".format(10, 10 + idx - 1)
            sheet['E{}'.format(10 + idx)] = "=SUM(E{}:E{})".format(10, 10 + idx - 1)
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

    def setupSheetBjbSyariah(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            amount = 0
            for line in p.input_line_ids:
                if line.code == 'BJBS':
                    amount = line.amount
            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = grade
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = amount
            idx += 1
        if idx > 0:
            sheet['D{}'.format(10 + idx)] = "=SUM(D{}:D{})".format(10, 10 + idx - 1)
            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

    def setupSheetPinbuk(self, payslips, sheet):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal {}'.format(datetime.strptime(payslips[1].date_from,"%Y-%M-%d").strftime("%d %b %Y"))
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            sheet.insert_rows(10 + idx)
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)].border = thin_border
            if p.employee_id.bank_account_id.acc_number:
                sheet['C{}'.format(10 + idx)] = p.employee_id.bank_account_id.acc_number
            else:
                sheet['C{}'.format(10 + idx)] = "-"
            sheet['D{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = "='POTONGAN GAJI PEGAWAI'!V{}".format(10 + idx)
            idx += 1
        if idx > 0:
            sheet['D{}'.format(10 + idx)] = "=SUM(D{}:D{})".format(10, 10 + idx - 1)

            sheet['B{}'.format(
                10 + idx + 2)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

    def setupSheetGaji(self, payslips, sheet, dataDplk, dataBpjsTk, dataBpjsKes, dataKH, dataAb):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name == "PEGAWAI TETAP":
                single = p.single_salary
            elif p.contract_id.type_id.name == "CALON PEGAWAI":
                single = (p.single_salary * 80) / 100
            else:
                single = p.contract_id.wage

            if p.contract_id.type_id.name == "CAPEG LAMA":
                dplk = 0
                pot_dplk = 0
            else:
                dplk = p.tunjangan_dplk
                pot_dplk = p.potongan_dplk
            if p.employee_id.tanggal_masuk != False:
                masa = ((datetime.now() - datetime.strptime(p.employee_id.tanggal_masuk,
                                                            "%Y-%M-%d")).days / 30) / 12
                tahun = int(masa)
                bulan = str(masa - tahun)[2:3]
            else:
                tahun = 0
                bulan = 0
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            sheet.insert_rows(10 + idx)
            # sheet['G{}'.format(10 + idx)].set_border =
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = p.employee_id.tanggal_pengangkatan
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = str(tahun) + ' Tahun'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = str(bulan) + ' Bulan'
            sheet['E{}'.format(10 + idx)].border = thin_border
            sheet['E{}'.format(10 + idx)] = grade
            sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)] = single
            sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['G{}'.format(10 + idx)].border = thin_border
            sheet['G{}'.format(10 + idx)] = "='DPLK'!E{}".format(10 + idx)
            sheet['H{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['H{}'.format(10 + idx)].border = thin_border
            sheet['H{}'.format(10 + idx)] = "='BPJS TK'!I{}".format(10 + idx)
            sheet['I{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['I{}'.format(10 + idx)].border = thin_border
            sheet['I{}'.format(10 + idx)] = "='BPJS KES'!E{}".format(10 + idx)
            sheet['J{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['J{}'.format(10 + idx)].border = thin_border
            sheet['J{}'.format(10 + idx)] = "=SUM(G{}:I{})".format(10 + idx, 10 + idx)
            sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['K{}'.format(10 + idx)].border = thin_border
            sheet['K{}'.format(10 + idx)] = "='PENGURANG KEHADIRAN'!G{}".format(11 + idx)
            sheet['L{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['L{}'.format(10 + idx)].border = thin_border
            sheet['L{}'.format(10 + idx)] = "='PENGURANG ABSENSI'!I{}".format(11 + idx)
            sheet['M{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['M{}'.format(10 + idx)].border = thin_border
            sheet['M{}'.format(10 + idx)] = "==K{}+L{}".format(10 + idx, 10 + idx)
            sheet['N{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['N{}'.format(10 + idx)].border = thin_border
            sheet['N{}'.format(10 + idx)] = "=F{}+J{}-M{}".format(10 + idx, 10 + idx, 10 + idx)
            sheet['O{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['O{}'.format(10 + idx)].border = thin_border
            sheet['O{}'.format(10 + idx)] = "='DPLK'!F{}".format(10 + idx)
            sheet['P{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['P{}'.format(10 + idx)].border = thin_border
            sheet['P{}'.format(10 + idx)] = "='BPJS TK'!L{}".format(10 + idx)
            sheet['Q{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['Q{}'.format(10 + idx)].border = thin_border
            sheet['Q{}'.format(10 + idx)] = "='BPJS KES'!F{}".format(10 + idx)
            sheet['R{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['R{}'.format(10 + idx)].border = thin_border
            sheet['R{}'.format(10 + idx)] = "='BJB SYARIAH'!D{}".format(10 + idx)
            sheet['S{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['S{}'.format(10 + idx)].border = thin_border
            sheet['S{}'.format(10 + idx)] = "='KOPEN'!F{}".format(10 + idx)
            sheet['T{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['T{}'.format(10 + idx)].border = thin_border
            sheet['T{}'.format(10 + idx)] = "='ZIEBAR'!F{}".format(10 + idx)
            sheet['U{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['U{}'.format(10 + idx)].border = thin_border
            sheet['U{}'.format(10 + idx)] = "=SUM(O{}:T{})".format(10 + idx, 10 + idx)
            sheet['V{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['V{}'.format(10 + idx)].border = thin_border
            sheet['V{}'.format(10 + idx)] = "=ROUND((F{} - M{} - U{}), 0)".format(10 + idx, 10 + idx, 10 + idx)
            idx += 1
        if idx > 0:
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['G{}'.format(10 + idx)] = "=SUM(G{}:G{})".format(10, 10 + idx - 1)
            sheet['H{}'.format(10 + idx)] = "=SUM(H{}:H{})".format(10, 10 + idx - 1)
            sheet['I{}'.format(10 + idx)] = "=SUM(I{}:I{})".format(10, 10 + idx - 1)
            sheet['J{}'.format(10 + idx)] = "=SUM(J{}:J{})".format(10, 10 + idx - 1)
            sheet['K{}'.format(10 + idx)] = "=SUM(K{}:K{})".format(10, 10 + idx - 1)
            sheet['L{}'.format(10 + idx)] = "=SUM(L{}:L{})".format(10, 10 + idx - 1)
            sheet['M{}'.format(10 + idx)] = "=SUM(M{}:M{})".format(10, 10 + idx - 1)
            sheet['N{}'.format(10 + idx)] = "=SUM(N{}:N{})".format(10, 10 + idx - 1)
            sheet['O{}'.format(10 + idx)] = "=SUM(O{}:O{})".format(10, 10 + idx - 1)
            sheet['P{}'.format(10 + idx)] = "=SUM(P{}:P{})".format(10, 10 + idx - 1)
            sheet['Q{}'.format(10 + idx)] = "=SUM(Q{}:Q{})".format(10, 10 + idx - 1)
            sheet['R{}'.format(10 + idx)] = "=SUM(R{}:R{})".format(10, 10 + idx - 1)
            sheet['S{}'.format(10 + idx)] = "=SUM(S{}:S{})".format(10, 10 + idx - 1)
            sheet['T{}'.format(10 + idx)] = "=SUM(T{}:T{})".format(10, 10 + idx - 1)
            sheet['U{}'.format(10 + idx)] = "=SUM(U{}:U{})".format(10, 10 + idx - 1)
            sheet['V{}'.format(10 + idx)] = "=SUM(V{}:V{})".format(10, 10 + idx - 1)

            sheet['E{}'.format(10 + idx + 3)] = "=DPLK!{}{}".format(dataDplk['col'], dataDplk['row'])
            sheet['E{}'.format(10 + idx + 4)] = "='BPJS TK'!{}{}".format(dataBpjsTk['col'], dataBpjsTk['row'])
            sheet['E{}'.format(10 + idx + 5)] = "='BPJS KES'!{}{}".format(dataBpjsKes['col'], dataBpjsKes['row'])
            sheet['F{}'.format(10 + idx + 3)] = "=DPLK!{}{}".format(dataDplk['col2'], dataDplk['row2'])
            sheet['F{}'.format(10 + idx + 4)] = "='BPJS TK'!{}{}".format(dataBpjsTk['col2'], dataBpjsTk['row2'])
            sheet['F{}'.format(10 + idx + 5)] = "='BPJS KES'!{}{}".format(dataBpjsKes['col2'], dataBpjsKes['row2'])
            sheet['G{}'.format(10 + idx + 3)] = "=E{}+F{}".format(10 + idx + 3, 10 + idx + 3)
            sheet['G{}'.format(10 + idx + 4)] = "=E{}+F{}".format(10 + idx + 4, 10 + idx + 4)
            sheet['G{}'.format(10 + idx + 5)] = "=E{}+F{}".format(10 + idx + 5, 10 + idx + 5)

            sheet['E{}'.format(10 + idx + 8)] = "='PENGURANG KEHADIRAN'!{}{}".format(dataKH['col'], dataKH['row'])
            sheet['E{}'.format(10 + idx + 9)] = "='PENGURANG ABSENSI'!{}{}".format(dataAb['col'], dataAb['row'])

            sheet['B{}'.format(
                10 + idx + 11)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

        return 10 + idx - 1

    def setupSheetpotongan(self, payslips, sheet, dataDplk, dataBpjsTk, dataBpjsKes, dataKH, dataAb):
        thin_border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))
        if payslips:
            # if self.nomor != False :
            #    sheet['A1'] = 'Lampiran Memo Nomor :' + self.nomor
            # else :
            #    sheet['A1'] = 'Lampiran Memo Nomor :'
            # sheet['A2'] = 'Tanggal ' + str(payslips[1].date_to[8:]) + ' ' + self.MONTHS[self.month - 1] + ' ' +str(payslips[1].date_to[:4])
            # sheet['A3'] = 'Perihal : Pembayaran Gaji Pegawai YKP bank bjb Bulan {} {}'.format(self.MONTHS[self.month - 1],
            #                                                                                  self.year)
            sheet['A6'] = 'BULAN {} {} PEGAWAI YKP bank bjb'.format(self.MONTHS[self.month - 1], self.year)
        idx = 0
        for p in payslips:
            if p.contract_id.type_id.name == "PEGAWAI TETAP":
                single = p.single_salary
            elif p.contract_id.type_id.name == "CALON PEGAWAI":
                single = (p.single_salary * 80) / 100
            else:
                single = p.contract_id.wage

            if p.contract_id.type_id.name == "CAPEG LAMA":
                dplk = 0
                pot_dplk = 0
            else:
                dplk = p.tunjangan_dplk
                pot_dplk = p.potongan_dplk
            if p.employee_id.tanggal_masuk != False:
                masa = ((datetime.now() - datetime.strptime(p.employee_id.tanggal_masuk,
                                                            "%Y-%M-%d")).days / 30) / 12
                tahun = int(masa)
                bulan = str(masa - tahun)[2:3]
            else:
                tahun = 0
                bulan = 0
            if p.contract_id.type_id.name != "PARTNERSHIP" and p.contract_id.type_id.name != "KONTRAK":
                if p.employee_id.grade and p.employee_id.jenis_pegawai:
                    grade = "G-" + p.employee_id.grade.name + "/" + p.employee_id.grade.level_id + " (" + p.employee_id.jenis_pegawai.name + ")"
            else:
                grade = p.employee_id.jenis_pegawai.name
            sheet.insert_rows(10 + idx)
            # sheet['G{}'.format(10 + idx)].set_border =
            sheet['A{}'.format(10 + idx)] = int('{}'.format(idx + 1))
            sheet['A{}'.format(10 + idx)].border = thin_border
            sheet['B{}'.format(10 + idx)] = p.employee_id.name
            sheet['B{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = p.employee_id.tanggal_pengangkatan
            sheet['C{}'.format(10 + idx)].border = thin_border
            sheet['C{}'.format(10 + idx)] = str(tahun) + ' Tahun'
            sheet['D{}'.format(10 + idx)].border = thin_border
            sheet['D{}'.format(10 + idx)] = str(bulan) + ' Bulan'
            sheet['E{}'.format(10 + idx)] = grade
            sheet['E{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['F{}'.format(10 + idx)].border = thin_border
            sheet['F{}'.format(10 + idx)] = single
            sheet['G{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['G{}'.format(10 + idx)].border = thin_border
            sheet['G{}'.format(10 + idx)] = "='DPLK'!E{}".format(10 + idx)  # dplk
            sheet['H{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['H{}'.format(10 + idx)].border = thin_border
            sheet['H{}'.format(10 + idx)] = "='BPJS TK'!I{}".format(10 + idx)
            sheet['I{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['I{}'.format(10 + idx)].border = thin_border
            sheet['I{}'.format(10 + idx)] = "='BPJS KES'!E{}".format(10 + idx)
            sheet['J{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['J{}'.format(10 + idx)].border = thin_border
            sheet['J{}'.format(10 + idx)] = "=SUM(G{}:I{})".format(10 + idx, 10 + idx)
            sheet['K{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['K{}'.format(10 + idx)].border = thin_border
            sheet['K{}'.format(10 + idx)] = "='PENGURANG KEHADIRAN'!G{}".format(11 + idx)
            sheet['L{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['L{}'.format(10 + idx)].border = thin_border
            sheet['L{}'.format(10 + idx)] = "='PENGURANG ABSENSI'!I{}".format(11 + idx)
            sheet['M{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['M{}'.format(10 + idx)].border = thin_border
            sheet['M{}'.format(10 + idx)] = "==K{}+L{}".format(10 + idx, 10 + idx)
            sheet['N{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['N{}'.format(10 + idx)].border = thin_border
            sheet['N{}'.format(10 + idx)] = "=F{}+J{}-M{}".format(10 + idx, 10 + idx, 10 + idx)
            sheet['O{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['O{}'.format(10 + idx)].border = thin_border
            sheet['O{}'.format(10 + idx)] = "='DPLK'!F{}".format(10 + idx)
            sheet['P{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['P{}'.format(10 + idx)].border = thin_border
            sheet['P{}'.format(10 + idx)] = "='BPJS TK'!L{}".format(10 + idx)
            sheet['Q{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['Q{}'.format(10 + idx)].border = thin_border
            sheet['Q{}'.format(10 + idx)] = "='BPJS KES'!F{}".format(10 + idx)
            sheet['R{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['R{}'.format(10 + idx)].border = thin_border
            sheet['R{}'.format(10 + idx)] = "='BJB SYARIAH'!D{}".format(10 + idx)
            sheet['S{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['S{}'.format(10 + idx)].border = thin_border
            sheet['S{}'.format(10 + idx)] = "='KOPEN'!F{}".format(10 + idx)
            sheet['T{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['T{}'.format(10 + idx)].border = thin_border
            sheet['T{}'.format(10 + idx)] = "='ZIEBAR'!F{}".format(10 + idx)
            sheet['U{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['U{}'.format(10 + idx)].border = thin_border
            sheet['U{}'.format(10 + idx)] = "=SUM(O{}:T{})".format(10 + idx, 10 + idx)
            sheet['V{}'.format(10 + idx)].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
            sheet['V{}'.format(10 + idx)].border = thin_border
            sheet['V{}'.format(10 + idx)] = "=ROUND((F{} - M{} - U{}), 0)".format(10 + idx, 10 + idx, 10 + idx)
            idx += 1
        if idx > 0:
            sheet['F{}'.format(10 + idx)] = "=SUM(F{}:F{})".format(10, 10 + idx - 1)
            sheet['G{}'.format(10 + idx)] = "=SUM(G{}:G{})".format(10, 10 + idx - 1)
            sheet['H{}'.format(10 + idx)] = "=SUM(H{}:H{})".format(10, 10 + idx - 1)
            sheet['I{}'.format(10 + idx)] = "=SUM(I{}:I{})".format(10, 10 + idx - 1)
            sheet['J{}'.format(10 + idx)] = "=SUM(J{}:J{})".format(10, 10 + idx - 1)
            sheet['K{}'.format(10 + idx)] = "=SUM(K{}:K{})".format(10, 10 + idx - 1)
            sheet['L{}'.format(10 + idx)] = "=SUM(L{}:L{})".format(10, 10 + idx - 1)
            sheet['M{}'.format(10 + idx)] = "=SUM(M{}:M{})".format(10, 10 + idx - 1)
            sheet['N{}'.format(10 + idx)] = "=SUM(N{}:N{})".format(10, 10 + idx - 1)
            sheet['O{}'.format(10 + idx)] = "=SUM(O{}:O{})".format(10, 10 + idx - 1)
            sheet['P{}'.format(10 + idx)] = "=SUM(P{}:P{})".format(10, 10 + idx - 1)
            sheet['Q{}'.format(10 + idx)] = "=SUM(Q{}:Q{})".format(10, 10 + idx - 1)
            sheet['R{}'.format(10 + idx)] = "=SUM(R{}:R{})".format(10, 10 + idx - 1)
            sheet['S{}'.format(10 + idx)] = "=SUM(S{}:S{})".format(10, 10 + idx - 1)
            sheet['T{}'.format(10 + idx)] = "=SUM(T{}:T{})".format(10, 10 + idx - 1)
            sheet['U{}'.format(10 + idx)] = "=SUM(U{}:U{})".format(10, 10 + idx - 1)
            sheet['V{}'.format(10 + idx)] = "=SUM(V{}:V{})".format(10, 10 + idx - 1)

            sheet['E{}'.format(10 + idx + 3)] = "=DPLK!{}{}".format(dataDplk['col'], dataDplk['row'])
            sheet['E{}'.format(10 + idx + 4)] = "='BPJS TK'!{}{}".format(dataBpjsTk['col'], dataBpjsTk['row'])
            sheet['E{}'.format(10 + idx + 5)] = "='BPJS KES'!{}{}".format(dataBpjsKes['col'], dataBpjsKes['row'])
            sheet['F{}'.format(10 + idx + 3)] = "=DPLK!{}{}".format(dataDplk['col2'], dataDplk['row2'])
            sheet['F{}'.format(10 + idx + 4)] = "='BPJS TK'!{}{}".format(dataBpjsTk['col2'], dataBpjsTk['row2'])
            sheet['F{}'.format(10 + idx + 5)] = "='BPJS KES'!{}{}".format(dataBpjsKes['col2'], dataBpjsKes['row2'])
            sheet['K{}'.format(10 + idx + 3)] = "=E{}+F{}".format(10 + idx + 3, 10 + idx + 3)
            sheet['K{}'.format(10 + idx + 4)] = "=E{}+F{}".format(10 + idx + 4, 10 + idx + 4)
            sheet['K{}'.format(10 + idx + 5)] = "=E{}+F{}".format(10 + idx + 5, 10 + idx + 5)

            sheet['E{}'.format(10 + idx + 8)] = "='PENGURANG KEHADIRAN'!{}{}".format(dataKH['col'], dataKH['row'])
            sheet['E{}'.format(10 + idx + 9)] = "='PENGURANG ABSENSI'!{}{}".format(dataAb['col'], dataAb['row'])

            sheet['B{}'.format(
                10 + idx + 11)] = '* Perhitungan ini dicetak melalui sistem dan tidak memerlukan tanda tangan'

        return 10 + idx - 1

    def _protectSheet(self, sheet):
        sheet.protection.password = "YakinBisa123!"
        sheet.protection.sheet = True
        sheet.protection.enable()

    def generate_excel(self):
        date_from = "{}-{}-{}".format(self.year, self.month if self.month > 9 else "0{}".format(self.month), "01")
        date_to = "{}-{}-{}".format(self.year, self.month if self.month > 9 else "0{}".format(self.month), "28")
        payslips = self.env['hr.payslip'].search([('date_to', '>=', date_from), ('date_to', '<=', date_to)])

        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/report_payslip.xlsx".format(path)
        wb = load_workbook(filename)
        dataDplk = self.setupSheetDPLK(payslips, wb['DPLK'])
        dataBpjsTk = self.setupSheetBPJSTK(payslips, wb['BPJS TK'])
        dataBpjsKes = self.setupSheetBpjsKes(payslips, wb['BPJS KES'])
        dataKH = self.setupSheetPenguranKehadiran(payslips, wb['PENGURANG KEHADIRAN'])
        dataAb = self.setupSheetPenguranAbsensi(payslips, wb['PENGURANG ABSENSI'])
        self.setupSheetKopen(payslips, wb['KOPEN'])
        self.setupSheetZiebar(payslips, wb['ZIEBAR'])
        self.setupSheetBjbSyariah(payslips, wb['BJB SYARIAH'])
        self.setupSheetPinbuk(payslips, wb['DAFTAR PINBUK'])
        self.setupSheetpotongan(payslips, wb['POTONGAN GAJI PEGAWAI'], dataDplk, dataBpjsTk, dataBpjsKes, dataKH,
                                dataAb)
        length = self.setupSheetGaji(payslips, wb['TUNJANGAN PEGAWAI (GAJI)'], dataDplk, dataBpjsTk, dataBpjsKes,
                                     dataKH, dataAb)
        for sheet in wb.worksheets:
            self._protectSheet(sheet)
        wb.save(filename="{}/../reports/report_payslip_{}_{}.xlsx".format(path, self.month, self.year))
        return {
            'name': 'Payslip',
            'type': 'ir.actions.act_url',
            'url': '/payroll/download/{}{}'.format(self.year, self.month),
            'target': 'self',
        }


class ptkp(models.Model):
    _name = "hr.ptkp"
    _rec_name = "kode"

    kode = fields.Char("Kode", required=True)
    nominal_bulan = fields.Float("Nominal Perbulan", required=True)
    nominal_tahun = fields.Float("Nominal Pertahun", required=True)


ptkp()


class pkp(models.Model):
    _name = "hr.pkp"
    _rec_name = "kode"

    kode = fields.Char("Kode", required=True)
    nominal_mix = fields.Float("Nominal Min", required=True)
    nominal_max = fields.Float("Nominal Max", required=True)
    pajak = fields.Float("Pajak (%)", required=True,  digits=(1,10))
    penambah = fields.Float('Penambah')

pkp()


class HrEmployee(models.Model):
    _name = 'hr.employee'
    _inherit = "hr.employee"

    status_pajak = fields.Many2one(comodel_name="hr.ptkp", string="Status Pajak")


HrEmployee()


class PajakWizard(models.TransientModel):
    _name = "hr.pajak.report.wizard"

    def generate_pajak(self):
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'id_ID')
        except:
            print('error setting locale')

        DATETIME_FORMAT = "%Y-%m-%d"
        d1 = datetime.strptime(self.date_from, DATETIME_FORMAT)
        d2 = datetime.strptime(self.date_to, DATETIME_FORMAT)
        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/report_pajak.xlsx".format(path)
        wb = load_workbook(filename)
        sheet = wb['Sheet1']
        idx = 0
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for employee_id in self.employee_ids:
            domain = [('date_from', '>=', d1.strftime(DATETIME_FORMAT)),
                      ('date_to', '<=', d2.strftime(DATETIME_FORMAT)), ('employee_id', '=', employee_id.id)]
            payslips = self.env['hr.payslip'].search(domain)
            # lembur
            overtimes = self.env['hr.overtime'].search(
                [('state', '=', 'validate_realisasi'), ('realisasi_date_from', '>=', d1.strftime(DATETIME_FORMAT)),
                 ('realisasi_date_from', '<=', d2.strftime(DATETIME_FORMAT)), ('employee_id', '=', employee_id.id)])
            lembur = 0
            for overtime in overtimes:
                lembur += overtime.uang_lembur
            # perdin
            perdins = self.env['hr.perdin.employee'].search(
                [('state', '=', 'validate'), ('date_from', '>=', d1.strftime(DATETIME_FORMAT)),
                 ('date_to', '<=', d2.strftime(DATETIME_FORMAT)), ('employee_id', '=', employee_id.id)])
            perdin = 0
            for obj_perdin in perdins:
                perdin += obj_perdin.uang_saku
            # pelatihan
            trainings = self.env['hr.training.participant'].search(
                [('state', '=', 'approve'), ('start_date', '>=', d1.strftime(DATETIME_FORMAT)),
                 ('end_date', '<=', d2.strftime(DATETIME_FORMAT)), ('name', '=', employee_id.id)])
            diklat = 0
            for training in trainings:
                diklat += training.total_uang_pelatihan
            gaji = 0
            asuransi = 0
            potongan = 0
            tunj_dplk = 0
            tunj_bpjstk = 0
            tunj_bpjskes = 0
            for payslip in payslips:
                gaji += payslip.single_salary
                tunj_dplk += payslip.tunjangan_dplk
                tunj_bpjstk += payslip.tunj_bpjstk
                tunj_bpjskes += payslip.tunj_bpjskes
                asuransi += payslip.tunjangan_dplk + payslip.tunj_bpjstk + payslip.tunj_bpjskes
                potongan += payslip.potongan_dplk + payslip.pot_bpjstk + payslip.pot_bpjskes
            if employee_id.status_pajak:
                ptkp = employee_id.status_pajak.nominal_tahun
            else:
                ptkp = 0
            sheet.insert_rows(5 + idx)
            sheet['A{}'.format(5 + idx)] = idx + 1
            sheet['A{}'.format(5 + idx)].border = thin_border
            sheet['B{}'.format(5 + idx)] = employee_id.name
            sheet['B{}'.format(5 + idx)].border = thin_border
            sheet['C{}'.format(5 + idx)] = employee_id.no_npwp if employee_id.no_npwp else ''
            sheet['C{}'.format(5 + idx)].border = thin_border
            sheet['D{}'.format(5 + idx)] = employee_id.nik if employee_id.nik else ''
            sheet['D{}'.format(5 + idx)].border = thin_border
            sheet['E{}'.format(5 + idx)] = employee_id.status_pajak.kode if employee_id.status_pajak else ''
            sheet['E{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)] = gaji
            sheet['F{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['G{}'.format(5 + idx)].border = thin_border
            sheet['G{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['H{}'.format(5 + idx)] = 0
            sheet['H{}'.format(5 + idx)].border = thin_border
            sheet['H{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['I{}'.format(5 + idx)] = 0
            sheet['I{}'.format(5 + idx)].border = thin_border
            sheet['I{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['J{}'.format(5 + idx)] = 0
            sheet['J{}'.format(5 + idx)].border = thin_border
            sheet['J{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['K{}'.format(5 + idx)] = lembur
            sheet['K{}'.format(5 + idx)].border = thin_border
            sheet['K{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['L{}'.format(5 + idx)] = diklat
            sheet['L{}'.format(5 + idx)].border = thin_border
            sheet['L{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['M{}'.format(5 + idx)] = perdin
            sheet['M{}'.format(5 + idx)].border = thin_border
            sheet['M{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['N{}'.format(5 + idx)] = "=SUM(H{}:M{})".format(5 + idx, 5 + idx)
            sheet['N{}'.format(5 + idx)].border = thin_border
            sheet['N{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['O{}'.format(5 + idx)] = "=SUM(H{}:J{})".format(5 + idx, 5 + idx)
            sheet['O{}'.format(5 + idx)].border = thin_border
            sheet['O{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['P{}'.format(5 + idx)] = 0
            sheet['P{}'.format(5 + idx)].border = thin_border
            sheet['P{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Q{}'.format(5 + idx)] = tunj_bpjstk
            sheet['Q{}'.format(5 + idx)].border = thin_border
            sheet['Q{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['R{}'.format(5 + idx)] = tunj_dplk
            sheet['R{}'.format(5 + idx)].border = thin_border
            sheet['R{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['S{}'.format(5 + idx)] = tunj_bpjskes
            sheet['S{}'.format(5 + idx)].border = thin_border
            sheet['S{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['T{}'.format(5 + idx)] = "=SUM(Q{}:S{})".format(5 + idx, 5 + idx)
            sheet['T{}'.format(5 + idx)].border = thin_border
            sheet['T{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['U{}'.format(5 + idx)] = 0
            sheet['U{}'.format(5 + idx)].border = thin_border
            sheet['U{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['V{}'.format(5 + idx)] = 0
            sheet['V{}'.format(5 + idx)].border = thin_border
            sheet['V{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['W{}'.format(5 + idx)] = "=F{} + G{} + O{} + T{}".format(5 + idx, 5 + idx, 5 + idx, 5 + idx)
            sheet['W{}'.format(5 + idx)].border = thin_border
            sheet['W{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['X{}'.format(5 + idx)] = "=F{} + G{} + N{} + T{} + U{} + P{} + V{}".format(5 + idx, 5 + idx, 5 + idx,
                                                                                             5 + idx, 5 + idx, 5 + idx,
                                                                                             5 + idx)
            sheet['X{}'.format(5 + idx)].border = thin_border
            sheet['X{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Y{}'.format(5 + idx)] = "=ROUND(IF((W{}*5/100)>6000000,6000000,(W{}*5/100)),0)".format(5 + idx,
                                                                                                          5 + idx)
            sheet['Y{}'.format(5 + idx)].border = thin_border
            sheet['Y{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Z{}'.format(5 + idx)] = "=ROUND(IF((X{}*5/100)>6000000,6000000,(X{}*5/100)),0)".format(5 + idx,
                                                                                                          5 + idx)
            sheet['Z{}'.format(5 + idx)].border = thin_border
            sheet['Z{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AA{}'.format(5 + idx)] = potongan
            sheet['AA{}'.format(5 + idx)].border = thin_border
            sheet['AA{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AB{}'.format(5 + idx)] = "=Y{} + AA{}".format(5 + idx, 5 + idx)
            sheet['AB{}'.format(5 + idx)].border = thin_border
            sheet['AB{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AC{}'.format(5 + idx)] = "=Z{} + AA{}".format(5 + idx, 5 + idx)
            sheet['AC{}'.format(5 + idx)].border = thin_border
            sheet['AC{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AD{}'.format(5 + idx)] = "=W{} - AB{}".format(5 + idx, 5 + idx)
            sheet['AD{}'.format(5 + idx)].border = thin_border
            sheet['AD{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AE{}'.format(5 + idx)] = "=X{} - AC{}".format(5 + idx, 5 + idx)
            sheet['AE{}'.format(5 + idx)].border = thin_border
            sheet['AE{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AF{}'.format(5 + idx)] = 0
            sheet['AF{}'.format(5 + idx)].border = thin_border
            sheet['AF{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AG{}'.format(5 + idx)] = "=AD{} + AF{}".format(5 + idx, 5 + idx)
            sheet['AG{}'.format(5 + idx)].border = thin_border
            sheet['AG{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AH{}'.format(5 + idx)] = "=AE{} + AF{}".format(5 + idx, 5 + idx)
            sheet['AH{}'.format(5 + idx)].border = thin_border
            sheet['AH{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AI{}'.format(5 + idx)] = ptkp
            sheet['AI{}'.format(5 + idx)].border = thin_border
            sheet['AI{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AJ{}'.format(5 + idx)] = "=ROUNDDOWN(IF(AG{}-AI{}>0,AG{}-AI{},0),-3)".format(5 + idx, 5 + idx,
                                                                                                5 + idx, 5 + idx)
            sheet['AJ{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AJ{}'.format(5 + idx)].border = thin_border
            sheet['AK{}'.format(5 + idx)] = "=ROUNDDOWN(IF(AH{}-AI{}>0,AH{}-AI{},0),-3)".format(5 + idx, 5 + idx,
                                                                                                5 + idx, 5 + idx)
            sheet['AK{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AK{}'.format(5 + idx)].border = thin_border
            sheet['AL{}'.format(
                5 + idx)] = "=ROUND((IF(AJ{}>50000000,((5/100*50000000)+(15/100*(AJ{}-50000000))),IF(AJ{}>0,(5/100*AJ{}),0)))*IF(C{}=\"\",1.5,1),0)".format(
                5 + idx, 5 + idx,
                5 + idx, 5 + idx, 5 + idx)
            _logger.info(
                "=ROUND((IF(AJ{}>50000000,((5/100*50000000)+(15/100*(AJ{}-50000000))),IF(AJ{}>0,(5/100*AJ{}),0)))*IF(C{}=\"\",1.5,1),0)".format(
                    5 + idx, 5 + idx,
                    5 + idx, 5 + idx, 5 + idx))
            sheet['AL{}'.format(5 + idx)].border = thin_border
            sheet['AL{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AM{}'.format(
                5 + idx)] = "=ROUND((IF(AK{}>50000000,((5/100*50000000)+(15/100*(AK{}-50000000))),IF(AK{}>0,(5/100*AK{}),0)))*IF(C{}=\"\",1.5,1),0)".format(
                5 + idx, 5 + idx,
                5 + idx, 5 + idx, 5 + idx)
            sheet['AM{}'.format(5 + idx)].border = thin_border
            sheet['AM{}'.format(5 + idx)].number_format = '#,##0.00'
            idx += 1
        if idx > 0:
            sheet.insert_rows(5 + idx)
            sheet['A{}'.format(5 + idx)].border = thin_border
            sheet['B{}'.format(5 + idx)].border = thin_border
            sheet['C{}'.format(5 + idx)].border = thin_border
            sheet['D{}'.format(5 + idx)].border = thin_border
            sheet['E{}'.format(5 + idx)] = "TOTAL"
            sheet['E{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)] = "=SUM(F{}:F{})".format(5, idx)
            sheet['F{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['G{}'.format(5 + idx)] = "=SUM(G{}:G{})".format(5, idx)
            sheet['G{}'.format(5 + idx)].border = thin_border
            sheet['G{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['H{}'.format(5 + idx)] = "=SUM(H{}:H{})".format(5, idx)
            sheet['H{}'.format(5 + idx)].border = thin_border
            sheet['H{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['I{}'.format(5 + idx)] = "=SUM(I{}:I{})".format(5, idx)
            sheet['I{}'.format(5 + idx)].border = thin_border
            sheet['I{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['J{}'.format(5 + idx)] = "=SUM(J{}:J{})".format(5, idx)
            sheet['J{}'.format(5 + idx)].border = thin_border
            sheet['J{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['K{}'.format(5 + idx)] = "=SUM(K{}:K{})".format(5, idx)
            sheet['K{}'.format(5 + idx)].border = thin_border
            sheet['K{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['L{}'.format(5 + idx)] = "=SUM(L{}:L{})".format(5, idx)
            sheet['L{}'.format(5 + idx)].border = thin_border
            sheet['L{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['M{}'.format(5 + idx)] = "=SUM(M{}:M{})".format(5, idx)
            sheet['M{}'.format(5 + idx)].border = thin_border
            sheet['M{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['N{}'.format(5 + idx)] = "=SUM(N{}:N{})".format(5, 5 + idx)
            sheet['N{}'.format(5 + idx)].border = thin_border
            sheet['N{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['O{}'.format(5 + idx)] = "=SUM(O{}:O{})".format(5, 5 + idx)
            sheet['O{}'.format(5 + idx)].border = thin_border
            sheet['O{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['P{}'.format(5 + idx)] = "=SUM(P{}:P{})".format(5, 5 + idx)
            sheet['P{}'.format(5 + idx)].border = thin_border
            sheet['P{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Q{}'.format(5 + idx)] = "=SUM(Q{}:Q{})".format(5, 5 + idx)
            sheet['Q{}'.format(5 + idx)].border = thin_border
            sheet['Q{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['R{}'.format(5 + idx)] = "=SUM(R{}:R{})".format(5, 5 + idx)
            sheet['R{}'.format(5 + idx)].border = thin_border
            sheet['R{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['S{}'.format(5 + idx)] = "=SUM(S{}:S{})".format(5, 5 + idx)
            sheet['S{}'.format(5 + idx)].border = thin_border
            sheet['S{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['T{}'.format(5 + idx)] = "=SUM(T{}:T{})".format(5, 5 + idx)
            sheet['T{}'.format(5 + idx)].border = thin_border
            sheet['T{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['U{}'.format(5 + idx)] = "=SUM(U{}:U{})".format(5, 5 + idx)
            sheet['U{}'.format(5 + idx)].border = thin_border
            sheet['U{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['V{}'.format(5 + idx)] = "=SUM(V{}:V{})".format(5, 5 + idx)
            sheet['V{}'.format(5 + idx)].border = thin_border
            sheet['V{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['W{}'.format(5 + idx)] = "=SUM(W{}:W{})".format(5, 5 + idx)
            sheet['W{}'.format(5 + idx)].border = thin_border
            sheet['W{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['X{}'.format(5 + idx)] = "=SUM(X{}:X{})".format(5, 5 + idx)
            sheet['X{}'.format(5 + idx)].border = thin_border
            sheet['X{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Y{}'.format(5 + idx)] = "=SUM(Y{}:Y{})".format(5, 5 + idx)
            sheet['Y{}'.format(5 + idx)].border = thin_border
            sheet['Y{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Z{}'.format(5 + idx)] = "=SUM(Z{}:Z{})".format(5, 5 + idx)
            sheet['Z{}'.format(5 + idx)].border = thin_border
            sheet['Z{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AA{}'.format(5 + idx)] = "=SUM(AA{}:AA{})".format(5, 5 + idx)
            sheet['AA{}'.format(5 + idx)].border = thin_border
            sheet['AA{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AB{}'.format(5 + idx)] = "=SUM(AB{}:AB{})".format(5, 5 + idx)
            sheet['AB{}'.format(5 + idx)].border = thin_border
            sheet['AB{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AC{}'.format(5 + idx)] = "=SUM(AC{}:AC{})".format(5, 5 + idx)
            sheet['AC{}'.format(5 + idx)].border = thin_border
            sheet['AC{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AD{}'.format(5 + idx)] = "=SUM(AD{}:AD{})".format(5, 5 + idx)
            sheet['AD{}'.format(5 + idx)].border = thin_border
            sheet['AD{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AE{}'.format(5 + idx)] = "=SUM(AE{}:AE{})".format(5, 5 + idx)
            sheet['AE{}'.format(5 + idx)].border = thin_border
            sheet['AE{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AF{}'.format(5 + idx)] = "=SUM(AF{}:AF{})".format(5, 5 + idx)
            sheet['AF{}'.format(5 + idx)].border = thin_border
            sheet['AF{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AG{}'.format(5 + idx)] = "=SUM(AG{}:AG{})".format(5, 5 + idx)
            sheet['AG{}'.format(5 + idx)].border = thin_border
            sheet['AG{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AH{}'.format(5 + idx)] = "=SUM(AH{}:AH{})".format(5, 5 + idx)
            sheet['AH{}'.format(5 + idx)].border = thin_border
            sheet['AH{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AI{}'.format(5 + idx)] = "=SUM(AI{}:AI{})".format(5, 5 + idx)
            sheet['AI{}'.format(5 + idx)].border = thin_border
            sheet['AI{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AJ{}'.format(5 + idx)] = "=SUM(AJ{}:AJ{})".format(5, 5 + idx)
            sheet['AJ{}'.format(5 + idx)].border = thin_border
            sheet['AJ{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AK{}'.format(5 + idx)] = "=SUM(AK{}:AK{})".format(5, 5 + idx)
            sheet['AK{}'.format(5 + idx)].border = thin_border
            sheet['AK{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AL{}'.format(5 + idx)] = "=SUM(AL{}:AL{})".format(5, 5 + idx)
            sheet['AL{}'.format(5 + idx)].border = thin_border
            sheet['AL{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['AM{}'.format(5 + idx)] = "=SUM(AM{}:AM{})".format(5, 5 + idx)
            sheet['AM{}'.format(5 + idx)].border = thin_border
            sheet['AM{}'.format(5 + idx)].number_format = '#,##0.00'

        wb.save(filename="{}/../reports/report_pajak_{}_{}.xlsx".format(path, d1.strftime(DATETIME_FORMAT),
                                                                        d2.strftime(DATETIME_FORMAT)))
        return {
            'name': 'Pajak',
            'type': 'ir.actions.act_url',
            'url': '/hr_pajak/download/{}/{}'.format(d1.strftime(DATETIME_FORMAT), d2.strftime(DATETIME_FORMAT)),
            'target': 'self',
        }

    date_from = fields.Date('Tanggal Awal')
    date_to = fields.Date('Tanggal Akhir')
    employee_ids = fields.Many2many('hr.employee')

class CodecPipeline(object):
    """Chains multiple codecs into a single encode/decode operation"""
    def __init__(self, *names, **kwds):
        self.default_errors = self._bind_kwds(**kwds)
        encoders = []
        decoders = []
        self.codecs = names
        for name in names:
            info = self._lookup_codec(name)
            encoders.append(info.encode)
            decoders.append(info.decode)
        self.encoders = encoders
        decoders.reverse()
        self.decoders = decoders

    def _bind_kwds(self, errors=None):
        if errors is None:
            errors = "strict"
        return errors

    def _lookup_codec(self, name):
        # Work around for http://bugs.python.org/issue15331 in 3.x
        try:
            return codecs.lookup(name)
        except LookupError:
            return codecs.lookup(name + "_codec")

    def __repr__(self):
        names = self.codecs
        errors = self.default_errors
        if not names:
            return "{}(errors={!r})".format(type(self).__name__, errors)
        return "{}({}, errors={!r})".format(type(self).__name__,
                                            ", ".join(map(repr, names)),
                                            errors)

    def encode(self, input, errors=None):
        """Apply all encoding operations in the pipeline"""
        if errors is None:
            errors = self.default_errors
        result = input
        for encode in self.encoders:
            result, __ = encode(result, errors)
        return result

    def decode(self, input, errors=None):
        """Apply all decoding operations in the pipeline"""
        if errors is None:
            errors = self.default_errors
        result = input
        for decode in self.decoders:
            result,__ = decode(result, errors)
        return result
class PediodePajak(models.Model):
    _name = 'hr.pajak.periode'

    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'juni', 'Juli', 'Agustus', 'September', 'Oktober',
              'November', 'Desember']

    def _compute_default(self):
        currDate = datetime.now()
        return 'Periode pajak {} {}'.format(self.months[currDate.month - 1], currDate.year)

    name = fields.Char('Periode Pajak', default=_compute_default)
    month = fields.Selection([
        (1, 'Januari'),
        (2, 'Februari'),
        (3, 'Maret'),
        (4, 'April'),
        (5, 'Mei'),
        (6, 'Juni'),
        (7, 'Juli'),
        (8, 'Agustus'),
        (9, 'September'),
        (10, 'Oktober'),
        (11, 'November'),
        (12, 'Desember'),
    ], default=datetime.now().month, string='Bulan')
    year = fields.Char('Tahun', default=datetime.now().year)
    pajak_ids = fields.One2many('hr.pajak', 'periode_id', 'Daftar Pajak')
    name_txt_file   = fields.Char('File Name', readonly=True)
    export_data     = fields.Binary("Export File")
    export_data_txt = fields.Binary("Export File")

    def _prepare_datas(self,name_file):
        dirpath = os.environ.get('HOME') or os.getcwd()
        abs_filepath = os.path.abspath(os.path.join(dirpath, str(name_file)))
        #abs_file = os.path.abspath(str(name_file))
        pph21_file = open(abs_filepath,"w")# w = write
        # create header
        pph21_file.write('Masa Pajak'+';'+'Tahun Pajak'+';'+'NPWP'+';'+'Nama'+';'+'Jumlah Bruto'+';'+'Jumlah PPh')
        for employee in self.pajak_ids :
            month = self.month
            year = self.year
            date_to = datetime(int(year), month, 28).strftime('%Y-%m-%d')
            date_one = datetime(int(year), month, 1).strftime('%Y-%m-%d')
            pysl = self.env['hr.payslip'].search(
                [('employee_id', '=', employee.employee_id.id), ('date_to', '<=', date_to),('date_to','>=', date_one)])
            #import pdb;pdb.set_trace()
            lembur = employee.lembur
            perdin = employee.perdin
            diklat = employee.diklat
            honorarium = employee.honorarium
            bonus = employee.bonus
            npwp = employee.employee_id.npwp or ""
            for payslip in pysl :
                pot_kehadiran = 0
                pot_absensi = 0
                gaji = 0
                jkk = 0
                jkm = 0
                jkes = 0
                tunj_pajak = 0
                pph21 = 0
                for detail in payslip.line_ids :
                    if detail.code == "PK" :
                        pot_kehadiran = detail.total
                    if detail.code == "PA" :
                        pot_absensi = detail.total
                gaji = payslip.single_salary + pot_kehadiran + pot_absensi
                jkk = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkk/100)
                jkm = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkm/100)
                jkes = payslip.tunj_bpjskes
                tunj_pajak = payslip.tunj_pajak
                bruto = gaji + jkk + jkm + jkes + tunj_pajak + lembur + perdin + diklat + honorarium + bonus
                pph21 = tunj_pajak
            pph21_file.write("\r\n")
            pph21_file.write(str(self.month)+';'+str(self.year)+';'+npwp+';'+employee.employee_id.name+';'+str(int(bruto))+';'+str(int(pph21)))
        pph21_file.close()
        return str(abs_filepath)

    def download_csv_pph21(self):
        self.env["ir.attachment"].search([('res_id','=',self.id)]).unlink()
        name_file = "pph21.csv"
        datas_file = self._prepare_datas(name_file)
        f_read =  open(datas_file, "r")# r = read
        file_data = f_read.read()
        #Pass your text file data using encoding.
        cp = CodecPipeline("utf8","base64")
        values = {
                    "name": "pph21",
                    "datas_fname": name_file,
                    "res_model": "hr.pajak.periode",
                    "res_id": self.id,
                    "type": "binary",
                    "public": True,
                    "datas": cp.encode(file_data),
                }
        #Using your data create as attachment
        attachment_id = self.env["ir.attachment"].sudo().create(values)
        self.export_data = cp.encode(file_data)
        #self.name = "%s%s" % ('report', '.txt')
        self.state = 'get'


    def download_excel_pph21(self):
        self.ensure_one()
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'id_ID')
        except:
            print('error setting locale')
        months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'juni', 'Juli', 'Agustus', 'September', 'Oktober',
              'November', 'Desember']
        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/report_pajak.xlsx".format(path)
        wb = load_workbook(filename)
        sheet = wb['Sheet1']
        idx = 0
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for pajak in self.pajak_ids:
            sheet.insert_rows(5 + idx)
            sheet['A{}'.format(5 + idx)] = idx + 1
            sheet['A{}'.format(5 + idx)].border = thin_border
            sheet['B{}'.format(5 + idx)] = pajak.employee_id.name
            sheet['B{}'.format(5 + idx)].border = thin_border
            sheet['C{}'.format(5 + idx)] = pajak.employee_id.no_npwp if pajak.employee_id.no_npwp else ''
            sheet['C{}'.format(5 + idx)].border = thin_border
            sheet['D{}'.format(5 + idx)] = pajak.employee_id.nik if pajak.employee_id.nik else ''
            sheet['D{}'.format(5 + idx)].border = thin_border
            sheet['E{}'.format(5 + idx)] = pajak.employee_id.status_pajak.kode if pajak.employee_id.status_pajak else ''
            sheet['E{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)] = pajak.gaji
            sheet['F{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['G{}'.format(5 + idx)] = pajak.tunj_pajak
            sheet['G{}'.format(5 + idx)].border = thin_border
            sheet['G{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['H{}'.format(5 + idx)] = pajak.tunj_lainya
            sheet['H{}'.format(5 + idx)].border = thin_border
            sheet['H{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['I{}'.format(5 + idx)] = 0
            sheet['I{}'.format(5 + idx)].border = thin_border
            sheet['I{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['J{}'.format(5 + idx)] = pajak.pre_asuransi
            sheet['J{}'.format(5 + idx)].border = thin_border
            sheet['J{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['K{}'.format(5 + idx)] = 0
            sheet['K{}'.format(5 + idx)].border = thin_border
            sheet['K{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['L{}'.format(5 + idx)] = 0
            sheet['L{}'.format(5 + idx)].border = thin_border
            sheet['L{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['M{}'.format(5 + idx)] = pajak.jum_bruto
            sheet['M{}'.format(5 + idx)].border = thin_border
            sheet['M{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['N{}'.format(5 + idx)] = pajak.biaya_jabatan
            sheet['N{}'.format(5 + idx)].border = thin_border
            sheet['N{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['O{}'.format(5 + idx)] = pajak.iuran
            sheet['O{}'.format(5 + idx)].border = thin_border
            sheet['O{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['P{}'.format(5 + idx)] = pajak.jum_pengurang
            sheet['P{}'.format(5 + idx)].border = thin_border
            sheet['P{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Q{}'.format(5 + idx)] = pajak.jum_neto
            sheet['Q{}'.format(5 + idx)].border = thin_border
            sheet['Q{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['R{}'.format(5 + idx)] = 0
            sheet['R{}'.format(5 + idx)].border = thin_border
            sheet['R{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['S{}'.format(5 + idx)] = pajak.jum_neto
            sheet['S{}'.format(5 + idx)].border = thin_border
            sheet['S{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['T{}'.format(5 + idx)] = pajak.ptkp
            sheet['T{}'.format(5 + idx)].border = thin_border
            sheet['T{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['U{}'.format(5 + idx)] = pajak.pkp_setahun
            sheet['U{}'.format(5 + idx)].border = thin_border
            sheet['U{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['V{}'.format(5 + idx)] = pajak.pph21
            sheet['V{}'.format(5 + idx)].border = thin_border
            sheet['V{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['W{}'.format(5 + idx)] = int(pajak.pph21_sebulan)
            sheet['W{}'.format(5 + idx)].border = thin_border
            sheet['W{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['X{}'.format(5 + idx)] = months[int(self.month)-1]
            sheet['X{}'.format(5 + idx)].border = thin_border
            sheet['X{}'.format(5 + idx)].number_format = '#,##0.00'
            idx += 1
        if idx > 0:
            sheet.insert_rows(5 + idx)
            sheet['A{}'.format(5 + idx)].border = thin_border
            sheet['B{}'.format(5 + idx)].border = thin_border
            sheet['C{}'.format(5 + idx)].border = thin_border
            sheet['D{}'.format(5 + idx)].border = thin_border
            sheet['E{}'.format(5 + idx)] = "TOTAL"
            sheet['E{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)] = "=SUM(F{}:F{})".format(5, 4 + idx)
            sheet['F{}'.format(5 + idx)].border = thin_border
            sheet['F{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['G{}'.format(5 + idx)] = "=SUM(G{}:G{})".format(5, 4 + idx)
            sheet['G{}'.format(5 + idx)].border = thin_border
            sheet['G{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['H{}'.format(5 + idx)] = "=SUM(H{}:H{})".format(5, 4 + idx)
            sheet['H{}'.format(5 + idx)].border = thin_border
            sheet['H{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['I{}'.format(5 + idx)] = "=SUM(I{}:I{})".format(5, 4 + idx)
            sheet['I{}'.format(5 + idx)].border = thin_border
            sheet['I{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['J{}'.format(5 + idx)] = "=SUM(J{}:J{})".format(5, 4 + idx)
            sheet['J{}'.format(5 + idx)].border = thin_border
            sheet['J{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['K{}'.format(5 + idx)] = "=SUM(K{}:K{})".format(5, 4 + idx)
            sheet['K{}'.format(5 + idx)].border = thin_border
            sheet['K{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['L{}'.format(5 + idx)] = "=SUM(L{}:L{})".format(5, 4 + idx)
            sheet['L{}'.format(5 + idx)].border = thin_border
            sheet['L{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['M{}'.format(5 + idx)] = "=SUM(M{}:M{})".format(5, 4 + idx)
            sheet['M{}'.format(5 + idx)].border = thin_border
            sheet['M{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['N{}'.format(5 + idx)] = "=SUM(N{}:N{})".format(5, 4 + idx)
            sheet['N{}'.format(5 + idx)].border = thin_border
            sheet['N{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['O{}'.format(5 + idx)] = "=SUM(O{}:O{})".format(5, 4 + idx)
            sheet['O{}'.format(5 + idx)].border = thin_border
            sheet['O{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['P{}'.format(5 + idx)] = "=SUM(P{}:P{})".format(5, 4 + idx)
            sheet['P{}'.format(5 + idx)].border = thin_border
            sheet['P{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['Q{}'.format(5 + idx)] = "=SUM(Q{}:Q{})".format(5, 4 + idx)
            sheet['Q{}'.format(5 + idx)].border = thin_border
            sheet['Q{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['R{}'.format(5 + idx)] = "=SUM(R{}:R{})".format(5, 4 + idx)
            sheet['R{}'.format(5 + idx)].border = thin_border
            sheet['R{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['S{}'.format(5 + idx)] = "=SUM(S{}:S{})".format(5, 4 + idx)
            sheet['S{}'.format(5 + idx)].border = thin_border
            sheet['S{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['T{}'.format(5 + idx)] = "=SUM(T{}:T{})".format(5, 4 + idx)
            sheet['T{}'.format(5 + idx)].border = thin_border
            sheet['T{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['U{}'.format(5 + idx)] = "=SUM(U{}:U{})".format(5, 4 + idx)
            sheet['U{}'.format(5 + idx)].border = thin_border
            sheet['U{}'.format(5 + idx)].number_format = '#,##0.00'
            sheet['V{}'.format(5 + idx)] = "=SUM(V{}:V{})".format(5, 4 + idx)
            sheet['V{}'.format(5 + idx)].border = thin_border
            sheet['V{}'.format(5 + idx)].number_format = '#,##0.00'
        currDate = datetime.now()
        wb.save(filename="{}/../reports/report_pajak_periode_{}_{}.xlsx".format(path, self.months[currDate.month - 1],
                                                                                currDate.year))
        return {
            'name': 'Pajak',
            'type': 'ir.actions.act_url',
            'url': '/hr_pajak/download/{}/{}'.format(self.months[currDate.month - 1], currDate.year),
            'target': 'self',
        }


class Pajak(models.Model):
    _name = 'hr.pajak'
    _description = 'Pajak PPH21'
    _rec_name = 'employee_id'

    @api.onchange('employee_id', 'year')
    def onchange_employee(self):
        month = self.periode_id.month
        year = self.periode_id.year
        date_to = datetime(int(year), month, 26).strftime('%Y-%m-%d')
        date_one = datetime(int(year), 1, 1).strftime('%Y-%m-%d')
        payslips = self.env['hr.payslip'].search(
            [('employee_id', '=', self.employee_id.id), ('date_to', '<=', date_to),('date_to','>=', date_one)])
        gaji = 0
        asuransi = 0
        potongan = 0
        salary = 0
        dplk = 0
        bpjstk = 0
        bpjskes = 0
        potongan_dplk = 0
        pot_bpjstk = 0
        pot_bpjskes = 0
        bpjstk_jkm = 0
        bpjstk_jkk = 0
        pot_kehadiran = 0
        pot_absensi = 0
        mth = 0
        mth_dari = 12
        self.dplk = 0
        self.bpjstk = 0
        self.bpjskes = 0
        self.pot_dplk = 0
        self.pot_bpjstk = 0
        self.pot_bpjskes = 0
        for payslip in payslips:
            pot_kehadiran = 0
            pot_absensi = 0
            salary = payslip.single_salary
            for detail in payslip.line_ids :
                if detail.code == "PK" :
                    pot_kehadiran = detail.total
                if detail.code == "PA" :
                    pot_absensi = detail.total
            #dplk = payslip.tunjangan_dplk
            bpjstk = payslip.tunj_bpjstk
            bpjskes = payslip.tunj_bpjskes
            potongan_dplk = payslip.potongan_dplk
            pot_bpjstk = payslip.pot_bpjstk
            #pot_bpjskes = payslip.pot_bpjskes
            gaji += salary + pot_kehadiran + pot_absensi 
            bpjstk_jkk = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkk/100)
            bpjstk_jkm = payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkm/100)
            asuransi += bpjstk_jkk + bpjstk_jkm + bpjskes
            #self.dplk = 0
            #self.bpjstk += bpjstk
            self.bpjstk_jkk += payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkk/100)
            self.bpjstk_jkm += payslip.single_salary * (payslip.employee_id.jenis_pegawai.jkm/100)
            self.bpjskes += bpjskes
            potongan += potongan_dplk + pot_bpjstk + pot_bpjskes
            self.pot_dplk += potongan_dplk
            self.pot_bpjstk += pot_bpjstk
            self.pot_bpjskes += pot_bpjskes
            months = datetime.strptime(payslip.date_to,"%Y-%m-%d").month
            if months < mth_dari :
                self.masa_dari = months
            if months > mth :
                self.masa_sampai = months 
            mth = months
            mth_dari = months
        for i in range(month, 12):
            gaji += salary
            asuransi += bpjstk_jkk + bpjstk_jkm + bpjskes
            #self.dplk += dplk
            #self.bpjstk += bpjstk
            self.bpjstk_jkk += bpjstk_jkk
            self.bpjstk_jkm += bpjstk_jkm
            self.bpjskes += bpjskes
            potongan += potongan_dplk + pot_bpjstk + pot_bpjskes
            self.pot_dplk += potongan_dplk
            self.pot_bpjstk += pot_bpjstk
            self.pot_bpjskes += pot_bpjskes

        DATETIME_FORMAT = "%Y-%m-%d"
        d1 = datetime(int(year), 1, 1).strftime(DATETIME_FORMAT)
        d2 = datetime(int(year), month, calendar.monthrange(int(year), month)[1]).strftime(DATETIME_FORMAT)
        # lembur
        overtimes = self.env['hr.overtime'].search(
            [('state', '=', 'validate_realisasi'), ('realisasi_date_from', '>=', d1),
             ('realisasi_date_from', '<=', d2), ('employee_id', '=', self.employee_id.id)])
        lembur = 0
        for overtime in overtimes:
            lembur += overtime.uang_lembur
        # perdin
        perdins = self.env['hr.perdin.employee'].search(
            [('state', '=', 'validate'), ('date_from', '>=', d1),
             ('date_to', '<=', d2), ('employee_id', '=', self.employee_id.id)])
        perdin = 0
        for obj_perdin in perdins:
            perdin += obj_perdin.uang_saku
        # pelatihan
        trainings = self.env['hr.training.participant'].search(
            [('state', '=', 'approve'), ('start_date', '>=', d1),
             ('end_date', '<=', d2), ('name', '=', self.employee_id.id)])
        diklat = 0
        for training in trainings:
            diklat += training.total_uang_pelatihan
        self.gaji = gaji
        #self.perdin = perdin
        #self.diklat = diklat
        #self.lembur = lembur
        self.perdin = 0
        self.diklat = 0
        self.lembur = 0
        self.tunj_lainya = 0
        #self.tunj_lainya = lembur + perdin + diklat
        self.status = self.employee_id.status_pajak.id
        self.pre_asuransi = asuransi
        self.iuran = potongan

    @api.onchange('status')
    def onchange_status(self):
        self.ptkp = self.status.nominal_tahun

    @api.onchange('lembur', 'perdin', 'diklat')
    def onchange_tunj_lainnya(self):
        self.tunj_lainya = self.lembur + self.perdin + self.diklat

    @api.onchange('dplk', 'bpjstk', 'bpjskes')
    def onchange_asuransi(self):
        self.pre_asuransi = self.bpjstk_jkk + self.bpjstk_jkm + self.bpjskes

    @api.onchange('pot_dplk', 'pot_bpjstk', 'pot_bpjskes')
    def onchange_iuran(self):
        self.iuran = self.pot_dplk + self.pot_bpjstk + self.pot_bpjskes

    @api.onchange('gaji', 'tunj_lainya', 'honorarium', 'pre_asuransi', 'bonus', 'iuran')
    def onchange_pennghasilan_all(self):
        self.jum_bruto = self.gaji + self.tunj_lainya + self.honorarium + self.pre_asuransi + self.bonus
        biaya_jabatan = 0.05 * self.jum_bruto
        self.biaya_jabatan = 6000000 if biaya_jabatan > 6000000 else biaya_jabatan
        self.jum_pengurang = self.iuran + self.biaya_jabatan

    @api.onchange('jum_pengurang')
    def onchange_jum_pengurang(self):
        self.jum_neto = self.jum_bruto - self.jum_pengurang
        self.jum_neto_pph21 = self.jum_neto + self.peng_neto_masa
        pkp_sthn = self.jum_neto_pph21 - self.ptkp 
        if pkp_sthn > 1000 : 
            self.pkp_setahun = str(pkp_sthn)[:len(str(pkp_sthn))-3]+"000"

    @api.onchange('penyesuaian')
    def onchange_penyesuaian(self):
        pph21_final = 0.0
        if self.periode_id.month == 12 :
                periods = self.env['hr.pajak.periode'].search(
                [('year', '=', self.periode_id.year), ('month', '<', 12)])
                for period in periods :
                    pajaks = self.env["hr.pajak"].search([("employee_id","=",self.employee_id.id)], limit=1)
                    for pajak in pajaks :
                        pph21_final += self.pph21_sebulan
                self.pph21_des = self.tunj_pajak - pph21_final + self.penyesuaian

    @api.onchange('pkp_setahun')
    def onchange_pkp(self):
        _logger.info('onchange pkp')
        _logger.info("pkp setahun : {}".format(self.pkp_setahun))
        rumus_pkp = self.env['hr.pkp'].search(
            [('nominal_mix', '<=', self.pkp_setahun), ('nominal_max', '>=', self.pkp_setahun)], limit=1)
        pph21_final = 0.0
        if rumus_pkp:
            _logger.info("nominal_mix : {}".format(rumus_pkp.nominal_mix))
            _logger.info("pajak : {}".format(rumus_pkp.pajak))
            _logger.info("penambah : {}".format(rumus_pkp.penambah))
            pajak_real = 0
            for pkp in self.env['hr.pkp'].search([]) :
                pajak = 0
                if self.pkp_setahun > pkp.nominal_max :
                    total_ptkp = pkp.nominal_max - pkp.nominal_mix
                    pajak = total_ptkp * pkp.pajak
                elif self.pkp_setahun <= pkp.nominal_max and self.pkp_setahun >= pkp.nominal_mix :
                    pjk = self.pkp_setahun - pkp.nominal_mix
                    pajak = pjk * pkp.pajak
                pajak_real = pajak_real + pajak
            self.tunj_pajak = pajak_real    
            _logger.info("tunj_pajak : {}".format(self.tunj_pajak))
            cek = True
            i = 0
            while cek == True and i < 100 :
                #if month != "12" :       
                #    self.jum_bruto = self.gaji + self.tunj_lainya + self.honorarium + self.pre_asuransi + self.bonus + self.tunj_pajak
                #else :
                #    tunj = self.env['pajak'].search([('month','!=','12'),('year','=',self.year)])
                #    for penambah in tunj :

                self.biaya_jabatan = (self.jum_bruto * 5)/100
                self.jum_pengurang = self.biaya_jabatan + self.iuran
                self.jum_neto = self.jum_bruto - self.jum_pengurang
                self.jum_neto_pph21 = self.jum_neto + self.peng_neto_masa
                #self.pkp_setahun = self.jum_neto_pph21 - self.ptkp
                pkp_sthn = self.jum_neto_pph21 - self.ptkp 
                if pkp_sthn > 1000 : 
                    self.pkp_setahun = str(pkp_sthn)[:len(str(pkp_sthn))-3]+"000"
                pajak_real = 0
                for pkp in self.env['hr.pkp'].search([]) :
                    pajak = 0
                    if self.pkp_setahun > pkp.nominal_max :
                        total_ptkp = pkp.nominal_max - pkp.nominal_mix
                        pajak = total_ptkp * pkp.pajak
                    elif self.pkp_setahun <= pkp.nominal_max and self.pkp_setahun >= pkp.nominal_mix :
                        pjk = self.pkp_setahun - pkp.nominal_mix
                        pajak = pjk * pkp.pajak
                    pajak_real = pajak_real + pajak
                self.pph21 = pajak_real
                #self.pph21 = (self.pkp_setahun - rumus_pkp.nominal_mix) * rumus_pkp.pajak
                if self.pph21 == self.tunj_pajak :
                    cek = False
                if self.pph21 != self.tunj_pajak :
                    self.tunj_pajak = self.pph21
                i += 1
            self.pph21_sebulan = self.pph21/12
            _logger.info("jum_bruto : {}".format(self.jum_bruto))
            
            #if self.periode_id.month == 12 :
            #    periods = self.env['hr.pajak.periode'].search(
            #    [('year', '=', self.periode_id.year), ('month', '<', 12)])
            #    for period in periods :
            #        pajaks = self.env["hr.pajak"].search([("employee_id","=",self.employee_id.id)], limit=1)
            #        for pajak in pajaks :
            #            pph21_final += self.pph21_sebulan
            #    self.pph21_des = self.tunj_pajak - pph21_final + self.penyesuaian
            ### Pengitungan pph21 Penghasilan Rutin ###
            pajak_real_rutin = 0
            pkp_setahun = 0
            #for pkp in self.env['hr.pkp'].search([]) :
            #    pajak = 0
            #    if (self.pkp_setahun - self.tunj_pajak - self.tunj_lainya - self.honorarium - self.bonus) > pkp.nominal_max :
            #        total_ptkp = pkp.nominal_max - pkp.nominal_mix
            #        pajak = total_ptkp * pkp.pajak
            #    elif self.pkp_setahun <= pkp.nominal_max and self.pkp_setahun >= pkp.nominal_mix :
            #        pjk = (self.pkp_setahun - self.tunj_pajak - self.tunj_lainya - self.honorarium - self.bonus) - pkp.nominal_mix
            #        pajak = pjk * pkp.pajak
            #    pajak_real_rutin = pajak_real_rutin + pajak
            cek_rutin = True
            i_rutin = 0
            pajak_sebelum = 0
            while cek_rutin == True and i_rutin < 100 :       
                jum_bruto = self.gaji + self.pre_asuransi + pajak_real_rutin
                biaya_jabatan = (jum_bruto * 5)/100
                jum_pengurang = biaya_jabatan + self.iuran
                jum_neto = jum_bruto -jum_pengurang
                jum_neto_pph21 = jum_neto + self.peng_neto_masa
                #self.pkp_setahun = self.jum_neto_pph21 - self.ptkp
                pkp_sthn = jum_neto_pph21 - self.ptkp 
                if pkp_sthn > 1000 : 
                    pkp_setahun = int(str(int(pkp_sthn))[:len(str(int(pkp_sthn)))-3]+"000")
                pajak_real_rutin = 0
                #import pdb;pdb.set_trace()
                for pkp in self.env['hr.pkp'].search([]) :
                    pajak = 0
                    if pkp_setahun > pkp.nominal_max :
                        total_ptkp = pkp.nominal_max - pkp.nominal_mix
                        pajak = total_ptkp * pkp.pajak
                    elif pkp_setahun <= pkp.nominal_max and pkp_setahun >= pkp.nominal_mix :
                        pjk = pkp_setahun - pkp.nominal_mix
                        pajak = pjk * pkp.pajak
                    pajak_real_rutin = pajak_real_rutin + pajak
                pph21 = pajak_real_rutin
                #self.pph21 = (self.pkp_setahun - rumus_pkp.nominal_mix) * rumus_pkp.pajak
                if pph21 == pajak_sebelum :
                    cek_rutin = False
                if pph21 != pajak_sebelum :
                    pajak_sebelum = pph21
                i_rutin += 1
            if self.periode_id.month == 12 :
                periods = self.env['hr.pajak.periode'].search(
                [('year', '=', self.periode_id.year), ('month', '<', 12)])
                for period in periods :
                    pajaks = self.env["hr.pajak"].search([("employee_id","=",self.employee_id.id)], limit=1)
                    for pajak in pajaks :
                        pph21_final += self.pph21_sebulan
                self.pph21_des = self.pph21 - pph21_final + self.penyesuaian
            if self.periode_id.month != 12 :
                self.pph21_sebulan = pph21/12 + (self.tunj_pajak - pph21)

    employee_id = fields.Many2one(comodel_name="hr.employee", string="Nama")
    status = fields.Many2one(comodel_name="hr.ptkp", string="Status Pajak")
    year = fields.Char("Tahun", default=datetime.now().year)
    gaji = fields.Integer("Gaji", default=0)
    tunj_pajak = fields.Integer("Tunjangan Pajak", default=0)
    perdin = fields.Integer('Perdin', default=1)
    lembur = fields.Integer('Lembur', default=1)
    diklat = fields.Integer('Diklat', default=1)
    tunj_lainya = fields.Integer("Tunjangan Lainya", default=0)
    honorarium = fields.Integer("Honorarium", default=0)
    dplk = fields.Integer("Dplk", default=0)
    bpjstk = fields.Integer("BPJS TK", default=0)
    bpjstk_jkk = fields.Integer("BPJS TK JKK", default=0)
    bpjstk_jkm = fields.Integer("BPJS TK JKM", default=0)
    bpjskes = fields.Integer("BPJS Kes", default=0)
    pre_asuransi = fields.Integer("Premi Asuransi", default=0)
    bonus = fields.Integer("Tantiem,Bonus,THR Dll", default=0)
    jum_bruto = fields.Integer("Penghasilan Bruto", default=0)
    pot_dplk = fields.Integer("Pot. Dplk", default=0)
    pot_bpjstk = fields.Integer("Pot. BPJS TK", default=0)
    pot_bpjskes = fields.Integer("Pot. BPJS Kes", default=0)
    iuran = fields.Integer('Iuran', default=0)
    biaya_jabatan = fields.Integer("Biaya Jabatan", default=0)
    jum_pengurang = fields.Integer("Jumlah Pengurang", default=0)
    jum_neto = fields.Integer("Penghasilan Neto", default=0)
    peng_neto_masa = fields.Integer("Penghasilan Neto Sebelumnya", default=0)
    jum_neto_pph21 = fields.Integer("Penghasilan Neto PPH21", default=0)
    ptkp = fields.Integer("PTKP", default=0)
    pkp_setahun = fields.Integer("PKP Setahun", default=0)
    pph21 = fields.Integer("PPH21", default=0)
    pph21_sebulan = fields.Integer("PPH21 Sebulan", default=0)
    pph21_des = fields.Integer("PPH21 Des", default=0)
    penyesuaian = fields.Integer("Penyesuaian")
    periode_id = fields.Many2one('hr.pajak.periode', string='Periode Pajak')
    masa_dari = fields.Integer("Masa Kerja Dari")
    masa_sampai = fields.Integer('Masa Kerja Samapai')


class ResourceMixin1(models.AbstractModel):
    _inherit = "resource.mixin"
    _description = 'Resource Mixin'

    def get_work_days_data(self, from_datetime, to_datetime, calendar=None):
        days_count = 0.0
        total_work_time = timedelta()
        calendar = calendar or self.resource_calendar_id
        calendar = calendar.with_context(no_tz_convert=self.env.context.get('no_tz_convert', False))
        for day_intervals in calendar._iter_work_intervals(
                from_datetime, to_datetime, self.resource_id.id,
                compute_leaves=True):
            theoric_hours = self.get_day_work_hours_count(day_intervals[0][0].date(), calendar=calendar)
            work_time = sum((interval[1] - interval[0] for interval in day_intervals), timedelta())
            total_work_time += work_time
            if theoric_hours:
                days_count += float_utils.round((work_time.total_seconds() / 3600 / theoric_hours) * 4) / 4
        return {
            'days': days_count,
            'hours': total_work_time.total_seconds() / 3600,
        }

    def iter_leaves(self, from_datetime, to_datetime, calendar=None):
        calendar = calendar or self.resource_calendar_id
        return calendar._iter_leave_intervals(from_datetime, to_datetime, self.resource_id.id)

    def _iter_leave_intervals(self, start_dt, end_dt, resource_id):
        """ Lists the current resource's leave intervals between the two provided
        datetimes (inclusive) expressed in UTC. """
        if not end_dt:
            end_dt = datetime.combine(start_dt.date(), datetime.time.max)

        start_dt = to_naive_user_tz(start_dt, self.env.user)
        end_dt = to_naive_user_tz(end_dt, self.env.user)

        for day in rrule.rrule(rrule.DAILY,
                               dtstart=start_dt,
                               until=end_dt.replace(hour=23, minute=59, second=59, microsecond=999999),
                               byweekday=self._get_weekdays()):
            start_time = datetime.time.min
            if day.date() == start_dt.date():
                start_time = start_dt.time()
            end_time = datetime.time.max
            if day.date() == end_dt.date() and end_dt.time() != datetime.time():
                end_time = end_dt.time()

            intervals = self._get_day_leave_intervals(
                day.date(),
                start_time,
                end_time,
                resource_id)

            if intervals:
                yield intervals
