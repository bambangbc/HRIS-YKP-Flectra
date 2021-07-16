# -*- coding: utf-8 -*-
import datetime
import os

from openpyxl import load_workbook

from flectra import models, fields, api, exceptions


class work_goal(models.Model):
    _name = 'hr.work.goal'
    _inherit = ['mail.thread', 'resource.mixin']

    def _get_default_employee(self):
        employee = self.env['hr.employee'].search([('user_id', '=', self.env.user.id)])
        return employee

    @api.depends('user_id')
    def _current_user_is_employee(self):
        for row in self:
            row.current_user_is_employee = self.env.user.id == row.user_id.id

    @api.depends('reviewer_user_id')
    def _current_user_is_reviewer(self):
        for row in self:
            row.current_user_is_reviewer = self.env.user.id == row.reviewer_user_id.id

    @api.depends('reviewer2_user_id')
    def _current_user_is_reviewer2(self):
        for row in self:
            row.current_user_is_reviewer2 = self.env.user.id == row.reviewer2_user_id.id

    def _total_nilai(self):
        for row in self:
            total_value = 0
            for financial in row.work_perspective_financial_ids:
                total_value += financial.total_nilai
            for customer in row.work_perspective_customer_ids:
                total_value += customer.total_nilai
            for bp in row.work_perspective_bp_ids:
                total_value += bp.total_nilai
            for people in row.work_perspective_people_ids:
                total_value += people.total_nilai
            row.total_value = total_value

    def _total_nilai_q1(self):
        for row in self:
            total_value = 0
            for financial in row.work_perspective_financial_q1_ids:
                total_value += financial.nilai_target_q1
            for customer in row.work_perspective_customer_q1_ids:
                total_value += customer.nilai_target_q1
            for bp in row.work_perspective_bp_q1_ids:
                total_value += bp.nilai_target_q1
            for people in row.work_perspective_people_q1_ids:
                total_value += people.nilai_target_q1
            row.total_value_q1 = total_value

    def _total_nilai_q2(self):
        for row in self:
            total_value = 0
            for financial in row.work_perspective_financial_q2_ids:
                total_value += financial.nilai_target_q2
            for customer in row.work_perspective_customer_q2_ids:
                total_value += customer.nilai_target_q2
            for bp in row.work_perspective_bp_q2_ids:
                total_value += bp.nilai_target_q2
            for people in row.work_perspective_people_q2_ids:
                total_value += people.nilai_target_q2
            row.total_value_q2 = total_value

    def _total_nilai_q3(self):
        for row in self:
            total_value = 0
            for financial in row.work_perspective_financial_q3_ids:
                total_value += financial.nilai_target_q3
            for customer in row.work_perspective_customer_q3_ids:
                total_value += customer.nilai_target_q3
            for bp in row.work_perspective_bp_q3_ids:
                total_value += bp.nilai_target_q3
            for people in row.work_perspective_people_q3_ids:
                total_value += people.nilai_target_q3
            row.total_value_q3 = total_value

    def _total_nilai_q4(self):
        for row in self:
            total_value = 0
            for financial in row.work_perspective_financial_q4_ids:
                total_value += financial.nilai_target_q4
            for customer in row.work_perspective_customer_q4_ids:
                total_value += customer.nilai_target_q4
            for bp in row.work_perspective_bp_q4_ids:
                total_value += bp.nilai_target_q4
            for people in row.work_perspective_people_q4_ids:
                total_value += people.nilai_target_q4
            row.total_value_q4 = total_value

    @api.onchange('reviewer_id', 'reviewer2_id')
    def on_change_reviewer(self):
        if self.reviewer_id:
            if self.reviewer_id.id == self.reviewer2_id.id:
                raise exceptions.ValidationError('Penilai 1 dan Penilai Pemutus tidak boleh orang yang sama')
        if self.reviewer2_id:
            if self.reviewer_id.id == self.reviewer2_id.id:
                raise exceptions.ValidationError('Penilai 1 dan Penilai Pemutus tidak boleh orang yang sama')

    def _compute_total_bobot(self):
        for row in self:
            total_bobot_finansial = self._total_bobot_sasaran(row.work_perspective_financial_ids, 'financial')
            total_bobot_customer = self._total_bobot_sasaran(row.work_perspective_customer_ids, 'customer')
            total_bobot_bp = self._total_bobot_sasaran(row.work_perspective_bp_ids, 'bp')
            total_bobot_people = self._total_bobot_sasaran(row.work_perspective_people_ids, 'people')
            row.total_bobot = total_bobot_finansial + total_bobot_customer + total_bobot_bp + total_bobot_people

    current_user_is_reviewer = fields.Boolean(compute=_current_user_is_reviewer)
    current_user_is_reviewer2 = fields.Boolean(compute=_current_user_is_reviewer2)
    current_user_is_employee = fields.Boolean(compute=_current_user_is_employee)
    name = fields.Many2one('hr.employee', string='Nama Pegawai', default=_get_default_employee)
    employee_name = fields.Char(related='name.name')
    user_id = fields.Many2one(related='name.user_id')
    # nik = fields.Char(related='name.nik', string='NIK')
    job_id = fields.Many2one(related='name.job_id', string='Jabatan')
    department_id = fields.Many2one(related='name.department_id', string='Department')
    goal_year = fields.Char('Tahun Sasaran', default=datetime.datetime.now().year)
    approval_date = fields.Date('Tanggal Penetapan', default=datetime.datetime.now())
    reviewer_id = fields.Many2one('hr.employee', string='Penilai 1', domain=[('user_id', '!=', False)])
    reviewer_name = fields.Char(related='reviewer_id.name')
    reviewer_job_id = fields.Many2one(related='reviewer_id.job_id', string='Jabatan')
    reviewer_penetapan_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer_user_id = fields.Many2one(related='reviewer_id.user_id')
    reviewer2_id = fields.Many2one('hr.employee', string='Penilai Pemutus', domain=[('user_id', '!=', False)])
    reviewer2_name = fields.Char(related='reviewer2_id.name')
    reviewer2_job_id = fields.Many2one(related='reviewer2_id.job_id', string='Jabatan')
    reviewer2_penetapan_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer2_user_id = fields.Many2one(related='reviewer2_id.user_id')
    total_bobot = fields.Float('Total Bobot', compute=_compute_total_bobot)
    note = fields.Html('Catatan')
    state = fields.Selection([
        ('draft', 'Penyusunan'),
        ('sent', 'Dikirm'),
        ('approve', 'Penetapan 1'),
        ('approve_pemutus', 'Pengisian Sasaran'),
        ('sent_penilaian', 'Dikirm'),
        ('assessment', 'Penilaian 1'),
        ('assessment_pemutus', 'Penilaian Pemutus'),
    ], default='draft')
    total_value = fields.Float('Total Nilai', compute=_total_nilai)
    total_value_q1 = fields.Float('Nilai Kuartal 1', compute=_total_nilai_q1)
    total_value_q2 = fields.Float('Nilai Kuartal 2', compute=_total_nilai_q2)
    total_value_q3 = fields.Float('Nilai Kuartal 3', compute=_total_nilai_q3)
    total_value_q4 = fields.Float('Nilai Kuartal 4', compute=_total_nilai_q4)
    self_signature = fields.Binary('Tanda Tangan')
    reviewer_signature = fields.Binary('Tanda Tangan')
    reviewer2_signature = fields.Binary('Tanda Tangan')
    self_signature_assesment = fields.Binary('Tanda Tangan')
    reviewer_signature_assesment = fields.Binary('Tanda Tangan')
    reviewer2_signature_assesment = fields.Binary('Tanda Tangan')
    self_assesment_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer_assesment_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer2_assesment_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    work_perspective_financial_ids = fields.One2many('hr.work.goal.perspective.financial', 'work_goal_id')
    work_perspective_financial_q1_ids = fields.One2many('hr.work.goal.perspective.financial', 'work_goal_id')
    work_perspective_financial_q2_ids = fields.One2many('hr.work.goal.perspective.financial', 'work_goal_id')
    work_perspective_financial_q3_ids = fields.One2many('hr.work.goal.perspective.financial', 'work_goal_id')
    work_perspective_financial_q4_ids = fields.One2many('hr.work.goal.perspective.financial', 'work_goal_id')
    work_perspective_customer_ids = fields.One2many('hr.work.goal.perspective.customer', 'work_goal_id')
    work_perspective_customer_q1_ids = fields.One2many('hr.work.goal.perspective.customer', 'work_goal_id')
    work_perspective_customer_q2_ids = fields.One2many('hr.work.goal.perspective.customer', 'work_goal_id')
    work_perspective_customer_q3_ids = fields.One2many('hr.work.goal.perspective.customer', 'work_goal_id')
    work_perspective_customer_q4_ids = fields.One2many('hr.work.goal.perspective.customer', 'work_goal_id')
    work_perspective_bp_ids = fields.One2many('hr.work.goal.perspective.bp', 'work_goal_id')
    work_perspective_bp_q1_ids = fields.One2many('hr.work.goal.perspective.bp', 'work_goal_id')
    work_perspective_bp_q2_ids = fields.One2many('hr.work.goal.perspective.bp', 'work_goal_id')
    work_perspective_bp_q3_ids = fields.One2many('hr.work.goal.perspective.bp', 'work_goal_id')
    work_perspective_bp_q4_ids = fields.One2many('hr.work.goal.perspective.bp', 'work_goal_id')
    work_perspective_people_ids = fields.One2many('hr.work.goal.perspective.people', 'work_goal_id')
    work_perspective_people_q1_ids = fields.One2many('hr.work.goal.perspective.people', 'work_goal_id')
    work_perspective_people_q2_ids = fields.One2many('hr.work.goal.perspective.people', 'work_goal_id')
    work_perspective_people_q3_ids = fields.One2many('hr.work.goal.perspective.people', 'work_goal_id')
    work_perspective_people_q4_ids = fields.One2many('hr.work.goal.perspective.people', 'work_goal_id')

    @api.multi
    def action_sent(self):
        self.write({'state': 'sent'})

    @api.multi
    def action_sent_penilaian(self):
        self.write({'state': 'sent_penilaian'})

    @api.multi
    def action_approve(self):
        self.write({'state': 'approve'})

    @api.multi
    def action_approve_pemutus(self):
        self.write({'state': 'approve_pemutus'})

    @api.multi
    def action_assesment(self):
        self.write({'state': 'assessment'})

    @api.multi
    def action_assesment_pemutus(self):
        self.write({'state': 'assessment_pemutus'})

    def _total_bobot_sasaran(self, data, type):
        model = 'hr.work.goal.perspective.sasaran.financial'
        if type == 'financial':
            model = 'hr.work.goal.perspective.sasaran.financial'
        if type == 'bp':
            model = 'hr.work.goal.perspective.sasaran.bp'
        if type == 'customer':
            model = 'hr.work.goal.perspective.sasaran.customer'
        if type == 'people':
            model = 'hr.work.goal.perspective.sasaran.people'
        sasaran = self.env[model].search([])
        selected_sasaran = {}
        for s in sasaran:
            for d in data:
                if isinstance(d, list):
                    if d[2].get('sasaran_id') == s.id and s.id not in selected_sasaran:
                        selected_sasaran[s.id] = s
                else:
                    if d.sasaran_id.id == s.id and s.id not in selected_sasaran:
                        selected_sasaran[s.id] = s
        total = 0
        for sel in selected_sasaran.values():
            if isinstance(sel, list) and len(sel) > 1:
                total += sel[1].bobot
            else:
                total += sel.bobot
        return total

    @api.model
    def create(self, vals):
        sasaran_finansial = vals.get('work_perspective_financial_ids')
        if sasaran_finansial is None:
            raise exceptions.ValidationError('Silahkan masukkan sasaran kerja financial')
        sasaran_customer = vals.get('work_perspective_customer_ids')
        if sasaran_customer is None:
            raise exceptions.ValidationError('Silahkan masukkan sasaran kerja customer')
        sasaran_bp = vals.get('work_perspective_bp_ids')
        if sasaran_bp is None:
            raise exceptions.ValidationError('Silahkan masukkan sasaran kerja business process')
        sasaran_people = vals.get('work_perspective_people_ids')
        if sasaran_people is None:
            raise exceptions.ValidationError('Silahkan masukkan sasaran kerja people development')

        total_bobot_finansial = self._total_bobot_sasaran(sasaran_finansial, 'financial')
        if total_bobot_finansial > 65:
            raise exceptions.ValidationError('total bobot sasaran finansial ({}%) > 65%'.format(total_bobot_finansial))

        total_bobot_customer = self._total_bobot_sasaran(sasaran_customer, 'customer')
        if total_bobot_customer > 65:
            raise exceptions.ValidationError('total bobot sasaran customer ({}%) > 65%'.format(total_bobot_customer))

        total_bobot_bp = self._total_bobot_sasaran(sasaran_bp, 'bp')
        if total_bobot_bp > 65:
            raise exceptions.ValidationError('total bobot sasaran business process ({}%) > 65%'.format(total_bobot_bp))

        total_bobot_people = self._total_bobot_sasaran(sasaran_people, 'people')
        if total_bobot_people > 65:
            raise exceptions.ValidationError(
                'total bobot sasaran people development ({}%) > 65%'.format(total_bobot_people))

        total_bobot_sasaran = total_bobot_finansial + total_bobot_customer + total_bobot_bp + total_bobot_people
        if total_bobot_sasaran > 100:
            raise exceptions.ValidationError('total bobot sasaran ({}%)  > 100%'.format(total_bobot_sasaran))
        if total_bobot_sasaran < 100:
            raise exceptions.ValidationError('total bobot sasaran ({}%)  < 100%'.format(total_bobot_sasaran))

        rec = super(work_goal, self).create(vals)
        return rec

    @api.multi
    def write(self, vals):
        count = self.env['hr.work.goal'].search_count(
            [('name', '=', vals.get('name')), ('goal_year', '=', vals.get('goal_year'))])
        if count > 0:
            raise exceptions.ValidationError(
                'Data Sasaran Kerja untuk tahun {} telah ada'.format(vals.get('goal_year')))
        rec = super(work_goal, self).write(vals)
        return rec

    def export_excel(self):
        filenames = []
        for row in self:
            filenames.append(self.generate_excel(row))

    def generate_excel(self, data):
        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/template_work_goal.xlsx".format(path)
        wb = load_workbook(filename)
        sheet = wb['sheet']
        sheet['B5'] = 'Nama Pegawai       : {}'.format(data.employee_name)
        sheet['B6'] = 'NIK.                            : {}'.format(data.name.nik)
        sheet['B7'] = 'Jabatan                    : {}'.format(data.job_id.name)
        sheet['B8'] = 'Unit Kerja                : {}'.format(data.department_id.name)
        sheet['F11'] = 'tahun : {}'.format(data.goal_year)
        sheet['T5'] = '{}'.format(data.reviewer_name)
        sheet['T6'] = '{}'.format(data.reviewer_job_id.name)
        sheet['T8'] = '{}'.format(data.reviewer2_name)
        sheet['T9'] = '{}'.format(data.reviewer2_job_id.name)
        row = 33
        sheet.insert_rows(row,
                          len(data.work_perspective_financial_ids) + len(data.work_perspective_customer_ids) + len(
                              data.work_perspective_bp_ids) + len(data.work_perspective_people_ids))
        sheet.merge_cells('B{}:E{}'.format(row, row))
        sheet['B{}'.format(row)] = 'Finansial'
        for finansial in data.work_perspective_financial_ids:
            sheet.merge_cells('F{}:M{}'.format(row, row))
            sheet['F{}'.format(row)] = finansial.name
            sheet.merge_cells('N{}:O{}'.format(row, row))
            sheet['N{}'.format(row)] = finansial.bobot
            sheet.merge_cells('P{}:Q{}'.format(row, row))
            sheet['P{}'.format(row)] = finansial.jenis
            sheet.merge_cells('R{}:S{}'.format(row, row))
            sheet['R{}'.format(row)] = finansial.target_semester1
            sheet.merge_cells('T{}:U{}'.format(row, row))
            sheet['T{}'.format(row)] = finansial.realisasi
            sheet.merge_cells('V{}:Z{}'.format(row, row))
            sheet['V{}'.format(row)] = 'Terlampir'
            sheet.merge_cells('AA{}:AB{}'.format(row, row))
            sheet['AA{}'.format(row)] = '{}%'.format(finansial.percent_realisasi)
            sheet.merge_cells('AC{}:AD{}'.format(row, row))
            sheet['AC{}'.format(row)] = finansial.nilai
            sheet.merge_cells('AE{}:AF{}'.format(row, row))
            sheet['AE{}'.format(row)] = finansial.nilai_target
            row += 1
        sheet.merge_cells('B{}:E{}'.format(row, row))
        sheet['B{}'.format(row)] = 'Customer'
        for customer in data.work_perspective_customer_ids:
            sheet.merge_cells('F{}:M{}'.format(row, row))
            sheet['F{}'.format(row)] = customer.name
            sheet.merge_cells('N{}:O{}'.format(row, row))
            sheet['N{}'.format(row)] = customer.bobot
            sheet.merge_cells('P{}:Q{}'.format(row, row))
            sheet['P{}'.format(row)] = customer.jenis
            sheet.merge_cells('R{}:S{}'.format(row, row))
            sheet['R{}'.format(row)] = customer.target_semester1
            sheet.merge_cells('T{}:U{}'.format(row, row))
            sheet['T{}'.format(row)] = customer.realisasi
            sheet.merge_cells('V{}:Z{}'.format(row, row))
            sheet['V{}'.format(row)] = 'Terlampir'
            sheet.merge_cells('AA{}:AB{}'.format(row, row))
            sheet['AA{}'.format(row)] = '{}%'.format(customer.percent_realisasi)
            sheet.merge_cells('AC{}:AD{}'.format(row, row))
            sheet['AC{}'.format(row)] = customer.nilai
            sheet.merge_cells('AE{}:AF{}'.format(row, row))
            sheet['AE{}'.format(row)] = customer.nilai_target
            row += 1
        sheet.merge_cells('B{}:E{}'.format(row, row))
        sheet['B{}'.format(row)] = 'Internal Business Process'
        for bp in data.work_perspective_bp_ids:
            sheet.merge_cells('F{}:M{}'.format(row, row))
            sheet['F{}'.format(row)] = bp.name
            sheet.merge_cells('N{}:O{}'.format(row, row))
            sheet['N{}'.format(row)] = bp.bobot
            sheet.merge_cells('P{}:Q{}'.format(row, row))
            sheet['P{}'.format(row)] = bp.jenis
            sheet.merge_cells('R{}:S{}'.format(row, row))
            sheet['R{}'.format(row)] = bp.target_semester1
            sheet.merge_cells('T{}:U{}'.format(row, row))
            sheet['T{}'.format(row)] = bp.realisasi
            sheet.merge_cells('V{}:Z{}'.format(row, row))
            sheet['V{}'.format(row)] = 'Terlampir'
            sheet.merge_cells('AA{}:AB{}'.format(row, row))
            sheet['AA{}'.format(row)] = '{}%'.format(bp.percent_realisasi)
            sheet.merge_cells('AC{}:AD{}'.format(row, row))
            sheet['AC{}'.format(row)] = bp.nilai
            sheet.merge_cells('AE{}:AF{}'.format(row, row))
            sheet['AE{}'.format(row)] = bp.nilai_target
            row += 1

        sheet.merge_cells('B{}:E{}'.format(row, row))
        sheet['B{}'.format(row)] = 'People Development'
        for people in data.work_perspective_people_ids:
            sheet.merge_cells('F{}:M{}'.format(row, row))
            sheet['F{}'.format(row)] = people.name
            sheet.merge_cells('N{}:O{}'.format(row, row))
            sheet['N{}'.format(row)] = people.bobot
            sheet.merge_cells('P{}:Q{}'.format(row, row))
            sheet['P{}'.format(row)] = people.jenis
            sheet.merge_cells('R{}:S{}'.format(row, row))
            sheet['R{}'.format(row)] = people.target_semester1
            sheet.merge_cells('T{}:U{}'.format(row, row))
            sheet['T{}'.format(row)] = people.realisasi
            sheet.merge_cells('V{}:Z{}'.format(row, row))
            sheet['V{}'.format(row)] = 'Terlampir'
            sheet.merge_cells('AA{}:AB{}'.format(row, row))
            sheet['AA{}'.format(row)] = '{}%'.format(people.percent_realisasi)
            sheet.merge_cells('AC{}:AD{}'.format(row, row))
            sheet['AC{}'.format(row)] = people.nilai
            sheet.merge_cells('AE{}:AF{}'.format(row, row))
            sheet['AE{}'.format(row)] = people.nilai_target
            row += 1

        # sheet['N{}'.format(row)] = '=SUM(N34:O{})'.format(row)
        # sheet['AE34:AF{}'.format(row)] = '=SUM(AE34:AF{}'.format(row)
        filename = 'pengisian_work_goal_{}_{}.xlsx'.format(data.employee_name.lower().replace(' ', '_'), data.goal_year)
        wb.save(filename="{}/../reports/{}".format(path, filename))
        return filename


class work_goal_perspective_financial_sasaran(models.Model):
    _name = 'hr.work.goal.perspective.sasaran.financial'

    name = fields.Char('Sasaran')
    bobot = fields.Float('Bobot')


class work_goal_perspective_customer_sasaran(models.Model):
    _name = 'hr.work.goal.perspective.sasaran.customer'

    name = fields.Char('Sasaran')
    bobot = fields.Float('Bobot')


class work_goal_perspective_bp_sasaran(models.Model):
    _name = 'hr.work.goal.perspective.sasaran.bp'

    name = fields.Char('Sasaran')
    bobot = fields.Float('Bobot')


class work_goal_perspective_bp_people(models.Model):
    _name = 'hr.work.goal.perspective.sasaran.people'

    name = fields.Char('Sasaran')
    bobot = fields.Float('Bobot')


class work_goal_perspective_financial(models.Model):
    _name = 'hr.work.goal.perspective.financial'
    _inherit = ['mail.thread', 'resource.mixin']

    def _total_target(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_target = row.target_quarter1 + row.target_quarter2 + row.target_quarter3 + row.target_quarter4

    def _total_realisasi(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_realisasi = row.realisasi_q1 + row.realisasi_q2 + row.realisasi_q3 + row.realisasi_q4

    @api.onchange('realisasi_q1')
    def onchange_percent_realisasi_q1(self):
        if self.polarity == '+' and self.target_quarter1 > 0:
            self.percent_realisasi_q1 = round((self.realisasi_q1 * 100) / self.target_quarter1, 2)
        else:
            if self.realisasi_q1 > 0:
                self.percent_realisasi_q1 = round((self.target_quarter1 * 100) / self.realisasi_q1, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q1),
             ('max_value', '>=', self.percent_realisasi_q1)], limit=1)
        if scale.id > 0:
            self.nilai_q1 = scale.value
        self.nilai_target_q1 = (self.bobot * self.nilai_q1) / 100

    @api.onchange('realisasi_q2')
    def onchange_percent_realisasi_q2(self):
        if self.polarity == '+' and self.target_quarter2 > 0:
            self.percent_realisasi_q2 = round((self.realisasi_q2 * 100) / self.target_quarter2, 2)
        else:
            if self.realisasi_q2 > 0:
                self.percent_realisasi_q2 = round((self.target_quarter2 * 100) / self.realisasi_q2, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q2),
             ('max_value', '>=', self.percent_realisasi_q2)], limit=1)
        if scale.id > 0:
            self.nilai_q2 = scale.value
        self.nilai_target_q2 = (self.bobot * self.nilai_q2) / 100

    @api.onchange('realisasi_q3')
    def onchange_percent_realisasi_q3(self):
        if self.polarity == '+' and self.target_quarter3 > 0:
            self.percent_realisasi_q3 = round((self.realisasi_q3 * 100) / self.target_quarter3, 2)
        else:
            if self.realisasi_q3 > 0:
                self.percent_realisasi_q3 = round((self.target_quarter3 * 100) / self.realisasi_q3, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q3),
             ('max_value', '>=', self.percent_realisasi_q3)], limit=1)
        if scale.id > 0:
            self.nilai_q3 = scale.value
        self.nilai_target_q3 = (self.bobot * self.nilai_q3) / 100

    @api.onchange('realisasi_q4')
    def onchange_percent_realisasi_q4(self):
        if self.polarity == '+' and self.target_quarter4 > 0:
            self.percent_realisasi_q4 = round((self.realisasi_q4 * 100) / self.target_quarter4, 2)
        else:
            if self.realisasi_q4 > 0:
                self.percent_realisasi_q4 = round((self.target_quarter4 * 100) / self.realisasi_q4, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q4),
             ('max_value', '>=', self.percent_realisasi_q4)], limit=1)
        if scale.id > 0:
            self.nilai_q4 = scale.value
        self.nilai_target_q4 = (self.bobot * self.nilai_q4) / 100

    def _hitung_Float_q1(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter1 > 0:
                row.percent_realisasi_q1 = round((row.realisasi_q1 * 100) / row.target_quarter1, 2)
            else:
                if row.realisasi_q1 > 0:
                    row.percent_realisasi_q1 = round((row.target_quarter1 * 100) / row.realisasi_q1, 2)

    def _hitung_Float_q2(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter2 > 0:
                row.percent_realisasi_q2 = round((row.realisasi_q2 * 100) / row.target_quarter2, 2)
            else:
                if row.realisasi_q2 > 0:
                    row.percent_realisasi_q2 = round((row.target_quarter2 * 100) / row.realisasi_q2, 2)

    def _hitung_Float_q3(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter3 > 0:
                row.percent_realisasi_q3 = round((row.realisasi_q3 * 100) / row.target_quarter3, 2)
            else:
                if row.realisasi_q3 > 0:
                    row.percent_realisasi_q3 = round((row.target_quarter3 * 100) / row.realisasi_q3, 2)

    def _hitung_Float_q4(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter4 > 0:
                row.percent_realisasi_q4 = round((row.realisasi_q4 * 100) / row.target_quarter4, 2)
            else:
                if row.realisasi_q4 > 0:
                    row.percent_realisasi_q4 = round((row.target_quarter4 * 100) / row.realisasi_q4, 2)

    def _hitung_scala_nilai_q1(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q1),
                 ('max_value', '>=', row.percent_realisasi_q1)], limit=1)
            if scale.id > 0:
                row.nilai_q1 = scale.value

    def _hitung_scala_nilai_q2(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q2),
                 ('max_value', '>=', row.percent_realisasi_q2)], limit=1)
            if scale.id > 0:
                row.nilai_q2 = scale.value

    def _hitung_scala_nilai_q3(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q3),
                 ('max_value', '>=', row.percent_realisasi_q3)], limit=1)
            if scale.id > 0:
                row.nilai_q3 = scale.value

    def _hitung_scala_nilai_q4(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q4),
                 ('max_value', '>=', row.percent_realisasi_q4)], limit=1)
            if scale.id > 0:
                row.nilai_q4 = scale.value

    def _hitung_target_nilai_q1(self):
        for row in self:
            row.nilai_target_q1 = (row.bobot * row.nilai_q1) / 100

    def _hitung_target_nilai_q2(self):
        for row in self:
            row.nilai_target_q2 = (row.bobot * row.nilai_q2) / 100

    def _hitung_target_nilai_q3(self):
        for row in self:
            row.nilai_target_q3 = (row.bobot * row.nilai_q3) / 100

    def _hitung_target_nilai_q4(self):
        for row in self:
            row.nilai_target_q4 = (row.bobot * row.nilai_q4) / 100

    def _hitung_scala_total_nilai(self):
        for row in self:
            row.total_nilai = (row.nilai_q1 + row.nilai_q2 + row.nilai_q3 + row.nilai_q4) / 4

    sasaran_id = fields.Many2one('hr.work.goal.perspective.sasaran.financial', string='Sasaran Kerja')
    work_goal_id = fields.Many2one('hr.work.goal')
    bobot_sasaran = fields.Float(related='sasaran_id.bobot')
    name = fields.Text('Uraian Target')
    jenis = fields.Selection([
        ('kl', 'Kualitatif'),
        ('kn', 'Kuantitatif')
    ], string='Jenis', default='kn')
    polarity = fields.Selection([
        ('+', 'Positif'),
        ('-', 'Negatif')
    ], string='Polaritas', default='+')
    target_semester1 = fields.Float('Target Semester 1')
    target_semester2 = fields.Float('Target Semester 2')
    target_quarter1 = fields.Float('Target Kuartal 1', track_visiblity='onchange')
    target_quarter2 = fields.Float('Target Kuartal 2', track_visiblity='onchange')
    target_quarter3 = fields.Float('Target Kuartal 3', track_visiblity='onchange')
    target_quarter4 = fields.Float('Target Kuartal 4', track_visiblity='onchange')
    total_target = fields.Float('Target', compute=_total_target)
    total_realisasi = fields.Float('Realisasi', compute=_total_realisasi)
    total_nilai = fields.Float("total_nilai", compute=_hitung_scala_total_nilai)
    bobot = fields.Float('Bobot')
    realisasi_q1 = fields.Float('Realisasi Kuartal 1')
    realisasi_q2 = fields.Float('Realisasi Kuartal 2')
    realisasi_q3 = fields.Float('Realisasi Kuartal 3')
    realisasi_q4 = fields.Float('Realisasi Kuartal 4')
    proof_q1 = fields.Text('Bukti Realisasi Q1')
    proof_q2 = fields.Text('Bukti Realisasi Q2')
    proof_q3 = fields.Text('Bukti Realisasi Q3')
    proof_q4 = fields.Text('Bukti Realisasi Q4')
    doc_proof_q1 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q2 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q3 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q4 = fields.Binary('Dokumen Realisasi Q1')
    filename_q1 = fields.Char('filename Q1')
    filename_q2 = fields.Char('filename Q2')
    filename_q3 = fields.Char('filename Q3')
    filename_q4 = fields.Char('filename Q4')
    percent_realisasi_q1 = fields.Float('% Realisasi Q1', default=0, compute=_hitung_Float_q1)
    percent_realisasi_q2 = fields.Float('% Realisasi Q2', default=0, compute=_hitung_Float_q2)
    percent_realisasi_q3 = fields.Float('% Realisasi Q3', default=0, compute=_hitung_Float_q3)
    percent_realisasi_q4 = fields.Float('% Realisasi Q4', default=0, compute=_hitung_Float_q4)
    nilai_q1 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q1)
    nilai_q2 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q2)
    nilai_q3 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q3)
    nilai_q4 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q4)
    nilai_target_q1 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q1)
    nilai_target_q2 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q2)
    nilai_target_q3 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q3)
    nilai_target_q4 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q4)


class work_goal_perspective_customer(models.Model):
    _name = 'hr.work.goal.perspective.customer'
    _inherit = ['mail.thread', 'resource.mixin']

    def _total_target(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_target = row.target_quarter1 + row.target_quarter2 + row.target_quarter3 + row.target_quarter4

    def _total_realisasi(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_realisasi = row.realisasi_q1 + row.realisasi_q2 + row.realisasi_q3 + row.realisasi_q4

    @api.onchange('realisasi_q1')
    def onchange_percent_realisasi_q1(self):
        if self.polarity == '+' and self.target_quarter1 > 0:
            self.percent_realisasi_q1 = round((self.realisasi_q1 * 100) / self.target_quarter1, 2)
        else:
            if self.realisasi_q1 > 0:
                self.percent_realisasi_q1 = round((self.target_quarter1 * 100) / self.realisasi_q1, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q1),
             ('max_value', '>=', self.percent_realisasi_q1)], limit=1)
        if scale.id > 0:
            self.nilai_q1 = scale.value
        self.nilai_target_q1 = (self.bobot * self.nilai_q1) / 100

    @api.onchange('realisasi_q2')
    def onchange_percent_realisasi_q2(self):
        if self.polarity == '+' and self.target_quarter2 > 0:
            self.percent_realisasi_q2 = round((self.realisasi_q2 * 100) / self.target_quarter2, 2)
        else:
            if self.realisasi_q2 > 0:
                self.percent_realisasi_q2 = round((self.target_quarter2 * 100) / self.realisasi_q2, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q2),
             ('max_value', '>=', self.percent_realisasi_q2)], limit=1)
        if scale.id > 0:
            self.nilai_q2 = scale.value
        self.nilai_target_q2 = (self.bobot * self.nilai_q2) / 100

    @api.onchange('realisasi_q3')
    def onchange_percent_realisasi_q3(self):
        if self.polarity == '+' and self.target_quarter3 > 0:
            self.percent_realisasi_q3 = round((self.realisasi_q3 * 100) / self.target_quarter3, 2)
        else:
            if self.realisasi_q3 > 0:
                self.percent_realisasi_q3 = round((self.target_quarter3 * 100) / self.realisasi_q3, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q3),
             ('max_value', '>=', self.percent_realisasi_q3)], limit=1)
        if scale.id > 0:
            self.nilai_q3 = scale.value
        self.nilai_target_q3 = (self.bobot * self.nilai_q3) / 100

    @api.onchange('realisasi_q4')
    def onchange_percent_realisasi_q4(self):
        if self.polarity == '+' and self.target_quarter4 > 0:
            self.percent_realisasi_q4 = round((self.realisasi_q4 * 100) / self.target_quarter4, 2)
        else:
            if self.realisasi_q4 > 0:
                self.percent_realisasi_q4 = round((self.target_quarter4 * 100) / self.realisasi_q4, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q4),
             ('max_value', '>=', self.percent_realisasi_q4)], limit=1)
        if scale.id > 0:
            self.nilai_q4 = scale.value
        self.nilai_target_q4 = (self.bobot * self.nilai_q4) / 100

    def _hitung_Float_q1(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter1 > 0:
                row.percent_realisasi_q1 = round((row.realisasi_q1 * 100) / row.target_quarter1, 2)
            else:
                if row.realisasi_q1 > 0:
                    row.percent_realisasi_q1 = round((row.target_quarter1 * 100) / row.realisasi_q1, 2)

    def _hitung_Float_q2(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter2 > 0:
                row.percent_realisasi_q2 = round((row.realisasi_q2 * 100) / row.target_quarter2, 2)
            else:
                if row.realisasi_q2 > 0:
                    row.percent_realisasi_q2 = round((row.target_quarter2 * 100) / row.realisasi_q2, 2)

    def _hitung_Float_q3(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter3 > 0:
                row.percent_realisasi_q3 = round((row.realisasi_q3 * 100) / row.target_quarter3, 2)
            else:
                if row.realisasi_q3 > 0:
                    row.percent_realisasi_q3 = round((row.target_quarter3 * 100) / row.realisasi_q3, 2)

    def _hitung_Float_q4(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter4 > 0:
                row.percent_realisasi_q4 = round((row.realisasi_q4 * 100) / row.target_quarter4, 2)
            else:
                if row.realisasi_q4 > 0:
                    row.percent_realisasi_q4 = round((row.target_quarter4 * 100) / row.realisasi_q4, 2)

    def _hitung_scala_nilai_q1(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q1),
                 ('max_value', '>=', row.percent_realisasi_q1)], limit=1)
            if scale.id > 0:
                row.nilai_q1 = scale.value

    def _hitung_scala_nilai_q2(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q2),
                 ('max_value', '>=', row.percent_realisasi_q2)], limit=1)
            if scale.id > 0:
                row.nilai_q2 = scale.value

    def _hitung_scala_nilai_q3(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q3),
                 ('max_value', '>=', row.percent_realisasi_q3)], limit=1)
            if scale.id > 0:
                row.nilai_q3 = scale.value

    def _hitung_scala_nilai_q4(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q4),
                 ('max_value', '>=', row.percent_realisasi_q4)], limit=1)
            if scale.id > 0:
                row.nilai_q4 = scale.value

    def _hitung_target_nilai_q1(self):
        for row in self:
            row.nilai_target_q1 = (row.bobot * row.nilai_q1) / 100

    def _hitung_target_nilai_q2(self):
        for row in self:
            row.nilai_target_q2 = (row.bobot * row.nilai_q2) / 100

    def _hitung_target_nilai_q3(self):
        for row in self:
            row.nilai_target_q3 = (row.bobot * row.nilai_q3) / 100

    def _hitung_target_nilai_q4(self):
        for row in self:
            row.nilai_target_q4 = (row.bobot * row.nilai_q4) / 100

    def _hitung_scala_total_nilai(self):
        for row in self:
            row.total_nilai = (row.nilai_q1 + row.nilai_q2 + row.nilai_q3 + row.nilai_q4) / 4

    sasaran_id = fields.Many2one('hr.work.goal.perspective.sasaran.customer', string='Sasaran Kerja')
    work_goal_id = fields.Many2one('hr.work.goal')
    bobot_sasaran = fields.Float(related='sasaran_id.bobot')
    name = fields.Text('Uraian Target')
    jenis = fields.Selection([
        ('kl', 'Kualitatif'),
        ('kn', 'Kuantitatif')
    ], string='Jenis', default='kn')
    polarity = fields.Selection([
        ('+', 'Positif'),
        ('-', 'Negatif')
    ], string='Polaritas', default='+')
    target_semester1 = fields.Float('Target Semester 1')
    target_semester2 = fields.Float('Target Semester 2')
    target_quarter1 = fields.Float('Target Kuartal 1', track_visiblity='onchange')
    target_quarter2 = fields.Float('Target Kuartal 2', track_visiblity='onchange')
    target_quarter3 = fields.Float('Target Kuartal 3', track_visiblity='onchange')
    target_quarter4 = fields.Float('Target Kuartal 4', track_visiblity='onchange')
    total_target = fields.Float('Target', compute=_total_target)
    total_realisasi = fields.Float('Realisasi', compute=_total_realisasi)
    total_nilai = fields.Float("total_nilai", compute=_hitung_scala_total_nilai)
    bobot = fields.Float('Bobot')
    realisasi_q1 = fields.Float('Realisasi Kuartal 1')
    realisasi_q2 = fields.Float('Realisasi Kuartal 2')
    realisasi_q3 = fields.Float('Realisasi Kuartal 3')
    realisasi_q4 = fields.Float('Realisasi Kuartal 4')
    proof_q1 = fields.Text('Bukti Realisasi Q1')
    proof_q2 = fields.Text('Bukti Realisasi Q2')
    proof_q3 = fields.Text('Bukti Realisasi Q3')
    proof_q4 = fields.Text('Bukti Realisasi Q4')
    doc_proof_q1 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q2 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q3 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q4 = fields.Binary('Dokumen Realisasi Q1')
    filename_q1 = fields.Char('filename Q1')
    filename_q2 = fields.Char('filename Q2')
    filename_q3 = fields.Char('filename Q3')
    filename_q4 = fields.Char('filename Q4')
    percent_realisasi_q1 = fields.Float('% Realisasi Q1', default=0, compute=_hitung_Float_q1)
    percent_realisasi_q2 = fields.Float('% Realisasi Q2', default=0, compute=_hitung_Float_q2)
    percent_realisasi_q3 = fields.Float('% Realisasi Q3', default=0, compute=_hitung_Float_q3)
    percent_realisasi_q4 = fields.Float('% Realisasi Q4', default=0, compute=_hitung_Float_q4)
    nilai_q1 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q1)
    nilai_q2 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q2)
    nilai_q3 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q3)
    nilai_q4 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q4)
    nilai_target_q1 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q1)
    nilai_target_q2 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q2)
    nilai_target_q3 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q3)
    nilai_target_q4 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q4)


class work_goal_perspective_bp(models.Model):
    _name = 'hr.work.goal.perspective.bp'
    _inherit = ['mail.thread', 'resource.mixin']

    def _total_target(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_target = row.target_quarter1 + row.target_quarter2 + row.target_quarter3 + row.target_quarter4

    def _total_realisasi(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_realisasi = row.realisasi_q1 + row.realisasi_q2 + row.realisasi_q3 + row.realisasi_q4

    @api.onchange('realisasi_q1')
    def onchange_percent_realisasi_q1(self):
        if self.polarity == '+' and self.target_quarter1 > 0:
            self.percent_realisasi_q1 = round((self.realisasi_q1 * 100) / self.target_quarter1, 2)
        else:
            if self.realisasi_q1 > 0:
                self.percent_realisasi_q1 = round((self.target_quarter1 * 100) / self.realisasi_q1, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q1),
             ('max_value', '>=', self.percent_realisasi_q1)], limit=1)
        if scale.id > 0:
            self.nilai_q1 = scale.value
        self.nilai_target_q1 = (self.bobot * self.nilai_q1) / 100

    @api.onchange('realisasi_q2')
    def onchange_percent_realisasi_q2(self):
        if self.polarity == '+' and self.target_quarter2 > 0:
            self.percent_realisasi_q2 = round((self.realisasi_q2 * 100) / self.target_quarter2, 2)
        else:
            if self.realisasi_q2 > 0:
                self.percent_realisasi_q2 = round((self.target_quarter2 * 100) / self.realisasi_q2, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q2),
             ('max_value', '>=', self.percent_realisasi_q2)], limit=1)
        if scale.id > 0:
            self.nilai_q2 = scale.value
        self.nilai_target_q2 = (self.bobot * self.nilai_q2) / 100

    @api.onchange('realisasi_q3')
    def onchange_percent_realisasi_q3(self):
        if self.polarity == '+' and self.target_quarter3 > 0:
            self.percent_realisasi_q3 = round((self.realisasi_q3 * 100) / self.target_quarter3, 2)
        else:
            if self.realisasi_q3 > 0:
                self.percent_realisasi_q3 = round((self.target_quarter3 * 100) / self.realisasi_q3, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q3),
             ('max_value', '>=', self.percent_realisasi_q3)], limit=1)
        if scale.id > 0:
            self.nilai_q3 = scale.value
        self.nilai_target_q3 = (self.bobot * self.nilai_q3) / 100

    @api.onchange('realisasi_q4')
    def onchange_percent_realisasi_q4(self):
        if self.polarity == '+' and self.target_quarter4 > 0:
            self.percent_realisasi_q4 = round((self.realisasi_q4 * 100) / self.target_quarter4, 2)
        else:
            if self.realisasi_q4 > 0:
                self.percent_realisasi_q4 = round((self.target_quarter4 * 100) / self.realisasi_q4, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q4),
             ('max_value', '>=', self.percent_realisasi_q4)], limit=1)
        if scale.id > 0:
            self.nilai_q4 = scale.value
        self.nilai_target_q4 = (self.bobot * self.nilai_q4) / 100

    def _hitung_Float_q1(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter1 > 0:
                row.percent_realisasi_q1 = round((row.realisasi_q1 * 100) / row.target_quarter1, 2)
            else:
                if row.realisasi_q1 > 0:
                    row.percent_realisasi_q1 = round((row.target_quarter1 * 100) / row.realisasi_q1, 2)

    def _hitung_Float_q2(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter2 > 0:
                row.percent_realisasi_q2 = round((row.realisasi_q2 * 100) / row.target_quarter2, 2)
            else:
                if row.realisasi_q2 > 0:
                    row.percent_realisasi_q2 = round((row.target_quarter2 * 100) / row.realisasi_q2, 2)

    def _hitung_Float_q3(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter3 > 0:
                row.percent_realisasi_q3 = round((row.realisasi_q3 * 100) / row.target_quarter3, 2)
            else:
                if row.realisasi_q3 > 0:
                    row.percent_realisasi_q3 = round((row.target_quarter3 * 100) / row.realisasi_q3, 2)

    def _hitung_Float_q4(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter4 > 0:
                row.percent_realisasi_q4 = round((row.realisasi_q4 * 100) / row.target_quarter4, 2)
            else:
                if row.realisasi_q4 > 0:
                    row.percent_realisasi_q4 = round((row.target_quarter4 * 100) / row.realisasi_q4, 2)

    def _hitung_scala_nilai_q1(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q1),
                 ('max_value', '>=', row.percent_realisasi_q1)], limit=1)
            if scale.id > 0:
                row.nilai_q1 = scale.value

    def _hitung_scala_nilai_q2(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q2),
                 ('max_value', '>=', row.percent_realisasi_q2)], limit=1)
            if scale.id > 0:
                row.nilai_q2 = scale.value

    def _hitung_scala_nilai_q3(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q3),
                 ('max_value', '>=', row.percent_realisasi_q3)], limit=1)
            if scale.id > 0:
                row.nilai_q3 = scale.value

    def _hitung_scala_nilai_q4(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q4),
                 ('max_value', '>=', row.percent_realisasi_q4)], limit=1)
            if scale.id > 0:
                row.nilai_q4 = scale.value

    def _hitung_target_nilai_q1(self):
        for row in self:
            row.nilai_target_q1 = (row.bobot * row.nilai_q1) / 100

    def _hitung_target_nilai_q2(self):
        for row in self:
            row.nilai_target_q2 = (row.bobot * row.nilai_q2) / 100

    def _hitung_target_nilai_q3(self):
        for row in self:
            row.nilai_target_q3 = (row.bobot * row.nilai_q3) / 100

    def _hitung_target_nilai_q4(self):
        for row in self:
            row.nilai_target_q4 = (row.bobot * row.nilai_q4) / 100

    def _hitung_scala_total_nilai(self):
        for row in self:
            row.total_nilai = (row.nilai_q1 + row.nilai_q2 + row.nilai_q3 + row.nilai_q4) / 4

    sasaran_id = fields.Many2one('hr.work.goal.perspective.sasaran.bp', string='Sasaran Kerja')
    work_goal_id = fields.Many2one('hr.work.goal')
    bobot_sasaran = fields.Float(related='sasaran_id.bobot')
    name = fields.Text('Uraian Target')
    jenis = fields.Selection([
        ('kl', 'Kualitatif'),
        ('kn', 'Kuantitatif')
    ], string='Jenis', default='kn')
    polarity = fields.Selection([
        ('+', 'Positif'),
        ('-', 'Negatif')
    ], string='Polaritas', default='+')
    target_semester1 = fields.Float('Target Semester 1')
    target_semester2 = fields.Float('Target Semester 2')
    target_quarter1 = fields.Float('Target Kuartal 1', track_visiblity='onchange')
    target_quarter2 = fields.Float('Target Kuartal 2', track_visiblity='onchange')
    target_quarter3 = fields.Float('Target Kuartal 3', track_visiblity='onchange')
    target_quarter4 = fields.Float('Target Kuartal 4', track_visiblity='onchange')
    total_target = fields.Float('Target', compute=_total_target)
    total_realisasi = fields.Float('Realisasi', compute=_total_realisasi)
    total_nilai = fields.Float("total_nilai", compute=_hitung_scala_total_nilai)
    bobot = fields.Float('Bobot')
    realisasi_q1 = fields.Float('Realisasi Kuartal 1')
    realisasi_q2 = fields.Float('Realisasi Kuartal 2')
    realisasi_q3 = fields.Float('Realisasi Kuartal 3')
    realisasi_q4 = fields.Float('Realisasi Kuartal 4')
    proof_q1 = fields.Text('Bukti Realisasi Q1')
    proof_q2 = fields.Text('Bukti Realisasi Q2')
    proof_q3 = fields.Text('Bukti Realisasi Q3')
    proof_q4 = fields.Text('Bukti Realisasi Q4')
    doc_proof_q1 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q2 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q3 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q4 = fields.Binary('Dokumen Realisasi Q1')
    filename_q1 = fields.Char('filename Q1')
    filename_q2 = fields.Char('filename Q2')
    filename_q3 = fields.Char('filename Q3')
    filename_q4 = fields.Char('filename Q4')
    percent_realisasi_q1 = fields.Float('% Realisasi Q1', default=0, compute=_hitung_Float_q1)
    percent_realisasi_q2 = fields.Float('% Realisasi Q2', default=0, compute=_hitung_Float_q2)
    percent_realisasi_q3 = fields.Float('% Realisasi Q3', default=0, compute=_hitung_Float_q3)
    percent_realisasi_q4 = fields.Float('% Realisasi Q4', default=0, compute=_hitung_Float_q4)
    nilai_q1 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q1)
    nilai_q2 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q2)
    nilai_q3 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q3)
    nilai_q4 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q4)
    nilai_target_q1 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q1)
    nilai_target_q2 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q2)
    nilai_target_q3 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q3)
    nilai_target_q4 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q4)


class work_goal_perspective_people(models.Model):
    _name = 'hr.work.goal.perspective.people'
    _inherit = ['mail.thread', 'resource.mixin']

    def _total_target(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_target = row.target_quarter1 + row.target_quarter2 + row.target_quarter3 + row.target_quarter4

    def _total_realisasi(self):
        for row in self:
            # row.total_target = row.target_semester1 + row.target_semester2
            row.total_realisasi = row.realisasi_q1 + row.realisasi_q2 + row.realisasi_q3 + row.realisasi_q4

    @api.onchange('realisasi_q1')
    def onchange_percent_realisasi_q1(self):
        if self.polarity == '+' and self.target_quarter1 > 0:
            self.percent_realisasi_q1 = round((self.realisasi_q1 * 100) / self.target_quarter1, 2)
        else:
            if self.realisasi_q1 > 0:
                self.percent_realisasi_q1 = round((self.target_quarter1 * 100) / self.realisasi_q1, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q1),
             ('max_value', '>=', self.percent_realisasi_q1)], limit=1)
        if scale.id > 0:
            self.nilai_q1 = scale.value
        self.nilai_target_q1 = (self.bobot * self.nilai_q1) / 100

    @api.onchange('realisasi_q2')
    def onchange_percent_realisasi_q2(self):
        if self.polarity == '+' and self.target_quarter2 > 0:
            self.percent_realisasi_q2 = round((self.realisasi_q2 * 100) / self.target_quarter2, 2)
        else:
            if self.realisasi_q2 > 0:
                self.percent_realisasi_q2 = round((self.target_quarter2 * 100) / self.realisasi_q2, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q2),
             ('max_value', '>=', self.percent_realisasi_q2)], limit=1)
        if scale.id > 0:
            self.nilai_q2 = scale.value
        self.nilai_target_q2 = (self.bobot * self.nilai_q2) / 100

    @api.onchange('realisasi_q3')
    def onchange_percent_realisasi_q3(self):
        if self.polarity == '+' and self.target_quarter3 > 0:
            self.percent_realisasi_q3 = round((self.realisasi_q3 * 100) / self.target_quarter3, 2)
        else:
            if self.realisasi_q3 > 0:
                self.percent_realisasi_q3 = round((self.target_quarter3 * 100) / self.realisasi_q3, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q3),
             ('max_value', '>=', self.percent_realisasi_q3)], limit=1)
        if scale.id > 0:
            self.nilai_q3 = scale.value
        self.nilai_target_q3 = (self.bobot * self.nilai_q3) / 100

    @api.onchange('realisasi_q4')
    def onchange_percent_realisasi_q4(self):
        if self.polarity == '+' and self.target_quarter4 > 0:
            self.percent_realisasi_q4 = round((self.realisasi_q4 * 100) / self.target_quarter4, 2)
        else:
            if self.realisasi_q4 > 0:
                self.percent_realisasi_q4 = round((self.target_quarter4 * 100) / self.realisasi_q4, 2)
        scale = self.env['hr.work.goal.scale'].search(
            [('type', '=', self.jenis), ('min_value', '<=', self.percent_realisasi_q4),
             ('max_value', '>=', self.percent_realisasi_q4)], limit=1)
        if scale.id > 0:
            self.nilai_q4 = scale.value
        self.nilai_target_q4 = (self.bobot * self.nilai_q4) / 100

    def _hitung_Float_q1(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter1 > 0:
                row.percent_realisasi_q1 = round((row.realisasi_q1 * 100) / row.target_quarter1, 2)
            else:
                if row.realisasi_q1 > 0:
                    row.percent_realisasi_q1 = round((row.target_quarter1 * 100) / row.realisasi_q1, 2)

    def _hitung_Float_q2(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter2 > 0:
                row.percent_realisasi_q2 = round((row.realisasi_q2 * 100) / row.target_quarter2, 2)
            else:
                if row.realisasi_q2 > 0:
                    row.percent_realisasi_q2 = round((row.target_quarter2 * 100) / row.realisasi_q2, 2)

    def _hitung_Float_q3(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter3 > 0:
                row.percent_realisasi_q3 = round((row.realisasi_q3 * 100) / row.target_quarter3, 2)
            else:
                if row.realisasi_q3 > 0:
                    row.percent_realisasi_q3 = round((row.target_quarter3 * 100) / row.realisasi_q3, 2)

    def _hitung_Float_q4(self):
        for row in self:
            if row.polarity == '+' and row.target_quarter4 > 0:
                row.percent_realisasi_q4 = round((row.realisasi_q4 * 100) / row.target_quarter4, 2)
            else:
                if row.realisasi_q4 > 0:
                    row.percent_realisasi_q4 = round((row.target_quarter4 * 100) / row.realisasi_q4, 2)

    def _hitung_scala_nilai_q1(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q1),
                 ('max_value', '>=', row.percent_realisasi_q1)], limit=1)
            if scale.id > 0:
                row.nilai_q1 = scale.value

    def _hitung_scala_nilai_q2(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q2),
                 ('max_value', '>=', row.percent_realisasi_q2)], limit=1)
            if scale.id > 0:
                row.nilai_q2 = scale.value

    def _hitung_scala_nilai_q3(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q3),
                 ('max_value', '>=', row.percent_realisasi_q3)], limit=1)
            if scale.id > 0:
                row.nilai_q3 = scale.value

    def _hitung_scala_nilai_q4(self):
        for row in self:
            scale = self.env['hr.work.goal.scale'].search(
                [('type', '=', row.jenis), ('min_value', '<=', row.percent_realisasi_q4),
                 ('max_value', '>=', row.percent_realisasi_q4)], limit=1)
            if scale.id > 0:
                row.nilai_q4 = scale.value

    def _hitung_target_nilai_q1(self):
        for row in self:
            row.nilai_target_q1 = (row.bobot * row.nilai_q1) / 100

    def _hitung_target_nilai_q2(self):
        for row in self:
            row.nilai_target_q2 = (row.bobot * row.nilai_q2) / 100

    def _hitung_target_nilai_q3(self):
        for row in self:
            row.nilai_target_q3 = (row.bobot * row.nilai_q3) / 100

    def _hitung_target_nilai_q4(self):
        for row in self:
            row.nilai_target_q4 = (row.bobot * row.nilai_q4) / 100

    def _hitung_scala_total_nilai(self):
        for row in self:
            row.total_nilai = (row.nilai_q1 + row.nilai_q2 + row.nilai_q3 + row.nilai_q4) / 4

    sasaran_id = fields.Many2one('hr.work.goal.perspective.sasaran.people', string='Sasaran Kerja')
    work_goal_id = fields.Many2one('hr.work.goal')
    bobot_sasaran = fields.Float(related='sasaran_id.bobot')
    name = fields.Char('Uraian Target')
    jenis = fields.Selection([
        ('kl', 'Kualitatif'),
        ('kn', 'Kuantitatif')
    ], string='Jenis', default='kn', )
    polarity = fields.Selection([
        ('+', 'Positif'),
        ('-', 'Negatif')
    ], string='Polaritas', default='+')
    target_semester1 = fields.Float('Target Semester 1')
    target_semester2 = fields.Float('Target Semester 2')
    target_quarter1 = fields.Float('Target Kuartal 1', track_visiblity='onchange')
    target_quarter2 = fields.Float('Target Kuartal 2', track_visiblity='onchange')
    target_quarter3 = fields.Float('Target Kuartal 3', track_visiblity='onchange')
    target_quarter4 = fields.Float('Target Kuartal 4', track_visiblity='onchange')
    total_target = fields.Float('Target', compute=_total_target)
    total_realisasi = fields.Float('Realisasi', compute=_total_realisasi)
    total_nilai = fields.Float("total_nilai", compute=_hitung_scala_total_nilai)
    bobot = fields.Float('Bobot')
    realisasi_q1 = fields.Float('Realisasi Kuartal 1')
    realisasi_q2 = fields.Float('Realisasi Kuartal 2')
    realisasi_q3 = fields.Float('Realisasi Kuartal 3')
    realisasi_q4 = fields.Float('Realisasi Kuartal 4')
    proof_q1 = fields.Text('Bukti Realisasi Q1')
    proof_q2 = fields.Text('Bukti Realisasi Q2')
    proof_q3 = fields.Text('Bukti Realisasi Q3')
    proof_q4 = fields.Text('Bukti Realisasi Q4')
    doc_proof_q1 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q2 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q3 = fields.Binary('Dokumen Realisasi Q1')
    doc_proof_q4 = fields.Binary('Dokumen Realisasi Q1')
    filename_q1 = fields.Char('filename Q1')
    filename_q2 = fields.Char('filename Q2')
    filename_q3 = fields.Char('filename Q3')
    filename_q4 = fields.Char('filename Q4')
    percent_realisasi_q1 = fields.Float('% Realisasi Q1', default=0, compute=_hitung_Float_q1)
    percent_realisasi_q2 = fields.Float('% Realisasi Q2', default=0, compute=_hitung_Float_q2)
    percent_realisasi_q3 = fields.Float('% Realisasi Q3', default=0, compute=_hitung_Float_q3)
    percent_realisasi_q4 = fields.Float('% Realisasi Q4', default=0, compute=_hitung_Float_q4)
    nilai_q1 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q1)
    nilai_q2 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q2)
    nilai_q3 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q3)
    nilai_q4 = fields.Float('Nilai', default=0, compute=_hitung_scala_nilai_q4)
    nilai_target_q1 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q1)
    nilai_target_q2 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q2)
    nilai_target_q3 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q3)
    nilai_target_q4 = fields.Float('Nilai Target', compute=_hitung_target_nilai_q4)


class work_kpi(models.Model):
    _name = 'hr.work.kpi'

    def _get_default_employee(self):
        employee = self.env['hr.employee'].search([('user_id', '=', self.env.user.id)])
        return employee

    @api.depends('user_id')
    def _current_user_is_employee(self):
        for row in self:
            row.current_user_is_employee = self.env.user.id == row.user_id.id

    @api.depends('reviewer_user_id')
    def _current_user_is_reviewer(self):
        for row in self:
            row.current_user_is_reviewer = self.env.user.id == row.reviewer_user_id.id

    @api.depends('reviewer2_user_id')
    def _current_user_is_reviewer2(self):
        for row in self:
            row.current_user_is_reviewer2 = self.env.user.id == row.reviewer2_user_id.id

    def _current_user_is_pengurus(self):
        for row in self:
            row.current_user_is_pengurus = self.env.user.has_group('hr_ykp_training.group_pengurus')

    def _hitung_self_value(self):
        for row in self:
            total_value = 0
            for indikator in row.kpi_detail_ids:
                total_value += indikator.kpi_weight_employee
            row.self_value = total_value

    def _hitung_self_value_category(self):
        for row in self:
            category = self.env['hr.work.kpi.weight.category'].search(
                [('min_value', '<', row.self_value), ('max_value', '>=', row.self_value)], limit=1)
            if category.id > 0:
                row.self_kpi_kategori = category.name

    def _hitung_penilai1_value(self):
        for row in self:
            total_value = 0
            for indikator in row.kpi_detail_ids:
                total_value += indikator.kpi_weight_reviewer
            row.penilai1_value = total_value

    def _hitung_penilai1_value_category(self):
        for row in self:
            category = self.env['hr.work.kpi.weight.category'].search(
                [('min_value', '<', row.penilai1_value), ('max_value', '>=', row.penilai1_value)], limit=1)
            if category.id > 0:
                row.penilai1_kpi_kategori = category.name

    def _hitung_penilai2_value(self):
        for row in self:
            total_value = 0
            for indikator in row.kpi_detail_ids:
                total_value += indikator.kpi_weight_reviewer2
            row.penilai2_value = total_value

    def _hitung_penilai2_value_category(self):
        for row in self:
            category = self.env['hr.work.kpi.weight.category'].search(
                [('min_value', '<', row.penilai2_value), ('max_value', '>=', row.penilai2_value)], limit=1)
            if category.id > 0:
                row.penilai2_kpi_kategori = category.name

    def _default_details(self):
        data = []
        kinerja = self.env['hr.work.kpi.indicator'].search([('type', '=', 'kinerja')])
        for k in kinerja:
            weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', k.id)])
            result = """
            <table style="width:100%;border-collapse: collapse;">
                <tr>
                    <th style="border: 1px solid #ddd;padding: 4px;">Indikator</th>
                    <th style="border: 1px solid #ddd;padding: 4px;">Rating</th>\
                    <th style="border: 1px solid #ddd;padding: 4px;">Kondite</th>
                </tr>
                {}
            </table/>
            """
            row = ""
            for w in weights:
                row += """
                        <tr>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                        </tr>
                        """.format(w.name, w.rating, w.nilai_konduite)
            result = result.format(row)
            # for w in weights:
            #     result += "rating : {} - nilai konduite : {}".format(w.rating, w.nilai_konduite) + "<br>"
            data.append((0, 0, {
                'type': 'kinerja',
                'indikator_penilaian_id': k.id,
                'kriteria_penilaian': result
            }))
        kompetensi = self.env['hr.work.kpi.indicator'].search([('type', '=', 'kompetensi')])
        for k in kompetensi:
            weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', k.id)])
            result = """
                        <table style="width:100%;border-collapse: collapse;">
                            <tr>
                                <th style="border: 1px solid #ddd;padding: 4px;">Indikator</th>
                                <th style="border: 1px solid #ddd;padding: 4px;">Rating</th>\
                                <th style="border: 1px solid #ddd;padding: 4px;">Kondite</th>
                            </tr>
                            {}
                        </table/>
                        """
            row = ""
            for w in weights:
                row += """
                        <tr>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                        </tr>
                        """.format(w.name, w.rating, w.nilai_konduite)
            result = result.format(row)
            data.append((0, 0, {
                'type': 'kompetensi',
                'indikator_penilaian_id': k.id,
                'kriteria_penilaian': result
            }))
        perilaku = self.env['hr.work.kpi.indicator'].search([('type', '=', 'perilaku')])
        for k in perilaku:
            weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', k.id)])
            result = """
                        <table style="width:100%;border-collapse: collapse;">
                            <tr>
                                <th style="border: 1px solid #ddd;padding: 4px;">Indikator</th>
                                <th style="border: 1px solid #ddd;padding: 4px;">Rating</th>\
                                <th style="border: 1px solid #ddd;padding: 4px;">Kondite</th>
                            </tr>
                            {}
                        </table/>
                        """
            row = ""
            for w in weights:
                row += """
                        <tr>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                            <td style="border: 1px solid #ddd;padding: 4px;">{}</td>
                        </tr>
                        """.format(w.name, w.rating, w.nilai_konduite)
            result = result.format(row)
            data.append((0, 0, {
                'type': 'perilaku',
                'indikator_penilaian_id': k.id,
                'kriteria_penilaian': result
            }))
        return data

    def _compute_career(self):
        for row in self:
            category = self.env['hr.work.kpi.weight.category'].search([('name', '=', row.penilai1_kpi_kategori)],
                                                                      limit=1)
            if category.id > 0:
                row.corporate_carrer_id = self.env['hr.corporate.career.scale'].search(
                    [('kpi_weight_id', '=', category.id), ('kpi_culcure_ids', '=', row.corporate_culture_id.id)],
                    limit=1)

    def _compute_career2(self):
        for row in self:
            category = self.env['hr.work.kpi.weight.category'].search([('name', '=', row.penilai2_kpi_kategori)],
                                                                      limit=1)
            if category.id > 0:
                row.corporate_carrer_id = self.env['hr.corporate.career.scale'].search(
                    [('kpi_weight_id', '=', category.id), ('kpi_culcure_ids', '=', row.corporate_culture_id.id)],
                    limit=1)

    @api.onchange('reviewer_id', 'reviewer2_id')
    def on_change_reviewer(self):
        if self.reviewer_id:
            if self.reviewer_id.id == self.reviewer2_id.id:
                raise exceptions.ValidationError('Penilai 1 dan Penilai Pemutus tidak boleh orang yang sama')
        if self.reviewer2_id:
            if self.reviewer_id.id == self.reviewer2_id.id:
                raise exceptions.ValidationError('Penilai 1 dan Penilai Pemutus tidak boleh orang yang sama')

    current_user_is_reviewer = fields.Boolean(compute=_current_user_is_reviewer)
    current_user_is_reviewer2 = fields.Boolean(compute=_current_user_is_reviewer2)
    current_user_is_employee = fields.Boolean(compute=_current_user_is_employee)
    current_user_is_pengurus = fields.Boolean(compute=_current_user_is_pengurus)
    name = fields.Many2one('hr.employee', string='Nama Pegawai', default=_get_default_employee)
    user_id = fields.Many2one(related='name.user_id')
    # nik = fields.Char(related='name.nik', string='NIK')
    job_id = fields.Many2one(related='name.job_id', string='Jabatan')
    department_id = fields.Many2one(related='name.department_id', string='Department')
    goal_year = fields.Char('Tahun Sasaran', default=datetime.datetime.now().year)
    kpi_detail_ids = fields.One2many('hr.work.kpi.detail', 'name', string='Daftar Indikator', default=_default_details)
    approval_date = fields.Date('Tanggal Pengisian', default=datetime.datetime.now())
    reviewer_id = fields.Many2one('hr.employee', string='Penilai 1', domain=[('user_id', '!=', False)])
    reviewer_job_id = fields.Many2one(related='reviewer_id.job_id', string='Jabatan')
    reviewer_assesment_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer_user_id = fields.Many2one(related='reviewer_id.user_id')
    reviewer2_id = fields.Many2one('hr.employee', string='Penilai Pemutus', domain=[('user_id', '!=', False)])
    reviewer2_job_id = fields.Many2one(related='reviewer2_id.job_id', string='Jabatan')
    reviewer2_assesment_date = fields.Date('Tanggal Penilaian', default=datetime.datetime.now())
    reviewer2_user_id = fields.Many2one(related='reviewer2_id.user_id')
    self_value = fields.Float('Nilai Sendiri', compute=_hitung_self_value)
    self_kpi_kategori = fields.Char('Kategori', compute=_hitung_self_value_category)
    self_signature = fields.Binary('Tanda Tangan')
    penilai1_value = fields.Float('Nilai Penilai 1', compute=_hitung_penilai1_value)
    penilai1_kpi_kategori = fields.Char('Kategori', compute=_hitung_penilai1_value_category)
    penilai1_notes = fields.Text('Catatan')
    corporate_culture1_id = fields.Many2one('hr.corporate.culture.scale', string='Corporate Culture')
    corporate_carrer1_id = fields.Many2one('hr.corporate.career.scale', string='Pengembangan Karir', store=True,
                                           compute=_compute_career)
    penilai1_signature = fields.Binary('Tanda Tangan')
    penilai2_value = fields.Float('Nilai Penilai Pemutus', compute=_hitung_penilai2_value)
    penilai2_kpi_kategori = fields.Char('Kategori', compute=_hitung_penilai2_value_category)
    penilai2_notes = fields.Text('Catatan')
    corporate_culture2_id = fields.Many2one('hr.corporate.culture.scale', string='Corporate Culture')
    corporate_carrer2_id = fields.Many2one('hr.corporate.career.scale', string='Pengembangan Karir', store=True,
                                           compute=_compute_career2)
    penilai2_signature = fields.Binary('Tanda Tangan')
    corporate_culture_id = fields.Many2one('hr.corporate.culture.scale', string='Predikat Corporate Culture')
    corporate_carrer_id = fields.Many2one('hr.corporate.career.scale', string='Pengembangan Karir', store=True)
    desc_corporate_culture = fields.Char(related='corporate_culture_id.desc')
    notes = fields.Html('Catatan Pengembangan Kompetensi Pegawai (Diklat) dari Penilai I/Penilai Pemutus :')
    notes2 = fields.Html('Catatan Faktor Penambah/Pengurang Nilai dari Pengurus :')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('sent', 'Dikirim'),
        ('assesment', 'Penilaian 1'),
        ('assesment_pemutus', 'Penilaian Pemutus'),
        ('assesment_pengurus', 'Penilaian Pengurus'),
    ], default='draft')
    semester = fields.Selection([
        ('semester_1', 'Semester 1'),
        ('semester_2', 'Semester 2'),
    ], string='Semester')
    quarter = fields.Selection([
        ('q1', 'Kuartal 1'),
        ('q2', 'Kuartal 2'),
        ('q3', 'Kuartal 3'),
        ('q4', 'Kuartal 4'),
    ], string='Kuartal', default='q1')

    @api.multi
    def action_sent(self):
        self.write({'state': 'sent'})
        not_valid = False
        for detail in self.kpi_detail_ids:
            if detail.kpi_weight_employee == 0:
                not_valid = True
                break
        if not_valid:
            raise exceptions.ValidationError('Silahkan terlebih dahulu di nilai setiap item indikator')

    @api.multi
    def action_assesment(self):
        self.write({'state': 'assesment'})
        not_valid = False
        for detail in self.kpi_detail_ids:
            if detail.kpi_weight_reviewer == 0:
                not_valid = True
                break
        if not_valid:
            raise exceptions.ValidationError('Silahkan terlebih dahulu di nilai setiap item indikator')

    @api.multi
    def action_assesment_pemutus(self):
        self.write({'state': 'assesment_pemutus'})
        not_valid = False
        for detail in self.kpi_detail_ids:
            if detail.kpi_weight_reviewer2 == 0:
                not_valid = True
                break
        if not_valid:
            raise exceptions.ValidationError('Silahkan terlebih dahulu di nilai setiap item indikator')

    @api.multi
    def action_assesment_pengurus(self):
        self.write({'state': 'assesment_pengurus'})

    @api.model
    def create(self, vals):
        semester = self._context['default_semester']
        count = self.env['hr.work.kpi'].search_count(
            [('name', '=', vals.get('name')), ('goal_year', '=', vals.get('goal_year')), ('semester', '=', semester)])
        if count > 0:
            raise exceptions.ValidationError(
                'Data KPI tahun {} untuk {} telah ada'.format(vals.get('goal_year'), semester))
        vals['kpi_detail_ids'] = self._default_details()
        rec = super(work_kpi, self).create(vals)
        return rec


class work_kpi_detail(models.Model):
    _name = 'hr.work.kpi.detail'

    @api.depends('name.user_id')
    def _current_user_is_employee(self):
        for row in self:
            row.current_user_is_employee = self.env.user.id == row.name.user_id.id

    @api.depends('name.reviewer_user_id')
    def _current_user_is_reviewer(self):
        for row in self:
            row.current_user_is_reviewer = self.env.user.id == row.name.reviewer_user_id.id

    @api.depends('name.reviewer2_user_id')
    def _current_user_is_reviewer2(self):
        for row in self:
            row.current_user_is_reviewer2 = self.env.user.id == row.name.reviewer2_user_id.id

    def _compute_kriteria_penilaian(self):
        for row in self:
            if row.indikator_penilaian_id:
                weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', row.indikator_penilaian_id.id)])
                result = '<table>{}</table/>'
                row = ""
                for w in weights:
                    row += """
                    <tr>
                        <td>{}</td>
                        <td>{}</td>
                        <td>{}</td>
                    </tr>
                    """.format(w.name, w.rating, w.nilai_konduite)
                result = result.format(row)
                row.kriteria_penilaian = result

    @api.onchange('kpi_weight_employee')
    def onchange_kpi_weight_employee(self):
        weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', self.indikator_penilaian_id.id)])
        max_kondite = 0
        min_kondite = 100
        for w in weights:
            if max_kondite < w.nilai_konduite:
                max_kondite = w.nilai_konduite
            if min_kondite > w.nilai_konduite:
                min_kondite = w.nilai_konduite
        if self.kpi_weight_employee > max_kondite:
            raise exceptions.ValidationError('Maksimum Nilai Kondite adalah : {}'.format(max_kondite))
        if self.kpi_weight_employee < min_kondite:
            raise exceptions.ValidationError('Minimum Nilai Kondite adalah : {}'.format(min_kondite))

    @api.onchange('kpi_weight_reviewer')
    def onchange_kpi_weight_reviewer(self):
        weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', self.indikator_penilaian_id.id)])
        max_kondite = 0
        min_kondite = 100
        for w in weights:
            if max_kondite < w.nilai_konduite:
                max_kondite = w.nilai_konduite
            if min_kondite > w.nilai_konduite:
                min_kondite = w.nilai_konduite
        if self.kpi_weight_employee > max_kondite:
            raise exceptions.ValidationError('Maksimum Nilai Kondite adalah : {}'.format(max_kondite))
        if self.kpi_weight_employee < min_kondite:
            raise exceptions.ValidationError('Minimum Nilai Kondite adalah : {}'.format(min_kondite))

    @api.onchange('kpi_weight_reviewer2')
    def onchange_kpi_weight_reviewer2(self):
        weights = self.env['hr.work.kpi.weight'].search([('indicator_id', '=', self.indikator_penilaian_id.id)])
        max_kondite = 0
        min_kondite = 100
        for w in weights:
            if max_kondite < w.nilai_konduite:
                max_kondite = w.nilai_konduite
            if min_kondite > w.nilai_konduite:
                min_kondite = w.nilai_konduite
        if self.kpi_weight_employee > max_kondite:
            raise exceptions.ValidationError('Maksimum Nilai Kondite adalah : {}'.format(max_kondite))
        if self.kpi_weight_employee < min_kondite:
            raise exceptions.ValidationError('Minimum Nilai Kondite adalah : {}'.format(min_kondite))

    current_user_is_reviewer = fields.Boolean(compute=_current_user_is_reviewer)
    current_user_is_reviewer2 = fields.Boolean(compute=_current_user_is_reviewer2)
    current_user_is_employee = fields.Boolean(compute=_current_user_is_employee)
    state = fields.Selection(related='name.state')
    name = fields.Many2one('hr.work.kpi', 'KPI')
    reviewer_user_id = fields.Many2one(related='name.reviewer_user_id')
    reviewer2_user_id = fields.Many2one(related='name.reviewer2_user_id')
    type = fields.Selection([
        ('kinerja', 'Kinerja'),
        ('kompetensi', 'Kompetensi'),
        ('perilaku', 'Perilaku'),
    ], string='Faktor Penilaian')
    indikator_penilaian_id = fields.Many2one('hr.work.kpi.indicator', string='Indikator Penilaian')
    kriteria_penilaian = fields.Html()
    bobot = fields.Float('Bobot')
    kpi_weight_employee = fields.Float('Hasil Nilai Pegawai')
    kpi_weight_reviewer = fields.Float('Hasil Nilai Penilai 1')
    kpi_weight_reviewer2 = fields.Float('Hasil Nilai Pemutus')
    kpi_weight_employee_id = fields.Many2one('hr.work.kpi.weight', string='Hasil Nilai Pegawai')
    kpi_weight_reviewer_id = fields.Many2one('hr.work.kpi.weight', string='Hasil Nilai Penilai 1')
    kpi_weight_reviewer2_id = fields.Many2one('hr.work.kpi.weight', string='Hasil Nilai Pemutus')


class work_goal_scale(models.Model):
    _name = 'hr.work.goal.scale'

    name = fields.Char('Kriteria Nilai')
    type = fields.Selection([
        ('kl', 'Kualitatif'),
        ('kn', 'Kuantitatif')
    ], string='Jenis Skala')
    min_value = fields.Float('Min Value')
    max_value = fields.Float('Max Value')
    value = fields.Integer('Skala Nilai')


class work_kpi_indicator(models.Model):
    _name = 'hr.work.kpi.indicator'

    name = fields.Char('Indikator Penilaian')
    desc = fields.Text('Deskripsi')
    type = fields.Selection([
        ('kinerja', 'Kinerja'),
        ('kompetensi', 'Kompetensi'),
        ('perilaku', 'Perilaku'),
    ], string='Faktor Penilaian')


class work_kpi_weight(models.Model):
    _name = 'hr.work.kpi.weight'

    def name_get(self):
        result = []
        for row in self:
            formatted_name = "rating : {} - nilai konduite : {}".format(row.rating, row.nilai_konduite)
            result.append((row.id, formatted_name))
        return result

    indicator_id = fields.Many2one('hr.work.kpi.indicator', string='Indikator Penilaian')
    name = fields.Text('Deskripsi')
    rating = fields.Float('Rating')
    nilai_konduite = fields.Float('Nilai Konduite')


class work_kpi_weight_category(models.Model):
    _name = 'hr.work.kpi.weight.category'

    name = fields.Char('Kategori')
    min_value = fields.Float('Min Value')
    max_value = fields.Float('Max Value')


class corporate_culture_scale(models.Model):
    _name = 'hr.corporate.culture.scale'

    name = fields.Char('Predikat')
    desc = fields.Char('Deskripsi')


class corporate_career_scale(models.Model):
    _name = 'hr.corporate.career.scale'

    name = fields.Char('Predikat')
    kpi_weight_id = fields.Many2one('hr.work.kpi.weight', string='Kategori')
    kpi_culcure_ids = fields.Many2many('hr.corporate.culture.scale', string='Culture')
