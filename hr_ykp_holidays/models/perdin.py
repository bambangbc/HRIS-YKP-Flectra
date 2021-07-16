import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

from flectra import _
from flectra import api, fields, models
from flectra.exceptions import UserError

SESSION_STATES = [('draft', 'Draft'), ("validate", "Approved"), ("refuse", "Refused"), ("set_to_draft", "Set To Draft")]


class Perdin(models.Model):
    _name = "hr.perdin"
    _inherit = ['mail.thread']
    _order = "date_from asc"
    _description = "Perdin"

    @api.model
    def create(self, vals):
        aray_bln = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
        sequence = self.env['ir.sequence'].next_by_code('hr.perdin') or 'Number not found !'
        momon = (datetime.strptime(vals['date_from'], "%Y-%m-%d").month) - 1
        years = str(datetime.strptime(vals['date_from'], "%Y-%m-%d").year)
        months = aray_bln[momon]
        vals['name'] = sequence + "/" + "S/P/YKP-bjb/" + months + "/" + years

        DATETIME_FORMAT = "%Y-%m-%d"
        from_dt = datetime.strptime(vals['date_from'], DATETIME_FORMAT)
        to_dt = datetime.strptime(vals['date_to'], DATETIME_FORMAT)
        timedelta = to_dt - from_dt
        diff_day = timedelta.days + 1
        vals['number_of_days_temp'] = diff_day

        return super(Perdin, self).create(vals)

    def _default_negara_tujuan(self):
        return self.env['res.country'].search([('id', '=', 100)], limit=1)

    @api.depends('approve')
    def _get_approve_employee(self):
        for row in self:
            if row.approve:
                employee = self.env['hr.employee'].search([('user_id', '=', row.approve.id)], limit=1)
                if employee:
                    row.approve_employee_id = employee.id

    name = fields.Char(string="Nomor Perdin", readonly=True)
    state = fields.Selection(string="State",
                             selection=SESSION_STATES,
                             required=True,
                             readonly=True,
                             default=SESSION_STATES[0][0], track_visibility='onchange')
    pejabat_berwenang = fields.Many2one(comodel_name="hr.employee", string="Pejabat Berwenang")
    jabatan = fields.Many2one(comodel_name="hr.job", string="Jabatan")
    tempat_berangkat = fields.Many2one(comodel_name="hr.city", string="Tempat Berangkat")
    negara_tujuan = fields.Many2one('res.country', string="Negara Tujuan", default=_default_negara_tujuan)
    propinsi_tujuan = fields.Many2one(comodel_name="hr.state", string="Propinsi Tujuan")
    tempat_tujuan = fields.Many2one(comodel_name="hr.city", string="Kota Tujuan")
    angkutan = fields.Char(string="Alat angkutan yang di gunakan")
    penanggung_fasilitas = fields.Char(string="Penanggung Fasilitas Perjalanan Dinas")
    datang_di = fields.Char(string="Datang Di")
    tanggal = fields.Date(string="Tanggal")
    date_from = fields.Date('Tanggal Berankat', readonly=True, index=True, copy=False,
                            states={'draft': [('readonly', False)], 'confirm': [('readonly', False)]},
                            track_visibility='onchange')
    date_to = fields.Date('Tanggal Kembali', readonly=True, copy=False,
                          states={'draft': [('readonly', False)], 'confirm': [('readonly', False)]},
                          track_visibility='onchange')
    number_of_days_temp = fields.Float('Jumlah Hari', track_visibility='onchange')
    tujuan_perdin = fields.Char(string="Tujuan Perdin")
    jenis_perdin = fields.Selection([
        ('collecting', 'Collecting'),
        ('reguler', 'Reguler'),
    ], string='Jenis')
    unti_kerja_penyelenggara = fields.Char("Unit Kerja Penyelenggara", default="YAYASAN KESEJAHTERAAN PEGAWAI BANK BJB",
                                           readonly=True)
    cost_centre = fields.Char("Cost Centre", default="YAYASAN KESEJAHTERAAN PEGAWAI", readonly=True)
    perdin_undangan = fields.Boolean("Perdin Undangan")
    approve = fields.Many2one(comodel_name="res.users", string="Approaal")
    approve_employee_id = fields.Many2one('hr.employee', compute=_get_approve_employee)
    employee_ids = fields.One2many("hr.perdin.employee", "perdin_id", "Employee's")
    alat_angkut = fields.Char("Alat Angkut yang digunakan")
    driver = fields.Many2one('hr.employee', string="Driver")

    @api.onchange('date_to')
    def _get_number_of_hours(self):
        """Returns an overtime hours."""
        if self.date_to and self.date_from:
            DATETIME_FORMAT = "%Y-%m-%d"
            from_dt = datetime.strptime(self.date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = timedelta.days + 1
            self.number_of_days_temp = diff_day

    @api.onchange('date_from')
    def _calc_date(self):
        if self.date_from and self.date_to:
            DATETIME_FORMAT = "%Y-%m-%d"
            from_dt = datetime.strptime(self.date_from, DATETIME_FORMAT)
            to_dt = datetime.strptime(self.date_to, DATETIME_FORMAT)
            timedelta = to_dt - from_dt
            diff_day = timedelta.days + 1
            self.number_of_days_temp = diff_day

    @api.multi
    def action_draft(self):
        if self.approve.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[0][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_validate(self):
        if self.approve.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[1][0]
            self.action_calculate_uang_saku()
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_refuse(self):
        if self.approve.id == self.env.uid or self.env.uid == 1:
            self.state = SESSION_STATES[2][0]
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))

    @api.multi
    def action_calculate_uang_saku(self):
        for employee in self.employee_ids:
            employee.hitung_uang_saku()

    def print_pengajuan_perdin(self):
        return {
            'type': 'ir.actions.report',
            'name': 'hr_ykp_holidays.print_form_perdin_pengajuan',
            'report_name': 'hr_ykp_holidays.print_form_perdin_pengajuan',
            'report_type': 'qweb-pdf',
            'model': 'hr.perdin',
            'datas': {
                'id': self.id,
            },
            'nodestroy': True
        }

    def print_report_perdin(self):
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'id_ID')
        except:
            print('error setting locale')

        path = os.path.dirname(os.path.realpath(__file__))
        filename = "{}/../reports/perdin_template.xlsx".format(path)
        wb = load_workbook(filename)
        sheet = wb['Sheet1']
        no = 1
        start_row = 3
        thin_border = Border(left=Side(style='medium'),
                             right=Side(style='medium'),
                             top=Side(style='medium'),
                             bottom=Side(style='medium'))

        for employee in self.employee_ids:
            sheet.insert_rows(start_row)
            allowance = self.env['hr.perdin.allowance'].search(
                [('jenis_pegawai', '=', employee.employee_id.jenis_pegawai.id)], limit=1)
            sheet['A{}'.format(start_row)] = no
            sheet['A{}'.format(start_row)].border = thin_border
            sheet['A{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['B{}'.format(start_row)] = employee.employee_id.name
            sheet['B{}'.format(start_row)].border = thin_border
            sheet['B{}'.format(start_row)].alignment = Alignment(wrap_text=True)
            sheet['B{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['C{}'.format(start_row)] = employee.employee_id.job_id.name
            sheet['C{}'.format(start_row)].border = thin_border
            sheet['C{}'.format(start_row)].alignment = Alignment(wrap_text=True)
            sheet['C{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['D{}'.format(start_row)] = self.tempat_tujuan.name
            sheet['D{}'.format(start_row)].border = thin_border
            sheet['D{}'.format(start_row)].alignment = Alignment(wrap_text=True)
            sheet['D{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['E{}'.format(start_row)] = "{} hari".format(self.number_of_days_temp)
            sheet['E{}'.format(start_row)].border = thin_border
            sheet['E{}'.format(start_row)].alignment = Alignment(wrap_text=True)
            sheet['E{}'.format(start_row)].font = Font(bold=True, size="8")
            nominal = 0
            if allowance:
                if self.negara_tujuan.id == 100:
                    nominal = allowance.nominal
                else:
                    nominal = allowance.nominal_luar
            sheet['F{}'.format(start_row)] = nominal if nominal > 0 else "-"
            sheet['F{}'.format(start_row)].border = thin_border
            sheet['F{}'.format(start_row)].font = Font(bold=True, size="8")
            exclude = False
            for excluded in allowance.location_exclude_ids:
                if excluded.id == self.propinsi_tujuan.id:
                    exclude = True
                    break
            sheet['G{}'.format(start_row)] = "-"
            sheet['G{}'.format(start_row)].border = thin_border
            sheet['G{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['H{}'.format(start_row)] = "-"
            sheet['H{}'.format(start_row)].border = thin_border
            sheet['H{}'.format(start_row)].font = Font(bold=True, size="8")
            if not exclude and self.negara_tujuan.id == 100:
                if 'jawa timur' in self.propinsi_tujuan.name.lower() or 'jawa tengah' in self.propinsi_tujuan.name.lower():
                    sheet['G{}'.format(start_row)] = (allowance.extra_jawa * nominal) / 100
                else:
                    sheet['H{}'.format(start_row)] = (allowance.extra * nominal) / 100
            sheet['I{}'.format(start_row)] = employee.uang_saku
            sheet['I{}'.format(start_row)].border = thin_border
            sheet['I{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['J{}'.format(start_row)] = "=I{}".format(start_row)
            sheet['J{}'.format(start_row)].border = thin_border
            sheet['J{}'.format(start_row)].font = Font(bold=True, size="8")
            no += 1
            start_row += 1
        if start_row > 3:
            sheet.insert_rows(start_row)
            sheet.merge_cells('A{}:I{}'.format(start_row, start_row))
            sheet['A{}'.format(start_row)] = 'Total'
            sheet['A{}'.format(start_row)].font = Font(bold=True, size="8")
            sheet['A{}'.format(start_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet['A{}'.format(start_row)].border = thin_border
            sheet['B{}'.format(start_row)].border = thin_border
            sheet['C{}'.format(start_row)].border = thin_border
            sheet['D{}'.format(start_row)].border = thin_border
            sheet['E{}'.format(start_row)].border = thin_border
            sheet['F{}'.format(start_row)].border = thin_border
            sheet['G{}'.format(start_row)].border = thin_border
            sheet['H{}'.format(start_row)].border = thin_border
            sheet['I{}'.format(start_row)].border = thin_border

            sheet['J{}'.format(start_row)] = "=SUM(J{}:J{})".format(3, start_row - 1)
            sheet['J{}'.format(start_row)].number_format = '#,##0.00'
            sheet['J{}'.format(start_row)].border = thin_border
            sheet['J{}'.format(start_row)].font = Font(bold=True, size="8")

        sheet.protection.password = "YakinBisa123!"
        sheet.protection.sheet = True
        sheet.protection.enable()
        wb.save(filename="{}/../reports/report_perdin_{}.xlsx".format(path, self.name.replace("/", "_")))
        return {
            'name': 'Report Lembur',
            'type': 'ir.actions.act_url',
            'url': '/hr_ykp_holidays/download/{}'.format(self.id),
            'target': 'self',
        }


class perdin_uang_saku(models.Model):
    _name = 'hr.perdin.allowance'
    _rec_name = 'jenis_pegawai'

    job_id = fields.Many2one(comodel_name="hr.job", string="Jabatan")
    jenis_pegawai = fields.Many2one('jenis.pegawai', string='Jenis Pegawai', required=True)
    nominal = fields.Float('Nominal')
    nominal_luar = fields.Float('Nominal Luar Negeri')
    extra = fields.Float('(%)tambahan jika diluar pulau', default=0.0)
    extra_jawa = fields.Float('(%)tambahan jika diluar pulau jawa', default=0.0)
    location_exclude_ids = fields.Many2many('hr.state', string='Lokasi Pengecualian')


class hr_overtime_employee(models.Model):
    _name = "hr.perdin.employee"
    _description = "Detail Employee"
    _rec_name = 'perdin_id'

    SESSION_STATES = [('draft', 'Draft'), ("validate", "Approved"), ("refuse", "Refused"),
                      ("set_to_draft", "Set To Draft")]

    perdin_id = fields.Many2one("hr.perdin", "Nomor Perdin", ondelete="cascade")
    tempat_tujuan = fields.Many2one(related='perdin_id.tempat_tujuan')
    tujuan_perdin = fields.Char(related='perdin_id.tujuan_perdin')
    image = fields.Binary("Photo", related="employee_id.image", readonly=True)
    employee_id = fields.Many2one("hr.employee", "Employee", required=True, track_visibility='onchange')
    user_id = fields.Many2one(related='employee_id.user_id')
    nomor = fields.Char(string="Nomor", related="perdin_id.name")
    nip = fields.Char("NIP", related="employee_id.nik", readonly=True)
    date_from = fields.Date(string="Tanggal Berangkat", related="perdin_id.date_from")
    date_to = fields.Date(string="Tanggal Kembali", related="perdin_id.date_to")
    state = fields.Selection(string="State",
                             selection=SESSION_STATES,
                             required=True,
                             readonly=True,
                             default=SESSION_STATES[0][0], related="perdin_id.state")
    uang_saku = fields.Float('Uang Saku', default=0.0)
    perdin_cost_ids = fields.One2many('hr.perdin.employee.cost', 'perdin_id', string='Pengeluran Perdin')
    allowance_id = fields.Many2one('hr.perdin.allowance')
    uang_saku_perhari = fields.Float("Uang saku perhari")
    extra_jawa = fields.Float("extra_jawa")
    extra = fields.Float("extra")

    @api.multi
    def hitung_uang_saku(self):
        negara_tujuan = self.perdin_id.negara_tujuan
        propinsi_tujuan = self.perdin_id.propinsi_tujuan
        num_days = self.perdin_id.number_of_days_temp
        jenis_pegawai = self.employee_id.jenis_pegawai
        allowance = self.env['hr.perdin.allowance'].search([('jenis_pegawai', '=', jenis_pegawai.id)], limit=1)
        if allowance:
            self.write({'allowance_id': allowance.id})
            exclude = False
            for excluded in allowance.location_exclude_ids:
                if excluded.id == propinsi_tujuan.id:
                    exclude = True
                    break
            if not exclude and negara_tujuan.id == 100:
                if 'jawa timur' in propinsi_tujuan.name.lower() or 'jawa tengah' in propinsi_tujuan.name.lower():
                    extra = allowance.extra_jawa / 100
                    self.write({'extra_jawa': extra * allowance.nominal})
                else:
                    extra = allowance.extra / 100
                    self.write({'extra': extra * allowance.nominal})
                self.write({'uang_saku_perhari': allowance.nominal, 'uang_saku': (allowance.nominal + (allowance.nominal * extra)) * num_days})
            elif exclude and negara_tujuan.id == 100:
                self.write({'uang_saku_perhari': allowance.nominal,'uang_saku': allowance.nominal * num_days})
            elif negara_tujuan.id != 100:
                self.write({'uang_saku_perhari': allowance.nominal_luar, 'uang_saku': allowance.nominal_luar * num_days})


class perdin_proof_cost(models.Model):
    _name = 'hr.perdin.employee.cost'

    perdin_id = fields.Many2one('hr.perdin.employee', string='Perdin', )
    tempat_tujuan = fields.Many2one(related='perdin_id.tempat_tujuan')
    tujuan_perdin = fields.Char(related='perdin_id.tujuan_perdin')
    employee_id = fields.Many2one(related='perdin_id.employee_id')
    user_id = fields.Many2one(related='perdin_id.user_id')
    user_login = fields.Many2one(default=lambda self: self.env.user.id, store=False)
    nominal = fields.Float('Biaya')
    state = fields.Selection([
        ('draft', 'Draft'), ("validate", "Approved")
    ], default='draft')
    doc_proof = fields.Binary('Bukti Pengeluaran')
    filename = fields.Char('filename')

    @api.multi
    def action_approve(self):
        if self.env.user.has_group('hr.group_hr_manager'):
            self.write({'state': 'validate'})
        else:
            raise UserError(_('Anda Tidak Memiliki Hak Akses INI.'))
